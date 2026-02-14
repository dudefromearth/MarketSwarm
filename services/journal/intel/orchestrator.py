# services/journal/intel/orchestrator.py
"""API server and main loop for the FOTW Trade Log service (v2)."""

import asyncio
import csv
import io
import json
import os
from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime, timedelta

from aiohttp import web
import openpyxl
from openpyxl.utils import get_column_letter

import redis.asyncio as redis

from .db_v2 import JournalDBv2, VersionConflictError
from .models_v2 import (
    TradeLog, Trade, TradeEvent, Symbol, Setting, Tag, Order,
    JournalEntry, JournalRetrospective, JournalTradeRef, JournalAttachment,
    PlaybookEntry, PlaybookSourceRef, Alert,
    PromptAlert, PromptAlertVersion, ReferenceStateSnapshot, PromptAlertTrigger,
    TrackedIdea, SelectorParams,
    RiskGraphStrategy, RiskGraphStrategyVersion, RiskGraphTemplate,
    Position, Leg, Fill, PositionEvent,
    MLDecision, PnLEvent, DailyPerformance, MLFeatureSnapshot,
    TrackedIdeaSnapshot, UserTradeAction, MLModel, MLExperiment,
    PositionJournalEntry,
    VALID_TAG_CATEGORIES, VALID_TAG_SCOPES, DEFAULT_SCOPES_BY_CATEGORY,
    EdgeLabSetup, EdgeLabHypothesis, EdgeLabOutcome,
)
from .analytics_v2 import AnalyticsV2
from .edge_lab_analytics import EdgeLabAnalytics
from .auth import JournalAuth, require_auth, optional_auth


class JournalOrchestrator:
    """REST API server for the FOTW Trade Log service."""

    def __init__(self, config: Dict[str, Any], logger):
        self.config = config
        self.logger = logger
        self.port = int(config.get('JOURNAL_PORT', 3002))
        self.db = JournalDBv2(config)
        self.analytics = AnalyticsV2(self.db)
        self.edge_lab_analytics = EdgeLabAnalytics(self.db)
        self.auth = JournalAuth(config, self.db._pool)

        # Redis for alerts sync
        redis_host = config.get('REDIS_HOST', 'localhost')
        redis_port = int(config.get('REDIS_PORT', 6379))
        self._redis: Optional[redis.Redis] = None
        self._redis_url = f"redis://{redis_host}:{redis_port}"

        # Journal attachments storage
        default_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'attachments')
        self.attachments_path = Path(config.get('JOURNAL_ATTACHMENTS_PATH', default_path))
        self.attachments_path.mkdir(parents=True, exist_ok=True)
        self.max_attachment_size = 10 * 1024 * 1024  # 10MB

    # Allowed origins for CORS with credentials
    ALLOWED_ORIGINS = [
        'http://localhost:5173',
        'http://localhost:5174',
        'http://localhost:5175',
        'http://127.0.0.1:5173',
        'http://127.0.0.1:5174',
        'http://127.0.0.1:5175',
    ]

    def _get_cors_origin(self, request: web.Request) -> str:
        """Get allowed origin for CORS response."""
        origin = request.headers.get('Origin', '')
        if origin in self.ALLOWED_ORIGINS:
            return origin
        return self.ALLOWED_ORIGINS[0]  # Default to first allowed origin

    def _json_response(self, data: Any, status: int = 200, request: web.Request = None) -> web.Response:
        """Create a JSON response with CORS headers."""
        origin = self._get_cors_origin(request) if request else self.ALLOWED_ORIGINS[0]
        return web.Response(
            text=json.dumps(data, default=str),
            status=status,
            content_type='application/json',
            headers={
                'Access-Control-Allow-Origin': origin,
                'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type, Authorization',
                'Access-Control-Allow-Credentials': 'true',
            }
        )

    def _error_response(self, message: str, status: int = 400, request: web.Request = None) -> web.Response:
        """Create an error JSON response."""
        return self._json_response({'success': False, 'error': message}, status, request)

    async def handle_options(self, request: web.Request) -> web.Response:
        """Handle CORS preflight requests."""
        origin = self._get_cors_origin(request)
        return web.Response(
            status=204,
            headers={
                'Access-Control-Allow-Origin': origin,
                'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type, Authorization',
                'Access-Control-Allow-Credentials': 'true',
            }
        )

    # ==================== Trade Log Endpoints ====================

    async def list_logs(self, request: web.Request) -> web.Response:
        """GET /api/logs - List trade logs for authenticated user.

        Query params:
        - state: Filter by lifecycle state (active, archived, retired, or comma-separated)
        - include_retired: Include retired logs (default: false)
        - include_inactive: Legacy param, treated as include archived+retired
        """
        try:
            # Get authenticated user (optional - returns all if no auth)
            user = await self.auth.get_request_user(request)
            user_id = user['id'] if user else None

            # Parse lifecycle state filter
            state_param = request.query.get('state', '')
            include_retired = request.query.get('include_retired', 'false') == 'true'
            include_inactive = request.query.get('include_inactive', 'false') == 'true'

            if state_param:
                # Explicit state filter
                states = [s.strip() for s in state_param.split(',')]
            elif include_inactive:
                # Legacy: include_inactive means all states except retired (unless also retired)
                states = ['active', 'archived']
                if include_retired:
                    states.append('retired')
            else:
                # Default: only active logs
                states = ['active', 'archived'] if include_retired else ['active']

            # Use new lifecycle-aware method if user_id available
            if user_id:
                logs = self.db.list_logs_by_state(user_id, states, include_retired)
            else:
                logs = self.db.list_logs(user_id=user_id, include_inactive=include_inactive)

            # Get summaries for each log
            log_data = []
            for log in logs:
                summary = self.db.get_log_summary(log.id, user_id)
                # Add lifecycle info to summary
                if summary:
                    summary['lifecycle_state'] = log.lifecycle_state
                    summary['archived_at'] = log.archived_at
                    summary['retired_at'] = log.retired_at
                    summary['retire_scheduled_at'] = log.retire_scheduled_at
                    summary['is_read_only'] = log.lifecycle_state != 'active'
                    summary['ml_included'] = bool(log.ml_included)
                    log_data.append(summary)

            # Include active log count for UI
            active_count = sum(1 for log in logs if log.lifecycle_state == 'active')

            return self._json_response({
                'success': True,
                'data': log_data,
                'count': len(log_data),
                'active_count': active_count
            })
        except Exception as e:
            self.logger.error(f"list_logs error: {e}")
            return self._error_response(str(e), 500)

    async def get_log(self, request: web.Request) -> web.Response:
        """GET /api/logs/:id - Get a single trade log with summary."""
        try:
            log_id = request.match_info['id']

            # Get authenticated user (optional)
            user = await self.auth.get_request_user(request)
            user_id = user['id'] if user else None

            summary = self.db.get_log_summary(log_id, user_id)

            if not summary:
                return self._error_response('Trade log not found', 404)

            return self._json_response({
                'success': True,
                'data': summary
            })
        except Exception as e:
            self.logger.error(f"get_log error: {e}")
            return self._error_response(str(e), 500)

    async def create_log(self, request: web.Request) -> web.Response:
        """POST /api/logs - Create a new trade log."""
        try:
            # Require authentication for creating logs
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            body = await request.json()

            if 'name' not in body:
                return self._error_response('name is required')
            if 'starting_capital' not in body:
                return self._error_response('starting_capital is required')

            # Convert dollars to cents if needed
            starting_capital = body['starting_capital']
            if isinstance(starting_capital, float) or starting_capital < 10000:
                # Assume dollars, convert to cents
                starting_capital = int(starting_capital * 100)

            risk_per_trade = body.get('risk_per_trade')
            if risk_per_trade and (isinstance(risk_per_trade, float) or risk_per_trade < 1000):
                risk_per_trade = int(risk_per_trade * 100)

            log = TradeLog(
                id=TradeLog.new_id(),
                name=body['name'],
                user_id=user['id'],  # Set owner from authenticated user
                starting_capital=starting_capital,
                risk_per_trade=risk_per_trade,
                max_position_size=body.get('max_position_size'),
                intent=body.get('intent'),
                constraints=json.dumps(body.get('constraints')) if body.get('constraints') else None,
                regime_assumptions=body.get('regime_assumptions'),
                notes=body.get('notes')
            )

            created = self.db.create_log(log)
            self.logger.info(f"Created trade log '{created.name}' for user {user['id']} with ${starting_capital/100:.2f} capital", emoji="📒")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201)

        except Exception as e:
            self.logger.error(f"create_log error: {e}")
            return self._error_response(str(e), 500)

    async def update_log(self, request: web.Request) -> web.Response:
        """PUT /api/logs/:id - Update trade log metadata (not starting params)."""
        try:
            log_id = request.match_info['id']
            body = await request.json()

            log = self.db.update_log(log_id, body)

            if not log:
                return self._error_response('Trade log not found', 404)

            return self._json_response({
                'success': True,
                'data': log.to_api_dict()
            })
        except Exception as e:
            self.logger.error(f"update_log error: {e}")
            return self._error_response(str(e), 500)

    async def delete_log(self, request: web.Request) -> web.Response:
        """DELETE /api/logs/:id - Soft delete a trade log (legacy, use archive instead)."""
        try:
            log_id = request.match_info['id']
            deleted = self.db.delete_log(log_id)

            if not deleted:
                return self._error_response('Trade log not found', 404)

            self.logger.info(f"Archived trade log {log_id}", emoji="📦")

            return self._json_response({
                'success': True,
                'message': 'Trade log archived'
            })
        except Exception as e:
            self.logger.error(f"delete_log error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Trade Log Lifecycle Management ====================

    async def archive_log(self, request: web.Request) -> web.Response:
        """POST /api/logs/:id/archive - Archive a trade log.

        Preconditions:
        - Log must be active
        - Log must have no open positions
        - Log must have no pending alerts
        """
        try:
            user = await require_auth(request, self.auth)
            if isinstance(user, web.Response):
                return user

            log_id = request.match_info['id']
            result = self.db.archive_log(log_id, user['id'])

            if not result.get('success'):
                status = 400
                if 'open position' in result.get('error', '').lower():
                    status = 409  # Conflict
                return self._json_response(result, status, request)

            self.logger.info(f"Archived trade log {log_id}", emoji="📦")

            # Publish lifecycle event for SSE
            await self._publish_lifecycle_event(user['id'], 'archived', {
                'log_id': log_id,
                'lifecycle_state': 'archived',
                'archived_at': result.get('archived_at'),
                'ml_included': False
            })

            return self._json_response({
                'success': True,
                'message': 'Trade log archived',
                'archived_at': result.get('archived_at')
            }, request=request)
        except Exception as e:
            self.logger.error(f"archive_log error: {e}")
            return self._error_response(str(e), 500, request)

    async def reactivate_log(self, request: web.Request) -> web.Response:
        """POST /api/logs/:id/reactivate - Reactivate an archived log.

        Checks active log limit before reactivating.
        """
        try:
            user = await require_auth(request, self.auth)
            if isinstance(user, web.Response):
                return user

            log_id = request.match_info['id']
            result = self.db.reactivate_log(log_id, user['id'])

            if not result.get('success'):
                return self._json_response(result, 400, request)

            self.logger.info(f"Reactivated trade log {log_id}", emoji="♻️")

            # Publish lifecycle event for SSE
            active_count = result.get('active_count', 0)
            cap_state = 'hard_warning' if active_count >= 10 else ('soft_warning' if active_count >= 5 else 'ok')
            await self._publish_lifecycle_event(user['id'], 'reactivated', {
                'log_id': log_id,
                'lifecycle_state': 'active',
                'reactivated_at': result.get('reactivated_at'),
                'active_log_count': active_count,
                'cap_state': cap_state
            })

            response = {
                'success': True,
                'message': 'Trade log reactivated',
                'reactivated_at': result.get('reactivated_at')
            }
            if result.get('warning'):
                response['warning'] = result['warning']

            return self._json_response(response, request=request)
        except Exception as e:
            self.logger.error(f"reactivate_log error: {e}")
            return self._error_response(str(e), 500, request)

    async def schedule_retire_log(self, request: web.Request) -> web.Response:
        """POST /api/logs/:id/retire - Schedule a log for permanent retirement.

        Optionally accepts:
        - grace_days: Number of days before retirement (default: 7)
        - force: If true, retire immediately (requires confirmation)
        - confirm_name: Log name for immediate retirement confirmation
        """
        try:
            user = await require_auth(request, self.auth)
            if isinstance(user, web.Response):
                return user

            log_id = request.match_info['id']
            body = await request.json() if request.body_exists else {}

            force = body.get('force', False)
            grace_days = int(body.get('grace_days', 7))

            if force:
                # Immediate retirement requires name confirmation
                confirm_name = body.get('confirm_name', '')
                log = self.db.get_log(log_id, user['id'])
                if not log:
                    return self._error_response('Log not found', 404, request)

                if confirm_name != log.name:
                    return self._error_response(
                        'To retire immediately, type the log name to confirm',
                        400, request
                    )

                result = self.db.retire_log(log_id, user['id'], force=True)
            else:
                result = self.db.schedule_retirement(log_id, user['id'], grace_days)

            if not result.get('success'):
                return self._json_response(result, 400, request)

            if force:
                self.logger.info(f"Permanently retired trade log {log_id}", emoji="🗄️")
                await self._publish_lifecycle_event(user['id'], 'retired', {
                    'log_id': log_id,
                    'retired_at': result.get('retired_at')
                })
            else:
                self.logger.info(f"Scheduled retirement for trade log {log_id} in {grace_days} days", emoji="⏳")
                await self._publish_lifecycle_event(user['id'], 'retire_scheduled', {
                    'log_id': log_id,
                    'retire_scheduled_at': result.get('retire_at'),
                    'grace_days_remaining': grace_days
                })

            return self._json_response({
                'success': True,
                **result
            }, request=request)
        except Exception as e:
            self.logger.error(f"schedule_retire_log error: {e}")
            return self._error_response(str(e), 500, request)

    async def cancel_retire_log(self, request: web.Request) -> web.Response:
        """DELETE /api/logs/:id/retire - Cancel a scheduled retirement."""
        try:
            user = await require_auth(request, self.auth)
            if isinstance(user, web.Response):
                return user

            log_id = request.match_info['id']
            result = self.db.cancel_retirement(log_id, user['id'])

            if not result.get('success'):
                return self._json_response(result, 400, request)

            self.logger.info(f"Cancelled retirement for trade log {log_id}", emoji="✅")
            await self._publish_lifecycle_event(user['id'], 'retire_cancelled', {
                'log_id': log_id,
                'lifecycle_state': 'archived'
            })

            return self._json_response({
                'success': True,
                'message': 'Retirement cancelled'
            }, request=request)
        except Exception as e:
            self.logger.error(f"cancel_retire_log error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_logs_for_import(self, request: web.Request) -> web.Response:
        """GET /api/logs/for-import - Get logs suitable for import targeting.

        Returns logs with metadata for import selection UI:
        - Active logs first, then archived
        - Includes open position count, ML status
        - Excludes retired logs
        """
        try:
            user = await require_auth(request, self.auth)
            if isinstance(user, web.Response):
                return user

            logs = self.db.get_logs_for_import_selection(user['id'])

            return self._json_response({
                'success': True,
                'data': logs
            }, request=request)
        except Exception as e:
            self.logger.error(f"get_logs_for_import error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_import_recommendation(self, request: web.Request) -> web.Response:
        """POST /api/logs/import-recommendation - Get recommendation for import targeting.

        Body:
        - target_log_id: Selected log ID
        - earliest_date: Earliest date in import (ISO string)
        - latest_date: Latest date in import (ISO string)

        Returns recommendation on whether to create an archived log instead.
        """
        try:
            user = await require_auth(request, self.auth)
            if isinstance(user, web.Response):
                return user

            body = await request.json()
            target_log_id = body.get('target_log_id')
            earliest_date = body.get('earliest_date')
            latest_date = body.get('latest_date')

            if not target_log_id:
                return self._error_response('target_log_id is required', 400, request)

            result = self.db.get_import_recommendation(
                user['id'],
                target_log_id,
                (earliest_date, latest_date)
            )

            return self._json_response({
                'success': True,
                **result
            }, request=request)
        except Exception as e:
            self.logger.error(f"get_import_recommendation error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_users_with_logs(self, request: web.Request) -> web.Response:
        """GET /api/users/with-logs - Get all user IDs that have trade logs.

        Used by Vexy's scheduled jobs to know which users to analyze.
        Internal endpoint, restricted to localhost.
        """
        try:
            # Restrict to internal access only
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            user_ids = self.db.get_users_with_logs()

            return self._json_response({
                'success': True,
                'user_ids': user_ids,
                'count': len(user_ids)
            }, request=request)
        except Exception as e:
            self.logger.error(f"get_users_with_logs error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_logs_health_metrics(self, request: web.Request) -> web.Response:
        """GET /api/logs/health - Get log health metrics for a user.

        Query params:
        - user_id: Required. User ID to get metrics for.

        Returns health metrics for each log including:
        - Activity timestamps (last trade, last import)
        - Trade counts (total, open)
        - Alert counts (pending)
        - ML inclusion status
        - Lifecycle state

        Used by Vexy's log_health_analyzer scheduled job.
        Internal endpoint, restricted to localhost.
        """
        try:
            # Restrict to internal access only
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            user_id = request.query.get('user_id')
            if not user_id:
                return self._error_response('user_id is required', 400, request)

            try:
                user_id = int(user_id)
            except ValueError:
                return self._error_response('user_id must be an integer', 400, request)

            logs = self.db.get_logs_health_metrics(user_id)

            return self._json_response({
                'success': True,
                'data': logs,
                'count': len(logs)
            }, request=request)
        except Exception as e:
            self.logger.error(f"get_logs_health_metrics error: {e}")
            return self._error_response(str(e), 500, request)

    # ==================== Trade Endpoints (Log-Scoped) ====================

    async def list_trades(self, request: web.Request) -> web.Response:
        """GET /api/logs/:logId/trades - List trades in a log."""
        try:
            log_id = request.match_info['logId']
            params = request.query

            trades = self.db.list_trades(
                log_id=log_id,
                status=params.get('status'),
                symbol=params.get('symbol'),
                strategy=params.get('strategy'),
                from_date=params.get('from'),
                to_date=params.get('to'),
                limit=int(params.get('limit', 100)),
                offset=int(params.get('offset', 0))
            )

            return self._json_response({
                'success': True,
                'data': [t.to_api_dict() for t in trades],
                'count': len(trades)
            })
        except Exception as e:
            self.logger.error(f"list_trades error: {e}")
            return self._error_response(str(e), 500)

    async def get_trade(self, request: web.Request) -> web.Response:
        """GET /api/trades/:id - Get a single trade with events."""
        try:
            trade_id = request.match_info['id']
            trade_data = self.db.get_trade_with_events(trade_id)

            if not trade_data:
                return self._error_response('Trade not found', 404)

            return self._json_response({
                'success': True,
                'data': trade_data
            })
        except Exception as e:
            self.logger.error(f"get_trade error: {e}")
            return self._error_response(str(e), 500)

    async def create_trade(self, request: web.Request) -> web.Response:
        """POST /api/logs/:logId/trades - Create a new trade."""
        try:
            log_id = request.match_info['logId']
            body = await request.json()

            # Verify log exists
            log = self.db.get_log(log_id)
            if not log:
                return self._error_response('Trade log not found', 404)

            # Entry mode: instant (default), freeform, or simulated
            entry_mode = body.get('entry_mode', 'instant')
            if entry_mode not in ('instant', 'freeform', 'simulated'):
                return self._error_response('entry_mode must be "instant", "freeform", or "simulated"', 400)

            # Set entry_time based on mode
            if entry_mode == 'freeform':
                # Freeform requires user-specified entry_time
                entry_time = body.get('entry_time')
                if not entry_time:
                    return self._error_response('entry_time is required for freeform trades', 400)
            else:
                entry_time = body.get('entry_time') or datetime.utcnow().isoformat()

            # Convert prices to cents if needed
            entry_price = body['entry_price']
            if isinstance(entry_price, float) or entry_price < 1000:
                entry_price = int(entry_price * 100)

            planned_risk = body.get('planned_risk')
            if planned_risk and (isinstance(planned_risk, float) or planned_risk < 1000):
                planned_risk = int(planned_risk * 100)

            max_profit = body.get('max_profit')
            if max_profit and (isinstance(max_profit, float) or max_profit < 1000):
                max_profit = int(max_profit * 100)

            max_loss = body.get('max_loss')
            if max_loss and (isinstance(max_loss, float) or max_loss < 1000):
                max_loss = int(max_loss * 100)

            # Determine initial status
            # Freeform can create already-closed trades if exit_time provided
            status = 'open'
            exit_time = body.get('exit_time')
            exit_price = body.get('exit_price')
            exit_spot = body.get('exit_spot')

            if entry_mode == 'freeform' and exit_time and exit_price is not None:
                status = 'closed'
                if isinstance(exit_price, float) or exit_price < 1000:
                    exit_price = int(exit_price * 100)

            trade = Trade(
                id=Trade.new_id(),
                log_id=log_id,
                symbol=body.get('symbol', 'SPX'),
                underlying=body.get('underlying', 'I:SPX'),
                strategy=body['strategy'],
                side=body['side'],
                strike=float(body['strike']),
                width=body.get('width'),
                dte=body.get('dte'),
                quantity=int(body.get('quantity', 1)),
                entry_time=entry_time,
                entry_price=entry_price,
                entry_spot=body.get('entry_spot'),
                entry_iv=body.get('entry_iv'),
                exit_time=exit_time if status == 'closed' else None,
                exit_price=exit_price if status == 'closed' else None,
                exit_spot=exit_spot if status == 'closed' else None,
                planned_risk=planned_risk or max_loss,  # Default to max_loss
                max_profit=max_profit,
                max_loss=max_loss,
                status=status,
                entry_mode=entry_mode,
                notes=body.get('notes'),
                tags=body.get('tags', []),
                source=body.get('source', 'manual'),
                playbook_id=body.get('playbook_id')
            )

            # Calculate P&L if closed
            if status == 'closed':
                trade.calculate_pnl()

            created = self.db.create_trade(trade)
            self.logger.info(f"Created {entry_mode} trade: {created.strategy} {created.strike} @ ${entry_price/100:.2f}", emoji="📝")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201)

        except KeyError as e:
            return self._error_response(f'Missing required field: {e}')
        except Exception as e:
            self.logger.error(f"create_trade error: {e}")
            return self._error_response(str(e), 500)

    async def update_trade(self, request: web.Request) -> web.Response:
        """PUT /api/trades/:id - Update a trade.

        Edit rules by entry_mode:
        - freeform: Full edit freedom (design sandbox)
        - instant: Protected fields only
        - simulated (before immutable_at): Full edit freedom (planning stage)
        - simulated (after immutable_at): Core immutable, notes/tags editable
        """
        try:
            trade_id = request.match_info['id']
            body = await request.json()

            # Get existing trade to check entry mode and immutability
            existing_trade = self.db.get_trade(trade_id)
            if not existing_trade:
                return self._error_response('Trade not found', 404, request)

            # Always protected fields
            protected = ['id', 'log_id', 'created_at']

            # Immutable core fields for locked simulated trades
            # These define the position at risk and cannot be changed after the fact
            sim_immutable_fields = [
                'entry_time', 'entry_price', 'entry_spot', 'entry_iv',
                'symbol', 'underlying', 'strategy', 'side', 'strike', 'width',
                'dte', 'quantity', 'entry_mode', 'immutable_at'
            ]

            # Check if simulated trade is locked (has immutable_at timestamp)
            is_locked_sim = (
                existing_trade.entry_mode == 'simulated' and
                existing_trade.immutable_at is not None
            )

            if is_locked_sim:
                # Check if trying to edit immutable fields
                attempted_core_edits = [k for k in body.keys() if k in sim_immutable_fields]
                if attempted_core_edits:
                    # Check if this is a correction request (has correction_reason)
                    correction_reason = body.get('correction_reason')
                    if correction_reason:
                        # Allow correction but log it
                        user = await self.auth.get_request_user(request)
                        user_id = user['id'] if user else None
                        for field in attempted_core_edits:
                            if field in body and field != 'correction_reason':
                                self.db.record_trade_correction(
                                    trade_id=trade_id,
                                    field_name=field,
                                    original_value=str(getattr(existing_trade, field, '')),
                                    corrected_value=str(body[field]),
                                    correction_reason=correction_reason,
                                    user_id=user_id
                                )
                    else:
                        return self._error_response(
                            f"Cannot modify core fields ({', '.join(attempted_core_edits)}) on locked simulated trade. "
                            "Provide 'correction_reason' for auditable corrections, or edit notes/tags instead.",
                            403, request
                        )

            # Filter updates based on protection rules
            updates = {k: v for k, v in body.items() if k not in protected and k != 'correction_reason'}

            # Handle tags serialization
            if 'tags' in updates:
                updates['tags'] = json.dumps(updates['tags'])

            trade = self.db.update_trade(trade_id, updates)

            if not trade:
                return self._error_response('Trade not found', 404, request)

            return self._json_response({
                'success': True,
                'data': trade.to_api_dict()
            }, request=request)
        except Exception as e:
            self.logger.error(f"update_trade error: {e}")
            return self._error_response(str(e), 500, request)

    async def add_adjustment(self, request: web.Request) -> web.Response:
        """POST /api/trades/:id/adjust - Add an adjustment event."""
        try:
            trade_id = request.match_info['id']
            body = await request.json()

            if 'price' not in body:
                return self._error_response('price is required')
            if 'quantity_change' not in body:
                return self._error_response('quantity_change is required')

            # Convert price to cents
            price = body['price']
            if isinstance(price, float) or price < 1000:
                price = int(price * 100)

            event = self.db.add_adjustment(
                trade_id=trade_id,
                price=price,
                quantity_change=int(body['quantity_change']),
                spot=body.get('spot'),
                notes=body.get('notes'),
                event_time=body.get('event_time')
            )

            if not event:
                return self._error_response('Trade not found or already closed', 404)

            self.logger.info(f"Adjusted trade {trade_id}: {body['quantity_change']:+d} @ ${price/100:.2f}", emoji="📝")

            return self._json_response({
                'success': True,
                'data': event.to_api_dict()
            })
        except Exception as e:
            self.logger.error(f"add_adjustment error: {e}")
            return self._error_response(str(e), 500)

    async def close_trade(self, request: web.Request) -> web.Response:
        """POST /api/trades/:id/close - Close a trade."""
        try:
            trade_id = request.match_info['id']
            body = await request.json()

            if 'exit_price' not in body:
                return self._error_response('exit_price is required')

            # Convert price to cents
            exit_price = body['exit_price']
            if isinstance(exit_price, float) or exit_price < 1000:
                exit_price = int(exit_price * 100)

            trade = self.db.close_trade(
                trade_id=trade_id,
                exit_price=exit_price,
                exit_spot=body.get('exit_spot'),
                exit_time=body.get('exit_time'),
                notes=body.get('notes')
            )

            if not trade:
                return self._error_response('Trade not found or already closed', 404)

            pnl_dollars = (trade.pnl or 0) / 100
            pnl_str = f"+${pnl_dollars:.2f}" if pnl_dollars >= 0 else f"-${abs(pnl_dollars):.2f}"
            self.logger.info(f"Closed trade {trade_id}: {pnl_str}", emoji="✅")

            return self._json_response({
                'success': True,
                'data': trade.to_api_dict()
            })
        except Exception as e:
            self.logger.error(f"close_trade error: {e}")
            return self._error_response(str(e), 500)

    async def delete_trade(self, request: web.Request) -> web.Response:
        """DELETE /api/trades/:id - Delete a trade."""
        try:
            trade_id = request.match_info['id']
            deleted = self.db.delete_trade(trade_id)

            if not deleted:
                return self._error_response('Trade not found', 404)

            self.logger.info(f"Deleted trade {trade_id}", emoji="🗑️")

            return self._json_response({
                'success': True,
                'message': 'Trade deleted'
            })
        except Exception as e:
            self.logger.error(f"delete_trade error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Order Queue Endpoints ====================

    async def list_orders(self, request: web.Request) -> web.Response:
        """GET /api/orders - List orders for authenticated user."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            # Optional filters
            status = request.query.get('status')
            order_type = request.query.get('type')
            symbol = request.query.get('symbol')

            orders = self.db.list_orders(
                user_id=user['id'],
                status=status,
                order_type=order_type,
                symbol=symbol
            )

            return self._json_response({
                'success': True,
                'data': [o.to_api_dict() for o in orders],
                'count': len(orders)
            }, request=request)
        except Exception as e:
            self.logger.error(f"list_orders error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_orders_active(self, request: web.Request) -> web.Response:
        """GET /api/orders/active - Get active orders summary."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            # Get pending orders
            pending_orders = self.db.list_pending_orders(user['id'])

            # Split by type
            pending_entries = [o.to_api_dict() for o in pending_orders if o.order_type == 'entry']
            pending_exits = [o.to_api_dict() for o in pending_orders if o.order_type == 'exit']

            return self._json_response({
                'success': True,
                'data': {
                    'pending_entries': pending_entries,
                    'pending_exits': pending_exits,
                    'total': len(pending_orders)
                }
            }, request=request)
        except Exception as e:
            self.logger.error(f"get_orders_active error: {e}")
            return self._error_response(str(e), 500, request)

    async def create_order(self, request: web.Request) -> web.Response:
        """POST /api/orders - Create a new order."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            body = await request.json()

            # Validate required fields
            order_type = body.get('order_type')
            if order_type not in ('entry', 'exit'):
                return self._error_response('order_type must be "entry" or "exit"', 400, request)

            symbol = body.get('symbol')
            direction = body.get('direction')
            limit_price = body.get('limit_price')

            if not symbol or not direction or limit_price is None:
                return self._error_response('symbol, direction, and limit_price are required', 400, request)

            if direction not in ('long', 'short'):
                return self._error_response('direction must be "long" or "short"', 400, request)

            # For exit orders, trade_id is required
            trade_id = body.get('trade_id')
            if order_type == 'exit' and not trade_id:
                return self._error_response('trade_id is required for exit orders', 400, request)

            order = self.db.create_order(
                user_id=user['id'],
                order_type=order_type,
                symbol=symbol,
                direction=direction,
                limit_price=float(limit_price),
                quantity=int(body.get('quantity', 1)),
                trade_id=trade_id,
                strategy=body.get('strategy'),
                stop_loss=float(body['stop_loss']) if body.get('stop_loss') else None,
                take_profit=float(body['take_profit']) if body.get('take_profit') else None,
                notes=body.get('notes'),
                expires_at=body.get('expires_at')
            )

            if not order:
                return self._error_response('Failed to create order', 500, request)

            self.logger.info(f"Created {order_type} order for {symbol} @ ${limit_price}", emoji="📋")

            return self._json_response({
                'success': True,
                'data': order.to_api_dict()
            }, 201, request)

        except Exception as e:
            self.logger.error(f"create_order error: {e}")
            return self._error_response(str(e), 500, request)

    async def cancel_order(self, request: web.Request) -> web.Response:
        """DELETE /api/orders/:id - Cancel an order."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            order_id = int(request.match_info['id'])
            cancelled = self.db.cancel_order(order_id, user['id'])

            if not cancelled:
                return self._error_response('Order not found or already processed', 404, request)

            self.logger.info(f"Cancelled order {order_id}", emoji="❌")

            return self._json_response({
                'success': True,
                'message': 'Order cancelled'
            }, request=request)
        except Exception as e:
            self.logger.error(f"cancel_order error: {e}")
            return self._error_response(str(e), 500, request)

    # ==================== Analytics Endpoints (Log-Scoped) ====================

    async def get_log_analytics(self, request: web.Request) -> web.Response:
        """GET /api/logs/:logId/analytics - Get full analytics for a log."""
        try:
            log_id = request.match_info['logId']
            analytics = self.analytics.get_full_analytics(log_id)

            if not analytics:
                return self._error_response('Trade log not found', 404)

            return self._json_response({
                'success': True,
                'data': analytics.to_api_dict()
            })
        except Exception as e:
            self.logger.error(f"get_log_analytics error: {e}")
            return self._error_response(str(e), 500)

    async def get_log_equity(self, request: web.Request) -> web.Response:
        """GET /api/logs/:logId/equity - Get equity curve for a log."""
        try:
            log_id = request.match_info['logId']
            params = request.query

            equity_points = self.analytics.get_equity_curve(
                log_id,
                from_date=params.get('from'),
                to_date=params.get('to')
            )

            return self._json_response({
                'success': True,
                'data': {
                    'equity': [p.to_dict() for p in equity_points]
                }
            })
        except Exception as e:
            self.logger.error(f"get_log_equity error: {e}")
            return self._error_response(str(e), 500)

    async def get_log_drawdown(self, request: web.Request) -> web.Response:
        """GET /api/logs/:logId/drawdown - Get drawdown curve for a log."""
        try:
            log_id = request.match_info['logId']
            params = request.query

            drawdown_points = self.analytics.get_drawdown_curve(
                log_id,
                from_date=params.get('from'),
                to_date=params.get('to')
            )

            return self._json_response({
                'success': True,
                'data': {
                    'drawdown': [p.to_dict() for p in drawdown_points]
                }
            })
        except Exception as e:
            self.logger.error(f"get_log_drawdown error: {e}")
            return self._error_response(str(e), 500)

    async def get_return_distribution(self, request: web.Request) -> web.Response:
        """GET /api/logs/:logId/distribution - Get return distribution histogram for a log."""
        try:
            log_id = request.match_info['logId']
            params = request.query

            # Bin size in dollars (default $50)
            bin_size_dollars = int(params.get('bin_size', 50))
            bin_size_cents = bin_size_dollars * 100

            distribution = self.analytics.get_return_distribution(log_id, bin_size_cents)

            return self._json_response({
                'success': True,
                'data': {
                    'distribution': distribution,
                    'bin_size_dollars': bin_size_dollars
                }
            })
        except Exception as e:
            self.logger.error(f"get_return_distribution error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Export/Import ====================

    # CSV/Excel column definitions
    EXPORT_COLUMNS = [
        'entry_time', 'symbol', 'underlying', 'strategy', 'side', 'strike', 'width',
        'dte', 'quantity', 'entry_price', 'entry_spot', 'exit_time', 'exit_price',
        'exit_spot', 'pnl', 'r_multiple', 'status', 'planned_risk', 'max_profit',
        'max_loss', 'notes', 'tags', 'source'
    ]

    async def export_trades(self, request: web.Request) -> web.Response:
        """GET /api/logs/:logId/export - Export trades as CSV or Excel."""
        try:
            log_id = request.match_info['logId']
            format_type = request.query.get('format', 'csv').lower()

            # Verify log exists
            log = self.db.get_log(log_id)
            if not log:
                return self._error_response('Trade log not found', 404)

            # Get all trades
            trades = self.db.list_trades(log_id=log_id, limit=100000)

            if format_type == 'xlsx':
                return await self._export_excel(trades, log.name)
            else:
                return await self._export_csv(trades, log.name)

        except Exception as e:
            self.logger.error(f"export_trades error: {e}")
            return self._error_response(str(e), 500)

    async def _export_csv(self, trades: List[Trade], log_name: str) -> web.Response:
        """Generate CSV file from trades."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow(self.EXPORT_COLUMNS)

        # Data rows
        for trade in trades:
            row = self._trade_to_export_row(trade)
            writer.writerow(row)

        csv_content = output.getvalue()
        filename = f"{log_name.replace(' ', '_')}_trades.csv"

        return web.Response(
            body=csv_content,
            content_type='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Access-Control-Allow-Origin': '*',
            }
        )

    async def _export_excel(self, trades: List[Trade], log_name: str) -> web.Response:
        """Generate Excel file from trades."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trades"

        # Header row with styling
        for col, header in enumerate(self.EXPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        for row_num, trade in enumerate(trades, 2):
            row_data = self._trade_to_export_row(trade)
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)

        # Auto-size columns
        for col in range(1, len(self.EXPORT_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{log_name.replace(' ', '_')}_trades.xlsx"

        return web.Response(
            body=output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Access-Control-Allow-Origin': '*',
            }
        )

    def _trade_to_export_row(self, trade: Trade) -> list:
        """Convert a trade to an export row with dollars instead of cents."""
        return [
            trade.entry_time,
            trade.symbol,
            trade.underlying,
            trade.strategy,
            trade.side,
            trade.strike,
            trade.width,
            trade.dte,
            trade.quantity,
            trade.entry_price / 100 if trade.entry_price else None,  # cents to dollars
            trade.entry_spot,
            trade.exit_time,
            trade.exit_price / 100 if trade.exit_price else None,  # cents to dollars
            trade.exit_spot,
            trade.pnl / 100 if trade.pnl is not None else None,  # cents to dollars
            trade.r_multiple,
            trade.status,
            trade.planned_risk / 100 if trade.planned_risk else None,
            trade.max_profit / 100 if trade.max_profit else None,
            trade.max_loss / 100 if trade.max_loss else None,
            trade.notes,
            ','.join(trade.tags) if trade.tags else '',
            trade.source
        ]

    async def import_trades(self, request: web.Request) -> web.Response:
        """POST /api/logs/:logId/import - Import trades from CSV or Excel."""
        try:
            log_id = request.match_info['logId']

            # Verify log exists
            log = self.db.get_log(log_id)
            if not log:
                return self._error_response('Trade log not found', 404)

            # Read multipart data
            reader = await request.multipart()
            field = await reader.next()

            if not field:
                return self._error_response('No file uploaded', 400)

            filename = field.filename or 'upload'
            content = await field.read()

            # Determine format from filename or content
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                trades_data = self._parse_excel(content)
            else:
                trades_data = self._parse_csv(content)

            if not trades_data:
                return self._error_response('No valid trades found in file', 400)

            # Create trades
            created_count = 0
            errors = []

            for i, row in enumerate(trades_data):
                try:
                    trade = self._row_to_trade(row, log_id)
                    if trade:
                        self.db.create_trade(trade)
                        created_count += 1
                except Exception as e:
                    errors.append(f"Row {i + 2}: {str(e)}")

            self.logger.info(f"Imported {created_count} trades into '{log.name}'", emoji="📥")

            return self._json_response({
                'success': True,
                'imported': created_count,
                'errors': errors[:10] if errors else [],  # Limit error messages
                'total_errors': len(errors)
            })

        except Exception as e:
            self.logger.error(f"import_trades error: {e}")
            return self._error_response(str(e), 500)

    def _parse_csv(self, content: bytes) -> List[dict]:
        """Parse CSV content into list of dicts."""
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            text = content.decode('latin-1')

        reader = csv.DictReader(io.StringIO(text))
        return list(reader)

    def _parse_excel(self, content: bytes) -> List[dict]:
        """Parse Excel content into list of dicts."""
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # First row is headers
        headers = [str(h).lower().strip() if h else '' for h in rows[0]]
        result = []

        for row in rows[1:]:
            if not any(row):  # Skip empty rows
                continue
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value
            result.append(row_dict)

        return result

    def _row_to_trade(self, row: dict, log_id: str) -> Optional[Trade]:
        """Convert an import row to a Trade object."""
        # Normalize keys to lowercase
        row = {k.lower().strip(): v for k, v in row.items() if k}

        # Required fields
        if not row.get('strategy') or not row.get('side') or not row.get('strike'):
            return None

        # Parse entry_time
        entry_time = row.get('entry_time')
        if isinstance(entry_time, datetime):
            entry_time = entry_time.isoformat()
        elif not entry_time:
            entry_time = datetime.utcnow().isoformat()

        # Parse exit_time
        exit_time = row.get('exit_time')
        if isinstance(exit_time, datetime):
            exit_time = exit_time.isoformat()
        elif exit_time == '' or exit_time is None:
            exit_time = None

        # Convert prices from dollars to cents
        def to_cents(val):
            if val is None or val == '':
                return None
            try:
                return int(float(val) * 100)
            except (ValueError, TypeError):
                return None

        entry_price = to_cents(row.get('entry_price'))
        exit_price = to_cents(row.get('exit_price'))
        pnl = to_cents(row.get('pnl'))
        planned_risk = to_cents(row.get('planned_risk'))
        max_profit = to_cents(row.get('max_profit'))
        max_loss = to_cents(row.get('max_loss'))

        # Parse numeric fields
        def to_float(val):
            if val is None or val == '':
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        def to_int(val):
            if val is None or val == '':
                return None
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return None

        # Parse tags
        tags_raw = row.get('tags', '')
        if isinstance(tags_raw, str):
            tags = [t.strip() for t in tags_raw.split(',') if t.strip()]
        else:
            tags = []

        # Determine status
        status = str(row.get('status', '')).lower()
        if status not in ('open', 'closed'):
            status = 'closed' if exit_price is not None else 'open'

        return Trade(
            id=Trade.new_id(),
            log_id=log_id,
            symbol=str(row.get('symbol', 'SPX')).upper(),
            underlying=str(row.get('underlying', 'I:SPX')),
            strategy=str(row.get('strategy', '')).lower(),
            side=str(row.get('side', '')).lower(),
            strike=to_float(row.get('strike')) or 0,
            width=to_int(row.get('width')),
            dte=to_int(row.get('dte')),
            quantity=to_int(row.get('quantity')) or 1,
            entry_time=entry_time,
            entry_price=entry_price or 0,
            entry_spot=to_float(row.get('entry_spot')),
            exit_time=exit_time,
            exit_price=exit_price,
            exit_spot=to_float(row.get('exit_spot')),
            pnl=pnl,
            r_multiple=to_float(row.get('r_multiple')),
            planned_risk=planned_risk,
            max_profit=max_profit,
            max_loss=max_loss,
            status=status,
            notes=str(row.get('notes', '')) if row.get('notes') else None,
            tags=tags,
            source=str(row.get('source', 'import'))
        )

    # ==================== Symbols ====================

    async def list_symbols(self, request: web.Request) -> web.Response:
        """GET /api/symbols - List all symbols."""
        try:
            params = request.query
            include_disabled = params.get('include_disabled', '').lower() == 'true'
            asset_type = params.get('asset_type')

            symbols = self.db.list_symbols(include_disabled=include_disabled)

            if asset_type:
                symbols = [s for s in symbols if s.asset_type == asset_type]

            return self._json_response({
                'success': True,
                'data': [s.to_dict() for s in symbols],
                'count': len(symbols)
            })
        except Exception as e:
            self.logger.error(f"list_symbols error: {e}")
            return self._error_response(str(e), 500)

    async def get_symbol(self, request: web.Request) -> web.Response:
        """GET /api/symbols/:symbol - Get a single symbol."""
        try:
            symbol = request.match_info['symbol']
            sym = self.db.get_symbol(symbol)

            if not sym:
                return self._error_response(f"Symbol {symbol} not found", 404)

            return self._json_response({
                'success': True,
                'data': sym.to_dict()
            })
        except Exception as e:
            self.logger.error(f"get_symbol error: {e}")
            return self._error_response(str(e), 500)

    async def add_symbol(self, request: web.Request) -> web.Response:
        """POST /api/symbols - Add a new symbol."""
        try:
            body = await request.json()

            required = ['symbol', 'name', 'asset_type', 'multiplier']
            for field in required:
                if field not in body:
                    return self._error_response(f"Missing required field: {field}", 400)

            # Check if symbol already exists
            existing = self.db.get_symbol(body['symbol'])
            if existing:
                return self._error_response(f"Symbol {body['symbol']} already exists", 409)

            symbol = Symbol(
                symbol=body['symbol'].upper(),
                name=body['name'],
                asset_type=body['asset_type'],
                multiplier=int(body['multiplier']),
                enabled=body.get('enabled', True)
            )

            created = self.db.add_symbol(symbol)

            return self._json_response({
                'success': True,
                'data': created.to_dict()
            }, 201)
        except Exception as e:
            self.logger.error(f"add_symbol error: {e}")
            return self._error_response(str(e), 500)

    async def update_symbol(self, request: web.Request) -> web.Response:
        """PUT /api/symbols/:symbol - Update a symbol."""
        try:
            symbol = request.match_info['symbol']
            body = await request.json()

            updated = self.db.update_symbol(symbol, body)
            if not updated:
                return self._error_response(f"Symbol {symbol} not found", 404)

            return self._json_response({
                'success': True,
                'data': updated.to_dict()
            })
        except Exception as e:
            self.logger.error(f"update_symbol error: {e}")
            return self._error_response(str(e), 500)

    async def delete_symbol(self, request: web.Request) -> web.Response:
        """DELETE /api/symbols/:symbol - Delete a user-added symbol."""
        try:
            symbol = request.match_info['symbol']

            # Check if it's a default symbol
            sym = self.db.get_symbol(symbol)
            if sym and sym.is_default:
                return self._error_response("Cannot delete default symbols", 403)

            deleted = self.db.delete_symbol(symbol)
            if not deleted:
                return self._error_response(f"Symbol {symbol} not found or is a default", 404)

            return self._json_response({
                'success': True,
                'message': f"Symbol {symbol} deleted"
            })
        except Exception as e:
            self.logger.error(f"delete_symbol error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Tags (Vocabulary System) ====================

    async def list_tags(self, request: web.Request) -> web.Response:
        """GET /api/tags - List all tags for the current user."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response("Authentication required", 401)

            params = request.query
            include_retired = params.get('include_retired', '').lower() == 'true'
            category = params.get('category', None)
            scope = params.get('scope', None)

            # Backward compat: day-texture → state
            if category == 'day-texture':
                category = 'state'

            # Check if user has any tags, seed if not
            tag_count = self.db.count_tags(user['id'])
            if tag_count == 0:
                self.db.seed_example_tags(user['id'])

            # Always ensure state (readiness) tags exist
            self.db.seed_day_texture_tags(user['id'])

            tags = self.db.list_tags(user['id'], include_retired=include_retired,
                                     category=category, scope=scope)

            return self._json_response({
                'success': True,
                'data': [t.to_api_dict() for t in tags],
                'count': len(tags)
            })
        except Exception as e:
            self.logger.error(f"list_tags error: {e}")
            return self._error_response(str(e), 500)

    async def get_tag(self, request: web.Request) -> web.Response:
        """GET /api/tags/:id - Get a single tag."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response("Authentication required", 401)

            tag_id = request.match_info['id']
            tag = self.db.get_tag(tag_id)

            if not tag:
                return self._error_response("Tag not found", 404)

            # Verify ownership
            if tag.user_id != user['id']:
                return self._error_response("Tag not found", 404)

            return self._json_response({
                'success': True,
                'data': tag.to_api_dict()
            })
        except Exception as e:
            self.logger.error(f"get_tag error: {e}")
            return self._error_response(str(e), 500)

    async def create_tag(self, request: web.Request) -> web.Response:
        """POST /api/tags - Create a new tag."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response("Authentication required", 401)

            body = await request.json()

            if 'name' not in body or not body['name'].strip():
                return self._error_response("Tag name is required", 400)

            name = body['name'].strip()

            # Validate category
            category = body.get('category', 'custom')
            if category not in VALID_TAG_CATEGORIES:
                return self._error_response(f"Invalid category: {category}", 400)

            # Validate visibility_scopes
            scopes = body.get('visibility_scopes', DEFAULT_SCOPES_BY_CATEGORY.get(category, ['journal']))
            if not isinstance(scopes, list) or len(scopes) == 0:
                return self._error_response("visibility_scopes must be a non-empty list", 400)
            if any(s not in VALID_TAG_SCOPES for s in scopes):
                return self._error_response(f"Invalid scope values. Valid: {sorted(VALID_TAG_SCOPES)}", 400)
            if 'global' in scopes and len(scopes) > 1:
                return self._error_response("'global' scope must be exclusive", 400)

            # Check for duplicate name
            existing = self.db.get_tag_by_name(user['id'], name)
            if existing:
                return self._error_response(f"Tag '{name}' already exists", 409)

            tag = Tag(
                id=Tag.new_id(),
                user_id=user['id'],
                name=name,
                description=body.get('description', '').strip() or None,
                category=category,
                visibility_scopes=scopes,
            )

            created = self.db.create_tag(tag)

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201)
        except Exception as e:
            self.logger.error(f"create_tag error: {e}")
            return self._error_response(str(e), 500)

    async def update_tag(self, request: web.Request) -> web.Response:
        """PUT /api/tags/:id - Update a tag."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response("Authentication required", 401)

            tag_id = request.match_info['id']
            tag = self.db.get_tag(tag_id)

            if not tag:
                return self._error_response("Tag not found", 404)

            # Verify ownership
            if tag.user_id != user['id']:
                return self._error_response("Tag not found", 404)

            body = await request.json()

            # Locked tags: only description can be updated
            if tag.is_locked:
                restricted = {'name', 'category', 'visibility_scopes'} & set(body.keys())
                if restricted:
                    return self._error_response("Locked tags can only have their description updated", 403)

            # System tags cannot be renamed
            if tag.system and 'name' in body:
                return self._error_response("System tags cannot be renamed", 403)

            # If renaming, check for duplicate
            if 'name' in body and body['name'].strip() != tag.name:
                existing = self.db.get_tag_by_name(user['id'], body['name'].strip())
                if existing:
                    return self._error_response(f"Tag '{body['name']}' already exists", 409)

            # Validate category if provided
            if 'category' in body and body['category'] not in VALID_TAG_CATEGORIES:
                return self._error_response(f"Invalid category: {body['category']}", 400)

            # Validate scopes if provided
            if 'visibility_scopes' in body:
                scopes = body['visibility_scopes']
                if not isinstance(scopes, list) or len(scopes) == 0:
                    return self._error_response("visibility_scopes must be a non-empty list", 400)
                if any(s not in VALID_TAG_SCOPES for s in scopes):
                    return self._error_response(f"Invalid scope values. Valid: {sorted(VALID_TAG_SCOPES)}", 400)
                if 'global' in scopes and len(scopes) > 1:
                    return self._error_response("'global' scope must be exclusive", 400)

            updates = {}
            if 'name' in body:
                updates['name'] = body['name'].strip()
            if 'description' in body:
                updates['description'] = body['description'].strip() or None
            if 'category' in body:
                updates['category'] = body['category']
            if 'visibility_scopes' in body:
                updates['visibility_scopes'] = body['visibility_scopes']

            updated = self.db.update_tag(tag_id, updates)

            return self._json_response({
                'success': True,
                'data': updated.to_api_dict()
            })
        except Exception as e:
            self.logger.error(f"update_tag error: {e}")
            return self._error_response(str(e), 500)

    async def retire_tag(self, request: web.Request) -> web.Response:
        """PUT /api/tags/:id/retire - Retire a tag (hide from suggestions)."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response("Authentication required", 401)

            tag_id = request.match_info['id']
            tag = self.db.get_tag(tag_id)

            if not tag:
                return self._error_response("Tag not found", 404)

            # Verify ownership
            if tag.user_id != user['id']:
                return self._error_response("Tag not found", 404)

            retired = self.db.retire_tag(tag_id)

            return self._json_response({
                'success': True,
                'data': retired.to_api_dict()
            })
        except Exception as e:
            self.logger.error(f"retire_tag error: {e}")
            return self._error_response(str(e), 500)

    async def restore_tag(self, request: web.Request) -> web.Response:
        """PUT /api/tags/:id/restore - Restore a retired tag."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response("Authentication required", 401)

            tag_id = request.match_info['id']
            tag = self.db.get_tag(tag_id)

            if not tag:
                return self._error_response("Tag not found", 404)

            # Verify ownership
            if tag.user_id != user['id']:
                return self._error_response("Tag not found", 404)

            restored = self.db.restore_tag(tag_id)

            return self._json_response({
                'success': True,
                'data': restored.to_api_dict()
            })
        except Exception as e:
            self.logger.error(f"restore_tag error: {e}")
            return self._error_response(str(e), 500)

    async def delete_tag(self, request: web.Request) -> web.Response:
        """DELETE /api/tags/:id - Delete a tag (only if unused)."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response("Authentication required", 401)

            tag_id = request.match_info['id']
            tag = self.db.get_tag(tag_id)

            if not tag:
                return self._error_response("Tag not found", 404)

            # Verify ownership
            if tag.user_id != user['id']:
                return self._error_response("Tag not found", 404)

            # System tags cannot be deleted
            if tag.system:
                return self._error_response("System tags cannot be deleted", 403)

            # Check if tag has been used
            if tag.usage_count > 0:
                return self._error_response(
                    f"Cannot delete tag with {tag.usage_count} uses. Retire it instead.",
                    400
                )

            deleted = self.db.delete_tag(tag_id)
            if not deleted:
                return self._error_response("Failed to delete tag", 500)

            return self._json_response({
                'success': True,
                'message': f"Tag '{tag.name}' deleted"
            })
        except Exception as e:
            self.logger.error(f"delete_tag error: {e}")
            return self._error_response(str(e), 500)

    async def seed_tags(self, request: web.Request) -> web.Response:
        """POST /api/tags/seed - Seed example tags for the current user."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response("Authentication required", 401)

            created = self.db.seed_example_tags(user['id'])

            return self._json_response({
                'success': True,
                'data': [t.to_api_dict() for t in created],
                'count': len(created),
                'message': f"Seeded {len(created)} example tags" if created else "Tags already exist"
            })
        except Exception as e:
            self.logger.error(f"seed_tags error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Settings ====================

    async def get_settings(self, request: web.Request) -> web.Response:
        """GET /api/settings - Get all settings."""
        try:
            params = request.query
            scope = params.get('scope', 'global')
            category = params.get('category')

            if category:
                settings = self.db.get_settings_by_category(category, scope)
                data = {s.key: s.get_value() for s in settings}
            else:
                data = self.db.get_all_settings(scope)

            return self._json_response({
                'success': True,
                'data': data,
                'scope': scope
            })
        except Exception as e:
            self.logger.error(f"get_settings error: {e}")
            return self._error_response(str(e), 500)

    async def get_setting(self, request: web.Request) -> web.Response:
        """GET /api/settings/:key - Get a single setting."""
        try:
            key = request.match_info['key']
            params = request.query
            scope = params.get('scope', 'global')
            log_id = params.get('log_id')

            # If log_id provided, get effective setting (with per-log override)
            if log_id:
                value = self.db.get_effective_setting(key, log_id)
                return self._json_response({
                    'success': True,
                    'data': {'key': key, 'value': value, 'effective': True}
                })

            setting = self.db.get_setting(key, scope)
            if not setting:
                return self._error_response(f"Setting {key} not found", 404)

            return self._json_response({
                'success': True,
                'data': {
                    'key': setting.key,
                    'value': setting.get_value(),
                    'category': setting.category,
                    'scope': setting.scope,
                    'description': setting.description
                }
            })
        except Exception as e:
            self.logger.error(f"get_setting error: {e}")
            return self._error_response(str(e), 500)

    async def set_setting(self, request: web.Request) -> web.Response:
        """PUT /api/settings/:key - Set a setting value."""
        try:
            key = request.match_info['key']
            body = await request.json()

            if 'value' not in body:
                return self._error_response("Missing required field: value", 400)
            if 'category' not in body:
                return self._error_response("Missing required field: category", 400)

            scope = body.get('scope', 'global')
            description = body.get('description')

            setting = self.db.set_setting(
                key=key,
                value=body['value'],
                category=body['category'],
                scope=scope,
                description=description
            )

            return self._json_response({
                'success': True,
                'data': {
                    'key': setting.key,
                    'value': setting.get_value(),
                    'category': setting.category,
                    'scope': setting.scope
                }
            })
        except Exception as e:
            self.logger.error(f"set_setting error: {e}")
            return self._error_response(str(e), 500)

    async def delete_setting(self, request: web.Request) -> web.Response:
        """DELETE /api/settings/:key - Delete a setting."""
        try:
            key = request.match_info['key']
            params = request.query
            scope = params.get('scope', 'global')

            deleted = self.db.delete_setting(key, scope)
            if not deleted:
                return self._error_response(f"Setting {key} not found", 404)

            return self._json_response({
                'success': True,
                'message': f"Setting {key} deleted"
            })
        except Exception as e:
            self.logger.error(f"delete_setting error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Journal Entry Endpoints ====================

    async def list_journal_entries(self, request: web.Request) -> web.Response:
        """GET /api/journal/entries - List journal entries for authenticated user."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            params = request.query
            entries = self.db.list_entries(
                user_id=user['id'],
                from_date=params.get('from'),
                to_date=params.get('to'),
                playbook_only=params.get('playbook_only', '').lower() == 'true',
                limit=int(params.get('limit', 100)),
                offset=int(params.get('offset', 0))
            )

            return self._json_response({
                'success': True,
                'data': [e.to_api_dict() for e in entries],
                'count': len(entries)
            })
        except Exception as e:
            self.logger.error(f"list_journal_entries error: {e}")
            return self._error_response(str(e), 500)

    async def get_journal_entry(self, request: web.Request) -> web.Response:
        """GET /api/journal/entries/:id - Get a single journal entry with attachments and trade refs."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_entry(entry_id)

            if not entry:
                return self._error_response('Entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            # Get related data
            trade_refs = self.db.list_trade_refs('entry', entry_id)
            attachments = self.db.list_attachments('entry', entry_id)

            data = entry.to_api_dict()
            data['trade_refs'] = [r.to_api_dict() for r in trade_refs]
            data['attachments'] = [a.to_api_dict() for a in attachments]

            return self._json_response({
                'success': True,
                'data': data
            })
        except Exception as e:
            self.logger.error(f"get_journal_entry error: {e}")
            return self._error_response(str(e), 500)

    async def get_journal_entry_by_date(self, request: web.Request) -> web.Response:
        """GET /api/journal/entries/date/:date - Get entry by date (YYYY-MM-DD)."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            date = request.match_info['date']
            entry = self.db.get_entry_by_date(user['id'], date)

            if not entry:
                return self._error_response('Entry not found', 404)

            # Get related data
            trade_refs = self.db.list_trade_refs('entry', entry.id)
            attachments = self.db.list_attachments('entry', entry.id)

            data = entry.to_api_dict()
            data['trade_refs'] = [r.to_api_dict() for r in trade_refs]
            data['attachments'] = [a.to_api_dict() for a in attachments]

            return self._json_response({
                'success': True,
                'data': data
            })
        except Exception as e:
            self.logger.error(f"get_journal_entry_by_date error: {e}")
            return self._error_response(str(e), 500)

    async def create_journal_entry(self, request: web.Request) -> web.Response:
        """POST /api/journal/entries - Create or update entry (upsert by date)."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            body = await request.json()

            if 'entry_date' not in body:
                return self._error_response('entry_date is required')

            # Get tags and track which ones are new
            new_tags = body.get('tags', [])

            entry = JournalEntry(
                id=JournalEntry.new_id(),
                user_id=user['id'],
                entry_date=body['entry_date'],
                content=body.get('content'),
                is_playbook_material=body.get('is_playbook_material', False),
                tags=new_tags
            )

            # Check existing entry to diff tags for usage tracking
            existing = self.db.get_entry_by_date(user['id'], body['entry_date'])
            old_tags = existing.tags if existing else []

            result = self.db.upsert_entry(entry)

            # Increment usage for newly added tags
            added_tags = set(new_tags) - set(old_tags)
            for tag_id in added_tags:
                self.db.increment_tag_usage(tag_id)
            self.logger.info(f"Upserted journal entry for {body['entry_date']}", emoji="📓")

            # Get related data to return complete entry
            trade_refs = self.db.list_trade_refs('entry', result.id)
            attachments = self.db.list_attachments('entry', result.id)

            data = result.to_api_dict()
            data['trade_refs'] = [r.to_api_dict() for r in trade_refs]
            data['attachments'] = [a.to_api_dict() for a in attachments]

            return self._json_response({
                'success': True,
                'data': data
            }, 201)
        except Exception as e:
            self.logger.error(f"create_journal_entry error: {e}")
            return self._error_response(str(e), 500)

    async def update_journal_entry(self, request: web.Request) -> web.Response:
        """PUT /api/journal/entries/:id - Update an entry."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_entry(entry_id)

            if not entry:
                return self._error_response('Entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            body = await request.json()
            updated = self.db.update_entry(entry_id, body)

            return self._json_response({
                'success': True,
                'data': updated.to_api_dict()
            })
        except Exception as e:
            self.logger.error(f"update_journal_entry error: {e}")
            return self._error_response(str(e), 500)

    async def delete_journal_entry(self, request: web.Request) -> web.Response:
        """DELETE /api/journal/entries/:id - Delete an entry."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_entry(entry_id)

            if not entry:
                return self._error_response('Entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            # Delete attachment files
            attachments = self.db.list_attachments('entry', entry_id)
            for att in attachments:
                try:
                    file_path = Path(att.file_path)
                    if file_path.exists():
                        file_path.unlink()
                except Exception:
                    pass  # Best effort deletion

            deleted = self.db.delete_entry(entry_id)
            self.logger.info(f"Deleted journal entry {entry_id}", emoji="🗑️")

            return self._json_response({
                'success': True,
                'message': 'Entry deleted'
            })
        except Exception as e:
            self.logger.error(f"delete_journal_entry error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Journal Retrospective Endpoints ====================

    async def list_retrospectives(self, request: web.Request) -> web.Response:
        """GET /api/journal/retrospectives - List retrospectives."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            params = request.query
            retros = self.db.list_retrospectives(
                user_id=user['id'],
                retro_type=params.get('type'),
                from_date=params.get('from'),
                to_date=params.get('to'),
                playbook_only=params.get('playbook_only', '').lower() == 'true'
            )

            return self._json_response({
                'success': True,
                'data': [r.to_api_dict() for r in retros],
                'count': len(retros)
            })
        except Exception as e:
            self.logger.error(f"list_retrospectives error: {e}")
            return self._error_response(str(e), 500)

    async def get_retrospective(self, request: web.Request) -> web.Response:
        """GET /api/journal/retrospectives/:id - Get a single retrospective."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            retro_id = request.match_info['id']
            retro = self.db.get_retrospective(retro_id)

            if not retro:
                return self._error_response('Retrospective not found', 404)

            if retro.user_id != user['id']:
                return self._error_response('Access denied', 403)

            # Get related data
            trade_refs = self.db.list_trade_refs('retrospective', retro_id)
            attachments = self.db.list_attachments('retrospective', retro_id)

            data = retro.to_api_dict()
            data['trade_refs'] = [r.to_api_dict() for r in trade_refs]
            data['attachments'] = [a.to_api_dict() for a in attachments]

            return self._json_response({
                'success': True,
                'data': data
            })
        except Exception as e:
            self.logger.error(f"get_retrospective error: {e}")
            return self._error_response(str(e), 500)

    async def get_retrospective_by_period(self, request: web.Request) -> web.Response:
        """GET /api/journal/retrospectives/:type/:periodStart - Get by type and period."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            retro_type = request.match_info['type']
            period_start = request.match_info['periodStart']

            if retro_type not in ('weekly', 'monthly'):
                return self._error_response('Invalid retrospective type', 400)

            retro = self.db.get_retrospective_by_period(user['id'], retro_type, period_start)

            if not retro:
                return self._error_response('Retrospective not found', 404)

            # Get related data
            trade_refs = self.db.list_trade_refs('retrospective', retro.id)
            attachments = self.db.list_attachments('retrospective', retro.id)

            data = retro.to_api_dict()
            data['trade_refs'] = [r.to_api_dict() for r in trade_refs]
            data['attachments'] = [a.to_api_dict() for a in attachments]

            return self._json_response({
                'success': True,
                'data': data
            })
        except Exception as e:
            self.logger.error(f"get_retrospective_by_period error: {e}")
            return self._error_response(str(e), 500)

    async def create_retrospective(self, request: web.Request) -> web.Response:
        """POST /api/journal/retrospectives - Create a retrospective."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            body = await request.json()

            required = ['retro_type', 'period_start', 'period_end']
            for field in required:
                if field not in body:
                    return self._error_response(f'{field} is required')

            if body['retro_type'] not in ('weekly', 'monthly'):
                return self._error_response('retro_type must be weekly or monthly')

            # Check if one already exists for this period
            existing = self.db.get_retrospective_by_period(
                user['id'], body['retro_type'], body['period_start']
            )
            if existing:
                return self._error_response('Retrospective already exists for this period', 409)

            retro = JournalRetrospective(
                id=JournalRetrospective.new_id(),
                user_id=user['id'],
                retro_type=body['retro_type'],
                period_start=body['period_start'],
                period_end=body['period_end'],
                content=body.get('content'),
                is_playbook_material=body.get('is_playbook_material', False)
            )

            created = self.db.create_retrospective(retro)
            self.logger.info(f"Created {body['retro_type']} retrospective for {body['period_start']}", emoji="📅")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201)
        except Exception as e:
            self.logger.error(f"create_retrospective error: {e}")
            return self._error_response(str(e), 500)

    async def update_retrospective(self, request: web.Request) -> web.Response:
        """PUT /api/journal/retrospectives/:id - Update a retrospective."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            retro_id = request.match_info['id']
            retro = self.db.get_retrospective(retro_id)

            if not retro:
                return self._error_response('Retrospective not found', 404)

            if retro.user_id != user['id']:
                return self._error_response('Access denied', 403)

            body = await request.json()
            updated = self.db.update_retrospective(retro_id, body)

            return self._json_response({
                'success': True,
                'data': updated.to_api_dict()
            })
        except Exception as e:
            self.logger.error(f"update_retrospective error: {e}")
            return self._error_response(str(e), 500)

    async def delete_retrospective(self, request: web.Request) -> web.Response:
        """DELETE /api/journal/retrospectives/:id - Delete a retrospective."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            retro_id = request.match_info['id']
            retro = self.db.get_retrospective(retro_id)

            if not retro:
                return self._error_response('Retrospective not found', 404)

            if retro.user_id != user['id']:
                return self._error_response('Access denied', 403)

            # Delete attachment files
            attachments = self.db.list_attachments('retrospective', retro_id)
            for att in attachments:
                try:
                    file_path = Path(att.file_path)
                    if file_path.exists():
                        file_path.unlink()
                except Exception:
                    pass

            deleted = self.db.delete_retrospective(retro_id)
            self.logger.info(f"Deleted retrospective {retro_id}", emoji="🗑️")

            return self._json_response({
                'success': True,
                'message': 'Retrospective deleted'
            })
        except Exception as e:
            self.logger.error(f"delete_retrospective error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Journal Trade Reference Endpoints ====================

    async def list_entry_trade_refs(self, request: web.Request) -> web.Response:
        """GET /api/journal/entries/:id/trades - List trade refs for an entry."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_entry(entry_id)

            if not entry:
                return self._error_response('Entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            refs = self.db.list_trade_refs('entry', entry_id)

            return self._json_response({
                'success': True,
                'data': [r.to_api_dict() for r in refs],
                'count': len(refs)
            })
        except Exception as e:
            self.logger.error(f"list_entry_trade_refs error: {e}")
            return self._error_response(str(e), 500)

    async def add_entry_trade_ref(self, request: web.Request) -> web.Response:
        """POST /api/journal/entries/:id/trades - Add trade reference to entry."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_entry(entry_id)

            if not entry:
                return self._error_response('Entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            body = await request.json()
            if 'trade_id' not in body:
                return self._error_response('trade_id is required')

            ref = JournalTradeRef(
                id=JournalTradeRef.new_id(),
                source_type='entry',
                source_id=entry_id,
                trade_id=body['trade_id'],
                note=body.get('note')
            )

            created = self.db.add_trade_ref(ref)

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201)
        except Exception as e:
            self.logger.error(f"add_entry_trade_ref error: {e}")
            return self._error_response(str(e), 500)

    async def list_retro_trade_refs(self, request: web.Request) -> web.Response:
        """GET /api/journal/retrospectives/:id/trades - List trade refs for a retrospective."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            retro_id = request.match_info['id']
            retro = self.db.get_retrospective(retro_id)

            if not retro:
                return self._error_response('Retrospective not found', 404)

            if retro.user_id != user['id']:
                return self._error_response('Access denied', 403)

            refs = self.db.list_trade_refs('retrospective', retro_id)

            return self._json_response({
                'success': True,
                'data': [r.to_api_dict() for r in refs],
                'count': len(refs)
            })
        except Exception as e:
            self.logger.error(f"list_retro_trade_refs error: {e}")
            return self._error_response(str(e), 500)

    async def add_retro_trade_ref(self, request: web.Request) -> web.Response:
        """POST /api/journal/retrospectives/:id/trades - Add trade reference to retrospective."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            retro_id = request.match_info['id']
            retro = self.db.get_retrospective(retro_id)

            if not retro:
                return self._error_response('Retrospective not found', 404)

            if retro.user_id != user['id']:
                return self._error_response('Access denied', 403)

            body = await request.json()
            if 'trade_id' not in body:
                return self._error_response('trade_id is required')

            ref = JournalTradeRef(
                id=JournalTradeRef.new_id(),
                source_type='retrospective',
                source_id=retro_id,
                trade_id=body['trade_id'],
                note=body.get('note')
            )

            created = self.db.add_trade_ref(ref)

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201)
        except Exception as e:
            self.logger.error(f"add_retro_trade_ref error: {e}")
            return self._error_response(str(e), 500)

    async def delete_trade_ref(self, request: web.Request) -> web.Response:
        """DELETE /api/journal/trade-refs/:id - Remove a trade reference."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            ref_id = request.match_info['id']

            # We don't have direct ownership check on refs, but the delete is safe
            deleted = self.db.delete_trade_ref(ref_id)
            if not deleted:
                return self._error_response('Trade reference not found', 404)

            return self._json_response({
                'success': True,
                'message': 'Trade reference deleted'
            })
        except Exception as e:
            self.logger.error(f"delete_trade_ref error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Journal Attachment Endpoints ====================

    def _get_storage_path(self, user_id: int, filename: str) -> Path:
        """Generate storage path for an attachment."""
        now = datetime.utcnow()
        import uuid
        unique_name = f"{uuid.uuid4()}_{filename}"
        path = self.attachments_path / str(user_id) / str(now.year) / f"{now.month:02d}"
        path.mkdir(parents=True, exist_ok=True)
        return path / unique_name

    async def list_entry_attachments(self, request: web.Request) -> web.Response:
        """GET /api/journal/entries/:id/attachments - List attachments for an entry."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_entry(entry_id)

            if not entry:
                return self._error_response('Entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            attachments = self.db.list_attachments('entry', entry_id)

            return self._json_response({
                'success': True,
                'data': [a.to_api_dict() for a in attachments],
                'count': len(attachments)
            })
        except Exception as e:
            self.logger.error(f"list_entry_attachments error: {e}")
            return self._error_response(str(e), 500)

    async def upload_entry_attachment(self, request: web.Request) -> web.Response:
        """POST /api/journal/entries/:id/attachments - Upload attachment to entry."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_entry(entry_id)

            if not entry:
                return self._error_response('Entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            # Read multipart data
            reader = await request.multipart()
            field = await reader.next()

            if not field:
                return self._error_response('No file uploaded', 400)

            filename = field.filename or 'upload'
            content_type = field.headers.get('Content-Type', 'application/octet-stream')

            # Read file with size limit
            chunks = []
            size = 0
            while True:
                chunk = await field.read_chunk()
                if not chunk:
                    break
                size += len(chunk)
                if size > self.max_attachment_size:
                    return self._error_response(f'File too large (max {self.max_attachment_size // 1024 // 1024}MB)', 400)
                chunks.append(chunk)

            content = b''.join(chunks)

            # Save file
            storage_path = self._get_storage_path(user['id'], filename)
            with open(storage_path, 'wb') as f:
                f.write(content)

            # Create attachment record
            attachment = JournalAttachment(
                id=JournalAttachment.new_id(),
                source_type='entry',
                source_id=entry_id,
                filename=filename,
                file_path=str(storage_path),
                mime_type=content_type,
                file_size=size
            )

            created = self.db.create_attachment(attachment)
            self.logger.info(f"Uploaded attachment '{filename}' to entry {entry_id}", emoji="📎")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201)
        except Exception as e:
            self.logger.error(f"upload_entry_attachment error: {e}")
            return self._error_response(str(e), 500)

    async def list_retro_attachments(self, request: web.Request) -> web.Response:
        """GET /api/journal/retrospectives/:id/attachments - List attachments for a retrospective."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            retro_id = request.match_info['id']
            retro = self.db.get_retrospective(retro_id)

            if not retro:
                return self._error_response('Retrospective not found', 404)

            if retro.user_id != user['id']:
                return self._error_response('Access denied', 403)

            attachments = self.db.list_attachments('retrospective', retro_id)

            return self._json_response({
                'success': True,
                'data': [a.to_api_dict() for a in attachments],
                'count': len(attachments)
            })
        except Exception as e:
            self.logger.error(f"list_retro_attachments error: {e}")
            return self._error_response(str(e), 500)

    async def upload_retro_attachment(self, request: web.Request) -> web.Response:
        """POST /api/journal/retrospectives/:id/attachments - Upload attachment to retrospective."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            retro_id = request.match_info['id']
            retro = self.db.get_retrospective(retro_id)

            if not retro:
                return self._error_response('Retrospective not found', 404)

            if retro.user_id != user['id']:
                return self._error_response('Access denied', 403)

            # Read multipart data
            reader = await request.multipart()
            field = await reader.next()

            if not field:
                return self._error_response('No file uploaded', 400)

            filename = field.filename or 'upload'
            content_type = field.headers.get('Content-Type', 'application/octet-stream')

            # Read file with size limit
            chunks = []
            size = 0
            while True:
                chunk = await field.read_chunk()
                if not chunk:
                    break
                size += len(chunk)
                if size > self.max_attachment_size:
                    return self._error_response(f'File too large (max {self.max_attachment_size // 1024 // 1024}MB)', 400)
                chunks.append(chunk)

            content = b''.join(chunks)

            # Save file
            storage_path = self._get_storage_path(user['id'], filename)
            with open(storage_path, 'wb') as f:
                f.write(content)

            # Create attachment record
            attachment = JournalAttachment(
                id=JournalAttachment.new_id(),
                source_type='retrospective',
                source_id=retro_id,
                filename=filename,
                file_path=str(storage_path),
                mime_type=content_type,
                file_size=size
            )

            created = self.db.create_attachment(attachment)
            self.logger.info(f"Uploaded attachment '{filename}' to retrospective {retro_id}", emoji="📎")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201)
        except Exception as e:
            self.logger.error(f"upload_retro_attachment error: {e}")
            return self._error_response(str(e), 500)

    async def download_attachment(self, request: web.Request) -> web.Response:
        """GET /api/journal/attachments/:id - Download an attachment."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            attachment_id = request.match_info['id']
            attachment = self.db.get_attachment(attachment_id)

            if not attachment:
                return self._error_response('Attachment not found', 404)

            # Verify ownership through parent entry/retrospective
            if attachment.source_type == 'entry':
                entry = self.db.get_entry(attachment.source_id)
                if not entry or entry.user_id != user['id']:
                    return self._error_response('Access denied', 403)
            else:
                retro = self.db.get_retrospective(attachment.source_id)
                if not retro or retro.user_id != user['id']:
                    return self._error_response('Access denied', 403)

            file_path = Path(attachment.file_path)
            if not file_path.exists():
                return self._error_response('File not found', 404)

            with open(file_path, 'rb') as f:
                content = f.read()

            return web.Response(
                body=content,
                content_type=attachment.mime_type or 'application/octet-stream',
                headers={
                    'Content-Disposition': f'attachment; filename="{attachment.filename}"',
                    'Access-Control-Allow-Origin': '*',
                }
            )
        except Exception as e:
            self.logger.error(f"download_attachment error: {e}")
            return self._error_response(str(e), 500)

    async def delete_attachment(self, request: web.Request) -> web.Response:
        """DELETE /api/journal/attachments/:id - Delete an attachment."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            attachment_id = request.match_info['id']
            attachment = self.db.get_attachment(attachment_id)

            if not attachment:
                return self._error_response('Attachment not found', 404)

            # Verify ownership through parent entry/retrospective
            if attachment.source_type == 'entry':
                entry = self.db.get_entry(attachment.source_id)
                if not entry or entry.user_id != user['id']:
                    return self._error_response('Access denied', 403)
            else:
                retro = self.db.get_retrospective(attachment.source_id)
                if not retro or retro.user_id != user['id']:
                    return self._error_response('Access denied', 403)

            # Delete file
            try:
                file_path = Path(attachment.file_path)
                if file_path.exists():
                    file_path.unlink()
            except Exception:
                pass  # Best effort

            # Delete record
            self.db.delete_attachment(attachment_id)
            self.logger.info(f"Deleted attachment {attachment_id}", emoji="🗑️")

            return self._json_response({
                'success': True,
                'message': 'Attachment deleted'
            })
        except Exception as e:
            self.logger.error(f"delete_attachment error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Calendar View Endpoints (Temporal Gravity) ====================

    async def get_calendar_month(self, request: web.Request) -> web.Response:
        """GET /api/journal/calendar/:year/:month - Get month view with entry indicators."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            year = int(request.match_info['year'])
            month = int(request.match_info['month'])

            if month < 1 or month > 12:
                return self._error_response('Invalid month', 400)

            data = self.db.get_calendar_month(user['id'], year, month)

            return self._json_response({
                'success': True,
                'data': data
            })
        except ValueError:
            return self._error_response('Invalid year or month', 400)
        except Exception as e:
            self.logger.error(f"get_calendar_month error: {e}")
            return self._error_response(str(e), 500)

    async def get_calendar_week(self, request: web.Request) -> web.Response:
        """GET /api/journal/calendar/week/:weekStart - Get week view with entry indicators."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            week_start = request.match_info['weekStart']

            # Validate date format
            try:
                datetime.strptime(week_start, '%Y-%m-%d')
            except ValueError:
                return self._error_response('Invalid date format (use YYYY-MM-DD)', 400)

            data = self.db.get_calendar_week(user['id'], week_start)

            return self._json_response({
                'success': True,
                'data': data
            })
        except Exception as e:
            self.logger.error(f"get_calendar_week error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Playbook Endpoints ====================

    async def list_playbook_entries(self, request: web.Request) -> web.Response:
        """GET /api/playbook/entries - List playbook entries with optional filters."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            # Parse query params
            entry_type = request.query.get('type')
            status = request.query.get('status')
            search = request.query.get('search')
            limit = int(request.query.get('limit', 100))
            offset = int(request.query.get('offset', 0))

            entries = self.db.list_playbook_entries(
                user['id'],
                entry_type=entry_type,
                status=status,
                search=search,
                limit=limit,
                offset=offset
            )

            # Get source counts for each entry
            entries_with_sources = []
            for entry in entries:
                data = entry.to_api_dict()
                sources = self.db.list_playbook_source_refs(entry.id)
                data['source_count'] = len(sources)
                entries_with_sources.append(data)

            return self._json_response({
                'success': True,
                'data': entries_with_sources,
                'count': len(entries_with_sources)
            })
        except Exception as e:
            self.logger.error(f"list_playbook_entries error: {e}")
            return self._error_response(str(e), 500)

    async def get_playbook_entry(self, request: web.Request) -> web.Response:
        """GET /api/playbook/entries/:id - Get a playbook entry with sources."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_playbook_entry(entry_id)

            if not entry:
                return self._error_response('Playbook entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            # Get sources
            sources = self.db.list_playbook_source_refs(entry_id)

            data = entry.to_api_dict()
            data['sources'] = [s.to_api_dict() for s in sources]

            return self._json_response({
                'success': True,
                'data': data
            })
        except Exception as e:
            self.logger.error(f"get_playbook_entry error: {e}")
            return self._error_response(str(e), 500)

    async def create_playbook_entry(self, request: web.Request) -> web.Response:
        """POST /api/playbook/entries - Create a new playbook entry."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            body = await request.json()

            # Validate required fields
            if not body.get('title'):
                return self._error_response('Title is required', 400)
            if not body.get('entry_type'):
                return self._error_response('Entry type is required', 400)

            valid_types = ['pattern', 'rule', 'warning', 'filter', 'constraint']
            if body['entry_type'] not in valid_types:
                return self._error_response(f'Invalid entry type. Must be one of: {valid_types}', 400)

            entry = PlaybookEntry(
                id=PlaybookEntry.new_id(),
                user_id=user['id'],
                title=body['title'],
                entry_type=body['entry_type'],
                description=body.get('description', ''),
                status=body.get('status', 'draft')
            )

            created = self.db.create_playbook_entry(entry)

            # Add source references if provided
            sources = body.get('sources', [])
            for src in sources:
                if src.get('source_type') and src.get('source_id'):
                    ref = PlaybookSourceRef(
                        id=PlaybookSourceRef.new_id(),
                        playbook_entry_id=created.id,
                        source_type=src['source_type'],
                        source_id=src['source_id'],
                        note=src.get('note')
                    )
                    self.db.create_playbook_source_ref(ref)

            self.logger.info(f"Created playbook entry '{entry.title}'", emoji="📚")

            # Return with sources
            source_refs = self.db.list_playbook_source_refs(created.id)
            data = created.to_api_dict()
            data['sources'] = [s.to_api_dict() for s in source_refs]

            return self._json_response({
                'success': True,
                'data': data
            }, 201)
        except Exception as e:
            self.logger.error(f"create_playbook_entry error: {e}")
            return self._error_response(str(e), 500)

    async def update_playbook_entry(self, request: web.Request) -> web.Response:
        """PUT /api/playbook/entries/:id - Update a playbook entry."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_playbook_entry(entry_id)

            if not entry:
                return self._error_response('Playbook entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            body = await request.json()

            # Build updates
            updates = {}
            if 'title' in body:
                updates['title'] = body['title']
            if 'entry_type' in body:
                valid_types = ['pattern', 'rule', 'warning', 'filter', 'constraint']
                if body['entry_type'] not in valid_types:
                    return self._error_response(f'Invalid entry type. Must be one of: {valid_types}', 400)
                updates['entry_type'] = body['entry_type']
            if 'description' in body:
                updates['description'] = body['description']
            if 'status' in body:
                valid_statuses = ['draft', 'active', 'retired']
                if body['status'] not in valid_statuses:
                    return self._error_response(f'Invalid status. Must be one of: {valid_statuses}', 400)
                updates['status'] = body['status']

            if not updates:
                return self._error_response('No valid fields to update', 400)

            updated = self.db.update_playbook_entry(entry_id, updates)

            if not updated:
                return self._error_response('Failed to update entry', 500)

            # Return with sources
            sources = self.db.list_playbook_source_refs(entry_id)
            data = updated.to_api_dict()
            data['sources'] = [s.to_api_dict() for s in sources]

            return self._json_response({
                'success': True,
                'data': data
            })
        except Exception as e:
            self.logger.error(f"update_playbook_entry error: {e}")
            return self._error_response(str(e), 500)

    async def delete_playbook_entry(self, request: web.Request) -> web.Response:
        """DELETE /api/playbook/entries/:id - Delete a playbook entry."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_playbook_entry(entry_id)

            if not entry:
                return self._error_response('Playbook entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            self.db.delete_playbook_entry(entry_id)
            self.logger.info(f"Deleted playbook entry {entry_id}", emoji="🗑️")

            return self._json_response({
                'success': True,
                'message': 'Playbook entry deleted'
            })
        except Exception as e:
            self.logger.error(f"delete_playbook_entry error: {e}")
            return self._error_response(str(e), 500)

    async def list_playbook_sources(self, request: web.Request) -> web.Response:
        """GET /api/playbook/entries/:id/sources - List sources for a playbook entry."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_playbook_entry(entry_id)

            if not entry:
                return self._error_response('Playbook entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            sources = self.db.list_playbook_source_refs(entry_id)

            return self._json_response({
                'success': True,
                'data': [s.to_api_dict() for s in sources],
                'count': len(sources)
            })
        except Exception as e:
            self.logger.error(f"list_playbook_sources error: {e}")
            return self._error_response(str(e), 500)

    async def add_playbook_source(self, request: web.Request) -> web.Response:
        """POST /api/playbook/entries/:id/sources - Add a source reference."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            entry_id = request.match_info['id']
            entry = self.db.get_playbook_entry(entry_id)

            if not entry:
                return self._error_response('Playbook entry not found', 404)

            if entry.user_id != user['id']:
                return self._error_response('Access denied', 403)

            body = await request.json()

            if not body.get('source_type'):
                return self._error_response('source_type is required', 400)
            if not body.get('source_id'):
                return self._error_response('source_id is required', 400)

            valid_types = ['entry', 'retrospective', 'trade']
            if body['source_type'] not in valid_types:
                return self._error_response(f'Invalid source_type. Must be one of: {valid_types}', 400)

            ref = PlaybookSourceRef(
                id=PlaybookSourceRef.new_id(),
                playbook_entry_id=entry_id,
                source_type=body['source_type'],
                source_id=body['source_id'],
                note=body.get('note')
            )

            created = self.db.create_playbook_source_ref(ref)

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201)
        except Exception as e:
            self.logger.error(f"add_playbook_source error: {e}")
            return self._error_response(str(e), 500)

    async def delete_playbook_source(self, request: web.Request) -> web.Response:
        """DELETE /api/playbook/sources/:id - Delete a source reference."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            ref_id = request.match_info['id']
            # Note: We should verify ownership through the playbook entry
            # For now, just delete if exists

            deleted = self.db.delete_playbook_source_ref(ref_id)

            if not deleted:
                return self._error_response('Source reference not found', 404)

            return self._json_response({
                'success': True,
                'message': 'Source reference deleted'
            })
        except Exception as e:
            self.logger.error(f"delete_playbook_source error: {e}")
            return self._error_response(str(e), 500)

    async def get_flagged_playbook_material(self, request: web.Request) -> web.Response:
        """GET /api/playbook/flagged-material - Get all items flagged as playbook material."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401)

            data = self.db.list_flagged_playbook_material(user['id'])

            return self._json_response({
                'success': True,
                'data': data
            })
        except Exception as e:
            self.logger.error(f"get_flagged_playbook_material error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Alerts API ====================

    async def _get_redis(self) -> redis.Redis:
        """Get or create Redis connection."""
        if self._redis is None:
            self._redis = redis.from_url(self._redis_url)
        return self._redis

    async def _publish_alerts_sync(self, action: str, alert_id: str):
        """Publish alert sync message to Redis for Copilot to reload."""
        try:
            r = await self._get_redis()
            await r.publish('alerts:sync', json.dumps({
                'action': action,
                'alert_id': alert_id,
                'ts': datetime.utcnow().isoformat()
            }))
        except Exception as e:
            self.logger.warn(f"Failed to publish alerts:sync: {e}")

    async def list_alerts(self, request: web.Request) -> web.Response:
        """GET /api/alerts - List alerts for the current user."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            params = request.query
            enabled = None
            if 'enabled' in params:
                enabled = params.get('enabled', '').lower() == 'true'

            triggered = None
            if 'triggered' in params:
                triggered = params.get('triggered', '').lower() == 'true'

            alerts = self.db.list_alerts(
                user_id=user['id'],
                enabled=enabled,
                alert_type=params.get('type'),
                source_id=params.get('source_id'),
                triggered=triggered
            )

            return self._json_response({
                'success': True,
                'data': [a.to_api_dict() for a in alerts],
                'count': len(alerts)
            }, request=request)
        except Exception as e:
            self.logger.error(f"list_alerts error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_alert(self, request: web.Request) -> web.Response:
        """GET /api/alerts/:id - Get a single alert."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            alert_id = request.match_info['id']
            alert = self.db.get_alert(alert_id)

            if not alert:
                return self._error_response('Alert not found', 404, request)

            if alert.user_id != user['id']:
                return self._error_response('Alert not found', 404, request)

            return self._json_response({
                'success': True,
                'data': alert.to_api_dict()
            }, request=request)
        except Exception as e:
            self.logger.error(f"get_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def create_alert(self, request: web.Request) -> web.Response:
        """POST /api/alerts - Create a new alert."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            body = await request.json()

            # Validate required fields
            if 'type' not in body:
                return self._error_response('type is required', 400, request)
            if 'condition' not in body:
                return self._error_response('condition is required', 400, request)

            # Map intent_class based on type if not provided
            intent_class = body.get('intentClass', body.get('intent_class'))
            if not intent_class:
                alert_type = body['type']
                if alert_type in ('trade_closed', 'pattern_recurrence'):
                    intent_class = 'reflective'
                elif alert_type.startswith('ai_') or alert_type in ('portfolio_pnl', 'portfolio_trailing', 'greeks_threshold'):
                    intent_class = 'protective'
                else:
                    intent_class = 'informational'

            alert = Alert(
                id=Alert.new_id(),
                user_id=user['id'],
                type=body['type'],
                intent_class=intent_class,
                condition=body['condition'],
                target_value=body.get('targetValue', body.get('target_value')),
                behavior=body.get('behavior', 'once_only'),
                priority=body.get('priority', 'medium'),
                source_type=body.get('sourceType', body.get('source_type', 'symbol')),
                source_id=body.get('sourceId', body.get('source_id', '')),
                strategy_id=body.get('strategyId', body.get('strategy_id')),
                entry_debit=body.get('entryDebit', body.get('entry_debit')),
                min_profit_threshold=body.get('minProfitThreshold', body.get('min_profit_threshold')),
                zone_low=body.get('zoneLow', body.get('zone_low')),
                zone_high=body.get('zoneHigh', body.get('zone_high')),
                high_water_mark=body.get('highWaterMark', body.get('high_water_mark')),
                label=body.get('label'),
                color=body.get('color', '#3b82f6'),
            )

            created = self.db.create_alert(alert)

            # Publish sync message for Copilot
            await self._publish_alerts_sync('create', created.id)

            self.logger.info(f"Created alert {created.type} for user {user['id']}", emoji="🔔")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201, request)

        except Exception as e:
            self.logger.error(f"create_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def update_alert(self, request: web.Request) -> web.Response:
        """PATCH /api/alerts/:id - Update an alert."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            alert_id = request.match_info['id']
            alert = self.db.get_alert(alert_id)

            if not alert:
                return self._error_response('Alert not found', 404, request)

            if alert.user_id != user['id']:
                return self._error_response('Alert not found', 404, request)

            body = await request.json()

            # Map camelCase to snake_case for DB
            updates = {}
            field_map = {
                'targetValue': 'target_value',
                'intentClass': 'intent_class',
                'sourceType': 'source_type',
                'sourceId': 'source_id',
                'strategyId': 'strategy_id',
                'entryDebit': 'entry_debit',
                'minProfitThreshold': 'min_profit_threshold',
                'zoneLow': 'zone_low',
                'zoneHigh': 'zone_high',
                'aiConfidence': 'ai_confidence',
                'aiReasoning': 'ai_reasoning',
                'highWaterMark': 'high_water_mark',
                'triggeredAt': 'triggered_at',
                'triggerCount': 'trigger_count',
            }

            for key, value in body.items():
                db_key = field_map.get(key, key)
                updates[db_key] = value

            updated = self.db.update_alert(alert_id, updates)

            # Publish sync message for Copilot
            await self._publish_alerts_sync('update', alert_id)

            return self._json_response({
                'success': True,
                'data': updated.to_api_dict()
            }, request=request)
        except Exception as e:
            self.logger.error(f"update_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def delete_alert(self, request: web.Request) -> web.Response:
        """DELETE /api/alerts/:id - Delete an alert."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            alert_id = request.match_info['id']
            alert = self.db.get_alert(alert_id)

            if not alert:
                return self._error_response('Alert not found', 404, request)

            if alert.user_id != user['id']:
                return self._error_response('Alert not found', 404, request)

            deleted = self.db.delete_alert(alert_id)

            # Publish sync message for Copilot
            await self._publish_alerts_sync('delete', alert_id)

            self.logger.info(f"Deleted alert {alert_id}", emoji="🗑️")

            return self._json_response({
                'success': True,
                'message': 'Alert deleted'
            }, request=request)
        except Exception as e:
            self.logger.error(f"delete_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def reset_alert(self, request: web.Request) -> web.Response:
        """POST /api/alerts/:id/reset - Reset an alert's trigger state."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            alert_id = request.match_info['id']
            alert = self.db.get_alert(alert_id)

            if not alert:
                return self._error_response('Alert not found', 404, request)

            if alert.user_id != user['id']:
                return self._error_response('Alert not found', 404, request)

            reset = self.db.reset_alert(alert_id)

            # Publish sync message for Copilot
            await self._publish_alerts_sync('reset', alert_id)

            return self._json_response({
                'success': True,
                'data': reset.to_api_dict()
            }, request=request)
        except Exception as e:
            self.logger.error(f"reset_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def list_all_alerts_internal(self, request: web.Request) -> web.Response:
        """GET /api/internal/alerts - List all enabled alerts (for Copilot evaluation).

        This is an internal endpoint for service-to-service communication.
        No user auth required - only accessible from localhost.
        """
        try:
            # Basic security: only allow from localhost
            peername = request.transport.get_extra_info('peername')
            if peername:
                client_ip = peername[0]
                if client_ip not in ('127.0.0.1', '::1', 'localhost'):
                    return self._error_response('Internal endpoint only', 403, request)

            alerts = self.db.get_all_enabled_alerts()

            return self._json_response({
                'success': True,
                'data': [a.to_api_dict() for a in alerts],
                'count': len(alerts)
            }, request=request)
        except Exception as e:
            self.logger.error(f"list_all_alerts_internal error: {e}")
            return self._error_response(str(e), 500, request)

    # ==================== Prompt Alerts ====================

    async def _publish_prompt_alerts_sync(self, action: str, alert_id: str):
        """Publish prompt alert sync message to Redis for Copilot."""
        try:
            r = await self._get_redis()
            await r.publish('prompt_alerts:sync', json.dumps({
                'action': action,
                'alert_id': alert_id,
                'timestamp': datetime.utcnow().isoformat()
            }))
        except Exception as e:
            self.logger.warn(f"Failed to publish prompt_alerts:sync: {e}")

    async def list_prompt_alerts(self, request: web.Request) -> web.Response:
        """GET /api/prompt-alerts - List prompt alerts for the current user."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            params = request.query

            alerts = self.db.list_prompt_alerts(
                user_id=user['id'],
                strategy_id=params.get('strategyId') or params.get('strategy_id'),
                lifecycle_state=params.get('lifecycleState') or params.get('lifecycle_state'),
                orchestration_group_id=params.get('groupId') or params.get('orchestration_group_id')
            )

            # Optionally include reference state for each alert
            include_reference = params.get('includeReference', '').lower() == 'true'
            data = []
            for alert in alerts:
                alert_dict = alert.to_api_dict()
                if include_reference:
                    ref = self.db.get_reference_snapshot(alert.id)
                    if ref:
                        alert_dict['referenceState'] = ref.to_api_dict()
                data.append(alert_dict)

            return self._json_response({
                'success': True,
                'data': data,
                'count': len(data)
            }, request=request)
        except Exception as e:
            self.logger.error(f"list_prompt_alerts error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_prompt_alert(self, request: web.Request) -> web.Response:
        """GET /api/prompt-alerts/:id - Get a single prompt alert with history."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            alert_id = request.match_info['id']
            alert = self.db.get_prompt_alert(alert_id)

            if not alert:
                return self._error_response('Prompt alert not found', 404, request)

            if alert.user_id != user['id']:
                return self._error_response('Prompt alert not found', 404, request)

            # Get related data
            reference = self.db.get_reference_snapshot(alert_id)
            versions = self.db.get_prompt_alert_versions(alert_id)
            triggers = self.db.get_prompt_alert_triggers(alert_id)

            data = alert.to_api_dict()
            data['referenceState'] = reference.to_api_dict() if reference else None
            data['versions'] = [v.to_api_dict() for v in versions]
            data['triggers'] = [t.to_api_dict() for t in triggers]

            return self._json_response({
                'success': True,
                'data': data
            }, request=request)
        except Exception as e:
            self.logger.error(f"get_prompt_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def create_prompt_alert(self, request: web.Request) -> web.Response:
        """POST /api/prompt-alerts - Create a new prompt alert."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            body = await request.json()

            # Validate required fields
            if 'promptText' not in body and 'prompt_text' not in body:
                return self._error_response('promptText is required', 400, request)
            if 'strategyId' not in body and 'strategy_id' not in body:
                return self._error_response('strategyId is required', 400, request)

            prompt_text = body.get('promptText', body.get('prompt_text', ''))
            strategy_id = body.get('strategyId', body.get('strategy_id', ''))

            # Create the alert
            alert = PromptAlert(
                id=PromptAlert.new_id(),
                user_id=user['id'],
                strategy_id=strategy_id,
                prompt_text=prompt_text,
                prompt_version=1,
                confidence_threshold=body.get('confidenceThreshold', body.get('confidence_threshold', 'medium')),
                orchestration_mode=body.get('orchestrationMode', body.get('orchestration_mode', 'parallel')),
                orchestration_group_id=body.get('orchestrationGroupId', body.get('orchestration_group_id')),
                sequence_order=body.get('sequenceOrder', body.get('sequence_order', 0)),
                activates_after_alert_id=body.get('activatesAfterAlertId', body.get('activates_after_alert_id')),
            )

            # If parsed zones are provided (from AI parsing), store them
            if 'parsedReferenceLogic' in body or 'parsed_reference_logic' in body:
                alert.parsed_reference_logic = json.dumps(
                    body.get('parsedReferenceLogic', body.get('parsed_reference_logic'))
                )
            if 'parsedDeviationLogic' in body or 'parsed_deviation_logic' in body:
                alert.parsed_deviation_logic = json.dumps(
                    body.get('parsedDeviationLogic', body.get('parsed_deviation_logic'))
                )
            if 'parsedEvaluationMode' in body or 'parsed_evaluation_mode' in body:
                alert.parsed_evaluation_mode = body.get('parsedEvaluationMode', body.get('parsed_evaluation_mode'))
            if 'parsedStageThresholds' in body or 'parsed_stage_thresholds' in body:
                alert.parsed_stage_thresholds = json.dumps(
                    body.get('parsedStageThresholds', body.get('parsed_stage_thresholds'))
                )

            created = self.db.create_prompt_alert(alert)

            # Create initial version record
            version = PromptAlertVersion(
                id=PromptAlertVersion.new_id(),
                prompt_alert_id=created.id,
                version=1,
                prompt_text=prompt_text,
                parsed_zones=json.dumps({
                    'reference_logic': body.get('parsedReferenceLogic', body.get('parsed_reference_logic')),
                    'deviation_logic': body.get('parsedDeviationLogic', body.get('parsed_deviation_logic')),
                    'evaluation_mode': body.get('parsedEvaluationMode', body.get('parsed_evaluation_mode')),
                    'stage_thresholds': body.get('parsedStageThresholds', body.get('parsed_stage_thresholds')),
                })
            )
            self.db.create_prompt_alert_version(version)

            # Create reference state snapshot if provided
            if 'referenceState' in body or 'reference_state' in body:
                ref_data = body.get('referenceState', body.get('reference_state', {}))
                snapshot = ReferenceStateSnapshot(
                    id=ReferenceStateSnapshot.new_id(),
                    prompt_alert_id=created.id,
                    delta=ref_data.get('delta'),
                    gamma=ref_data.get('gamma'),
                    theta=ref_data.get('theta'),
                    expiration_breakevens=json.dumps(ref_data.get('expirationBreakevens')) if ref_data.get('expirationBreakevens') else None,
                    theoretical_breakevens=json.dumps(ref_data.get('theoreticalBreakevens')) if ref_data.get('theoreticalBreakevens') else None,
                    max_profit=ref_data.get('maxProfit'),
                    max_loss=ref_data.get('maxLoss'),
                    pnl_at_spot=ref_data.get('pnlAtSpot'),
                    spot_price=ref_data.get('spotPrice'),
                    vix=ref_data.get('vix'),
                    market_regime=ref_data.get('marketRegime'),
                    dte=ref_data.get('dte'),
                    debit=ref_data.get('debit'),
                    strike=ref_data.get('strike'),
                    width=ref_data.get('width'),
                    side=ref_data.get('side'),
                )
                self.db.create_reference_snapshot(snapshot)

            # Publish sync message for Copilot
            await self._publish_prompt_alerts_sync('create', created.id)

            self.logger.info(f"Created prompt alert for strategy {strategy_id}", emoji="📝")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201, request)

        except Exception as e:
            self.logger.error(f"create_prompt_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def update_prompt_alert(self, request: web.Request) -> web.Response:
        """PATCH /api/prompt-alerts/:id - Update a prompt alert (creates new version if text changes)."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            alert_id = request.match_info['id']
            alert = self.db.get_prompt_alert(alert_id)

            if not alert:
                return self._error_response('Prompt alert not found', 404, request)

            if alert.user_id != user['id']:
                return self._error_response('Prompt alert not found', 404, request)

            body = await request.json()

            # Map camelCase to snake_case
            updates = {}
            field_map = {
                'promptText': 'prompt_text',
                'confidenceThreshold': 'confidence_threshold',
                'orchestrationMode': 'orchestration_mode',
                'orchestrationGroupId': 'orchestration_group_id',
                'sequenceOrder': 'sequence_order',
                'activatesAfterAlertId': 'activates_after_alert_id',
                'lifecycleState': 'lifecycle_state',
                'currentStage': 'current_stage',
                'lastAiConfidence': 'last_ai_confidence',
                'lastAiReasoning': 'last_ai_reasoning',
                'lastEvaluationAt': 'last_evaluation_at',
                'parsedReferenceLogic': 'parsed_reference_logic',
                'parsedDeviationLogic': 'parsed_deviation_logic',
                'parsedEvaluationMode': 'parsed_evaluation_mode',
                'parsedStageThresholds': 'parsed_stage_thresholds',
            }

            prompt_text_changed = False
            new_prompt_text = None

            for key, value in body.items():
                db_key = field_map.get(key, key)

                # JSON-encode complex fields
                if db_key in ('parsed_reference_logic', 'parsed_deviation_logic', 'parsed_stage_thresholds'):
                    if isinstance(value, dict):
                        value = json.dumps(value)

                updates[db_key] = value

                # Track if prompt text changed
                if db_key == 'prompt_text' and value != alert.prompt_text:
                    prompt_text_changed = True
                    new_prompt_text = value

            # If prompt text changed, create new version
            if prompt_text_changed and new_prompt_text:
                new_version = alert.prompt_version + 1
                updates['prompt_version'] = new_version

                # Reset stage on prompt edit
                updates['current_stage'] = 'watching'

                # Create version record
                version = PromptAlertVersion(
                    id=PromptAlertVersion.new_id(),
                    prompt_alert_id=alert_id,
                    version=new_version,
                    prompt_text=new_prompt_text,
                    parsed_zones=json.dumps({
                        'reference_logic': body.get('parsedReferenceLogic', body.get('parsed_reference_logic')),
                        'deviation_logic': body.get('parsedDeviationLogic', body.get('parsed_deviation_logic')),
                        'evaluation_mode': body.get('parsedEvaluationMode', body.get('parsed_evaluation_mode')),
                        'stage_thresholds': body.get('parsedStageThresholds', body.get('parsed_stage_thresholds')),
                    })
                )
                self.db.create_prompt_alert_version(version)

            updated = self.db.update_prompt_alert(alert_id, updates)

            # Publish sync message for Copilot
            await self._publish_prompt_alerts_sync('update', alert_id)

            return self._json_response({
                'success': True,
                'data': updated.to_api_dict()
            }, request=request)
        except Exception as e:
            self.logger.error(f"update_prompt_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def delete_prompt_alert(self, request: web.Request) -> web.Response:
        """DELETE /api/prompt-alerts/:id - Delete a prompt alert."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            alert_id = request.match_info['id']
            alert = self.db.get_prompt_alert(alert_id)

            if not alert:
                return self._error_response('Prompt alert not found', 404, request)

            if alert.user_id != user['id']:
                return self._error_response('Prompt alert not found', 404, request)

            deleted = self.db.delete_prompt_alert(alert_id)

            # Publish sync message for Copilot
            await self._publish_prompt_alerts_sync('delete', alert_id)

            self.logger.info(f"Deleted prompt alert {alert_id}", emoji="🗑️")

            return self._json_response({
                'success': True,
                'deleted': deleted
            }, request=request)
        except Exception as e:
            self.logger.error(f"delete_prompt_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def list_prompt_alerts_internal(self, request: web.Request) -> web.Response:
        """GET /api/internal/prompt-alerts - List all active prompt alerts (for Copilot evaluation).

        Internal endpoint for service-to-service communication.
        """
        try:
            # Basic security: only allow from localhost
            peername = request.transport.get_extra_info('peername')
            if peername:
                client_ip = peername[0]
                if client_ip not in ('127.0.0.1', '::1', 'localhost'):
                    return self._error_response('Internal endpoint only', 403, request)

            alerts = self.db.get_active_prompt_alerts()

            # Include reference states for each alert
            data = []
            for alert in alerts:
                alert_dict = alert.to_api_dict()
                ref = self.db.get_reference_snapshot(alert.id)
                if ref:
                    alert_dict['referenceState'] = ref.to_api_dict()
                data.append(alert_dict)

            return self._json_response({
                'success': True,
                'data': data,
                'count': len(data)
            }, request=request)
        except Exception as e:
            self.logger.error(f"list_prompt_alerts_internal error: {e}")
            return self._error_response(str(e), 500, request)

    # ==================== Algo Alerts ====================

    async def _publish_algo_alerts_sync(self, action: str, alert_id: str):
        """Publish sync event for algo alerts."""
        try:
            r = self.buses.get('market') or self.buses.get('intel')
            if r:
                await r.publish('algo_alerts:sync', json.dumps({
                    'action': action,
                    'alert_id': alert_id,
                }))
        except Exception as e:
            self.logger.warn(f"Failed to publish algo_alerts:sync: {e}")

    async def list_algo_alerts(self, request: web.Request) -> web.Response:
        """GET /api/algo-alerts - List algo alerts for the current user."""
        try:
            user = await self.auth.get_request_user(request)
            user_id = user['id'] if user else None
            if not user_id:
                return self._error_response('Unauthorized', 401, request)

            status_filter = request.query.get('status')
            alerts = self.db.list_algo_alerts(user_id, status_filter)
            return self._json_response({
                'success': True,
                'data': alerts,
                'count': len(alerts)
            }, request=request)
        except Exception as e:
            self.logger.error(f"list_algo_alerts error: {e}")
            return self._error_response(str(e), 500, request)

    async def create_algo_alert(self, request: web.Request) -> web.Response:
        """POST /api/algo-alerts - Create a new algo alert."""
        try:
            user = await self.auth.get_request_user(request)
            user_id = user['id'] if user else None
            if not user_id:
                return self._error_response('Unauthorized', 401, request)

            data = await request.json()
            import uuid
            alert_id = str(uuid.uuid4())

            name = data.get('name', 'Untitled Algo Alert')
            mode = data.get('mode', 'entry')
            if mode not in ('entry', 'management'):
                return self._error_response('Invalid mode — must be entry or management', 400, request)

            filters = data.get('filters', [])
            if not filters:
                return self._error_response('At least one filter is required', 400, request)

            # Assign IDs to filters if missing
            for f in filters:
                if not f.get('id'):
                    f['id'] = str(uuid.uuid4())

            entry_constraints = data.get('entryConstraints')
            position_id = data.get('positionId')
            prompt_override = data.get('promptOverride')

            alert = self.db.create_algo_alert(
                alert_id=alert_id,
                user_id=user_id,
                name=name,
                mode=mode,
                filters=json.dumps(filters),
                entry_constraints=json.dumps(entry_constraints) if entry_constraints else None,
                position_id=position_id,
                prompt_override=prompt_override,
            )

            await self._publish_algo_alerts_sync('create', alert_id)

            return self._json_response({
                'success': True,
                'data': alert,
            }, status=201, request=request)
        except Exception as e:
            self.logger.error(f"create_algo_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def update_algo_alert_route(self, request: web.Request) -> web.Response:
        """PUT /api/algo-alerts/:id - Update an algo alert."""
        try:
            user = await self.auth.get_request_user(request)
            user_id = user['id'] if user else None
            if not user_id:
                return self._error_response('Unauthorized', 401, request)

            alert_id = request.match_info['id']
            data = await request.json()

            updates = {}
            if 'name' in data:
                updates['name'] = data['name']
            if 'status' in data:
                if data['status'] not in ('active', 'paused', 'frozen', 'archived'):
                    return self._error_response('Invalid status', 400, request)
                updates['status'] = data['status']
            if 'filters' in data:
                import uuid as _uuid
                for f in data['filters']:
                    if not f.get('id'):
                        f['id'] = str(_uuid.uuid4())
                updates['filters'] = json.dumps(data['filters'])
            if 'entryConstraints' in data:
                updates['entry_constraints'] = json.dumps(data['entryConstraints'])
            if 'positionId' in data:
                updates['position_id'] = data['positionId']
            if 'promptOverride' in data:
                updates['prompt_override'] = data['promptOverride']

            alert = self.db.update_algo_alert(alert_id, user_id, updates)
            if not alert:
                return self._error_response('Algo alert not found', 404, request)

            await self._publish_algo_alerts_sync('update', alert_id)

            return self._json_response({
                'success': True,
                'data': alert,
            }, request=request)
        except Exception as e:
            self.logger.error(f"update_algo_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def delete_algo_alert_route(self, request: web.Request) -> web.Response:
        """DELETE /api/algo-alerts/:id - Delete an algo alert."""
        try:
            user = await self.auth.get_request_user(request)
            user_id = user['id'] if user else None
            if not user_id:
                return self._error_response('Unauthorized', 401, request)

            alert_id = request.match_info['id']
            deleted = self.db.delete_algo_alert(alert_id, user_id)
            if not deleted:
                return self._error_response('Algo alert not found', 404, request)

            await self._publish_algo_alerts_sync('delete', alert_id)

            return self._json_response({
                'success': True,
                'message': 'Algo alert deleted'
            }, request=request)
        except Exception as e:
            self.logger.error(f"delete_algo_alert error: {e}")
            return self._error_response(str(e), 500, request)

    async def list_algo_proposals(self, request: web.Request) -> web.Response:
        """GET /api/algo-proposals - List proposals for the current user."""
        try:
            user = await self.auth.get_request_user(request)
            user_id = user['id'] if user else None
            if not user_id:
                return self._error_response('Unauthorized', 401, request)

            algo_alert_id = request.query.get('algoAlertId')
            status_filter = request.query.get('status')

            proposals = self.db.list_algo_proposals(user_id, algo_alert_id, status_filter)
            return self._json_response({
                'success': True,
                'data': proposals,
                'count': len(proposals)
            }, request=request)
        except Exception as e:
            self.logger.error(f"list_algo_proposals error: {e}")
            return self._error_response(str(e), 500, request)

    async def approve_algo_proposal(self, request: web.Request) -> web.Response:
        """POST /api/algo-proposals/:id/approve - Approve a proposal."""
        try:
            user = await self.auth.get_request_user(request)
            user_id = user['id'] if user else None
            if not user_id:
                return self._error_response('Unauthorized', 401, request)

            proposal_id = request.match_info['id']
            proposal = self.db.resolve_algo_proposal(proposal_id, user_id, 'approved')
            if not proposal:
                return self._error_response('Proposal not found', 404, request)

            # Clear pending proposal in evaluator (via sync)
            algo_alert_id = proposal.get('algoAlertId')
            if algo_alert_id:
                await self._publish_algo_alerts_sync('proposal_resolved', algo_alert_id)

            return self._json_response({
                'success': True,
                'data': proposal,
            }, request=request)
        except Exception as e:
            self.logger.error(f"approve_algo_proposal error: {e}")
            return self._error_response(str(e), 500, request)

    async def reject_algo_proposal(self, request: web.Request) -> web.Response:
        """POST /api/algo-proposals/:id/reject - Reject a proposal."""
        try:
            user = await self.auth.get_request_user(request)
            user_id = user['id'] if user else None
            if not user_id:
                return self._error_response('Unauthorized', 401, request)

            proposal_id = request.match_info['id']
            proposal = self.db.resolve_algo_proposal(proposal_id, user_id, 'rejected')
            if not proposal:
                return self._error_response('Proposal not found', 404, request)

            algo_alert_id = proposal.get('algoAlertId')
            if algo_alert_id:
                await self._publish_algo_alerts_sync('proposal_resolved', algo_alert_id)

            return self._json_response({
                'success': True,
                'data': proposal,
            }, request=request)
        except Exception as e:
            self.logger.error(f"reject_algo_proposal error: {e}")
            return self._error_response(str(e), 500, request)

    async def list_algo_alerts_internal(self, request: web.Request) -> web.Response:
        """GET /api/internal/algo-alerts - List all active algo alerts (for Copilot)."""
        try:
            alerts = self.db.get_active_algo_alerts()
            return self._json_response({
                'success': True,
                'data': alerts,
                'count': len(alerts)
            }, request=request)
        except Exception as e:
            self.logger.error(f"list_algo_alerts_internal error: {e}")
            return self._error_response(str(e), 500, request)

    async def update_algo_alert_status_internal(self, request: web.Request) -> web.Response:
        """PUT /api/internal/algo-alerts/:id/status - Update status (from Copilot)."""
        try:
            alert_id = request.match_info['id']
            data = await request.json()
            status = data.get('status')
            frozen_reason = data.get('frozen_reason')

            updated = self.db.update_algo_alert_status_internal(alert_id, status, frozen_reason)
            return self._json_response({
                'success': updated,
            }, request=request)
        except Exception as e:
            self.logger.error(f"update_algo_alert_status_internal error: {e}")
            return self._error_response(str(e), 500, request)

    async def create_algo_proposal_internal(self, request: web.Request) -> web.Response:
        """POST /api/internal/algo-proposals - Create proposal (from Copilot)."""
        try:
            data = await request.json()
            proposal = self.db.create_algo_proposal(data)
            return self._json_response({
                'success': True,
                'data': proposal,
            }, status=201, request=request)
        except Exception as e:
            self.logger.error(f"create_algo_proposal_internal error: {e}")
            return self._error_response(str(e), 500, request)

    # ==================== Trade Idea Tracking (Feedback Loop) ====================

    async def create_tracked_idea(self, request: web.Request) -> web.Response:
        """POST /api/internal/tracked-ideas - Create a tracked idea record (called on settlement)."""
        try:
            data = await request.json()

            # Build TrackedIdea from data
            idea = TrackedIdea(
                id=data['id'],
                symbol=data['symbol'],
                entry_rank=data['entry_rank'],
                entry_time=data['entry_time'],
                entry_ts=data['entry_ts'],
                entry_spot=data['entry_spot'],
                entry_vix=data['entry_vix'],
                entry_regime=data['entry_regime'],
                # Trade params (required)
                strategy=data['strategy'],
                side=data['side'],
                strike=data['strike'],
                width=data['width'],
                dte=data['dte'],
                debit=data['debit'],
                max_profit_theoretical=data['max_profit_theoretical'],
                # Time context (optional)
                entry_hour=data.get('entry_hour'),
                entry_day_of_week=data.get('entry_day_of_week'),
                # GEX context (optional)
                entry_gex_flip=data.get('entry_gex_flip'),
                entry_gex_call_wall=data.get('entry_gex_call_wall'),
                entry_gex_put_wall=data.get('entry_gex_put_wall'),
                # Trade params (optional)
                r2r_predicted=data.get('r2r_predicted'),
                campaign=data.get('campaign'),
                max_pnl=data['max_pnl'],
                max_pnl_time=data.get('max_pnl_time'),
                max_pnl_spot=data.get('max_pnl_spot'),
                max_pnl_dte=data.get('max_pnl_dte'),
                settlement_time=data['settlement_time'],
                settlement_spot=data['settlement_spot'],
                final_pnl=data['final_pnl'],
                is_winner=data['is_winner'],
                pnl_captured_pct=data.get('pnl_captured_pct'),
                r2r_achieved=data.get('r2r_achieved'),
                score_total=data.get('score_total'),
                score_regime=data.get('score_regime'),
                score_r2r=data.get('score_r2r'),
                score_convexity=data.get('score_convexity'),
                score_campaign=data.get('score_campaign'),
                score_decay=data.get('score_decay'),
                score_edge=data.get('score_edge'),
                params_version=data.get('params_version'),
                edge_cases=json.dumps(data.get('edge_cases', [])) if data.get('edge_cases') else None,
            )

            self.db.create_tracked_idea(idea)
            self.logger.info(f"Tracked idea created: {idea.id} rank={idea.entry_rank} winner={idea.is_winner}")

            return self._json_response({
                'success': True,
                'data': idea.to_api_dict()
            }, request=request)
        except Exception as e:
            self.logger.error(f"create_tracked_idea error: {e}")
            return self._error_response(str(e), 500, request)

    async def list_tracked_ideas(self, request: web.Request) -> web.Response:
        """GET /api/internal/tracked-ideas - List tracked ideas with filters."""
        try:
            params = request.query
            ideas = self.db.list_tracked_ideas(
                limit=int(params.get('limit', 100)),
                offset=int(params.get('offset', 0)),
                regime=params.get('regime'),
                strategy=params.get('strategy'),
                rank=int(params['rank']) if params.get('rank') else None,
                is_winner=params.get('is_winner') == 'true' if params.get('is_winner') else None,
                params_version=int(params['params_version']) if params.get('params_version') else None,
                start_date=params.get('start_date'),
                end_date=params.get('end_date'),
            )

            return self._json_response({
                'success': True,
                'data': [i.to_api_dict() for i in ideas],
                'count': len(ideas)
            }, request=request)
        except Exception as e:
            self.logger.error(f"list_tracked_ideas error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_tracking_analytics(self, request: web.Request) -> web.Response:
        """GET /api/internal/tracked-ideas/analytics - Get aggregated analytics."""
        try:
            params = request.query
            analytics = self.db.get_tracking_analytics(
                params_version=int(params['params_version']) if params.get('params_version') else None,
                start_date=params.get('start_date'),
                end_date=params.get('end_date'),
            )

            return self._json_response({
                'success': True,
                'data': analytics
            }, request=request)
        except Exception as e:
            self.logger.error(f"get_tracking_analytics error: {e}")
            return self._error_response(str(e), 500, request)

    async def list_selector_params(self, request: web.Request) -> web.Response:
        """GET /api/internal/selector-params - List all parameter versions."""
        try:
            include_retired = request.query.get('include_retired') == 'true'
            params_list = self.db.list_params(include_retired=include_retired)

            return self._json_response({
                'success': True,
                'data': [p.to_api_dict() for p in params_list],
                'count': len(params_list)
            }, request=request)
        except Exception as e:
            self.logger.error(f"list_selector_params error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_active_params(self, request: web.Request) -> web.Response:
        """GET /api/internal/selector-params/active - Get currently active parameters."""
        try:
            params = self.db.get_active_params()
            if not params:
                return self._error_response("No active parameters found", 404, request)

            return self._json_response({
                'success': True,
                'data': params.to_api_dict()
            }, request=request)
        except Exception as e:
            self.logger.error(f"get_active_params error: {e}")
            return self._error_response(str(e), 500, request)

    async def create_selector_params(self, request: web.Request) -> web.Response:
        """POST /api/internal/selector-params - Create new parameter version."""
        try:
            data = await request.json()

            params = SelectorParams(
                name=data.get('name'),
                description=data.get('description'),
                weights=json.dumps(data.get('weights', {})),
                regime_thresholds=json.dumps(data.get('regime_thresholds', {})) if data.get('regime_thresholds') else None,
                status=data.get('status', 'draft'),
            )

            created = self.db.create_params(params)
            self.logger.info(f"Selector params created: version={created.version} name={created.name}")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, request=request)
        except Exception as e:
            self.logger.error(f"create_selector_params error: {e}")
            return self._error_response(str(e), 500, request)

    async def activate_selector_params(self, request: web.Request) -> web.Response:
        """POST /api/internal/selector-params/{version}/activate - Activate a parameter version."""
        try:
            version = int(request.match_info['version'])
            success = self.db.activate_params(version)

            if not success:
                return self._error_response(f"Parameter version {version} not found", 404, request)

            self.logger.info(f"Selector params activated: version={version}")

            return self._json_response({
                'success': True,
                'message': f'Parameter version {version} activated'
            }, request=request)
        except Exception as e:
            self.logger.error(f"activate_selector_params error: {e}")
            return self._error_response(str(e), 500, request)

    # ==================== Legacy Endpoints (for backwards compatibility) ====================

    async def legacy_list_trades(self, request: web.Request) -> web.Response:
        """GET /api/trades - Legacy: List trades from first active log."""
        try:
            # Get user from request for proper filtering
            user = await self.auth.get_request_user(request)
            user_id = user['id'] if user else None

            logs = self.db.list_logs(user_id=user_id)
            if not logs:
                return self._json_response({
                    'success': True,
                    'data': [],
                    'count': 0
                })

            # Use first log for legacy endpoint
            log_id = logs[0].id
            params = request.query

            trades = self.db.list_trades(
                log_id=log_id,
                status=params.get('status'),
                symbol=params.get('symbol'),
                strategy=params.get('strategy'),
                from_date=params.get('from'),
                to_date=params.get('to'),
                limit=int(params.get('limit', 100)),
                offset=int(params.get('offset', 0))
            )

            return self._json_response({
                'success': True,
                'data': [t.to_api_dict() for t in trades],
                'count': len(trades)
            })
        except Exception as e:
            self.logger.error(f"legacy_list_trades error: {e}")
            return self._error_response(str(e), 500)

    async def legacy_create_trade(self, request: web.Request) -> web.Response:
        """POST /api/trades - Legacy: Create trade in first active log."""
        try:
            # Get user from request for proper filtering
            user = await self.auth.get_request_user(request)
            user_id = user['id'] if user else None

            logs = self.db.list_logs(user_id=user_id)
            if not logs:
                # Create default log for this user
                default_log = TradeLog(
                    id=TradeLog.new_id(),
                    name='Default Log',
                    user_id=user_id,
                    starting_capital=2500000,  # $25,000
                    intent='Auto-created default log'
                )
                self.db.create_log(default_log)
                logs = [default_log]

            log_id = logs[0].id
            body = await request.json()

            # Verify log exists
            log = self.db.get_log(log_id)
            if not log:
                return self._error_response('Trade log not found', 404)

            # Set defaults
            entry_time = body.get('entry_time') or datetime.utcnow().isoformat()

            # Convert prices to cents if needed
            entry_price = body['entry_price']
            if isinstance(entry_price, float) or entry_price < 1000:
                entry_price = int(entry_price * 100)

            planned_risk = body.get('planned_risk')
            if planned_risk and (isinstance(planned_risk, float) or planned_risk < 1000):
                planned_risk = int(planned_risk * 100)

            max_profit = body.get('max_profit')
            if max_profit and (isinstance(max_profit, float) or max_profit < 1000):
                max_profit = int(max_profit * 100)

            max_loss = body.get('max_loss')
            if max_loss and (isinstance(max_loss, float) or max_loss < 1000):
                max_loss = int(max_loss * 100)

            trade = Trade(
                id=Trade.new_id(),
                log_id=log_id,
                symbol=body.get('symbol', 'SPX'),
                underlying=body.get('underlying', 'I:SPX'),
                strategy=body['strategy'],
                side=body['side'],
                strike=float(body['strike']),
                width=body.get('width'),
                dte=body.get('dte'),
                quantity=int(body.get('quantity', 1)),
                entry_time=entry_time,
                entry_price=entry_price,
                entry_spot=body.get('entry_spot'),
                entry_iv=body.get('entry_iv'),
                planned_risk=planned_risk or max_loss,
                max_profit=max_profit,
                max_loss=max_loss,
                notes=body.get('notes'),
                tags=body.get('tags', []),
                source=body.get('source', 'manual'),
                playbook_id=body.get('playbook_id')
            )

            created = self.db.create_trade(trade)
            self.logger.info(f"[Legacy] Created trade: {created.strategy} {created.strike} @ ${entry_price/100:.2f}", emoji="📝")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201)

        except KeyError as e:
            return self._error_response(f'Missing required field: {e}')
        except Exception as e:
            self.logger.error(f"legacy_create_trade error: {e}")
            return self._error_response(str(e), 500)

    async def legacy_analytics(self, request: web.Request) -> web.Response:
        """GET /api/analytics - Legacy: Analytics from first active log."""
        try:
            logs = self.db.list_logs()
            if not logs:
                return self._json_response({
                    'success': True,
                    'data': {'summary': {}}
                })

            analytics = self.analytics.get_full_analytics(logs[0].id)

            return self._json_response({
                'success': True,
                'data': {
                    'summary': analytics.to_api_dict() if analytics else {}
                }
            })
        except Exception as e:
            self.logger.error(f"legacy_analytics error: {e}")
            return self._error_response(str(e), 500)

    async def legacy_equity(self, request: web.Request) -> web.Response:
        """GET /api/analytics/equity - Legacy: Equity from first active log."""
        try:
            logs = self.db.list_logs()
            if not logs:
                return self._json_response({
                    'success': True,
                    'data': {'equity': []}
                })

            equity_points = self.analytics.get_equity_curve(logs[0].id)

            return self._json_response({
                'success': True,
                'data': {
                    'equity': [p.to_dict() for p in equity_points]
                }
            })
        except Exception as e:
            self.logger.error(f"legacy_equity error: {e}")
            return self._error_response(str(e), 500)

    # ==================== Leaderboard API ====================

    def _get_period_key(self, period_type: str) -> str:
        """Get the current period key for a period type."""
        now = datetime.utcnow()
        if period_type == 'weekly':
            # ISO week format: 2026-W06
            return now.strftime('%G-W%V')
        elif period_type == 'monthly':
            # Month format: 2026-02
            return now.strftime('%Y-%m')
        else:
            return 'all'

    def _get_period_boundaries(self, period_type: str, period_key: str) -> tuple:
        """Get start and end dates for a period."""
        if period_type == 'weekly':
            # Parse ISO week: 2026-W06
            year, week = period_key.split('-W')
            # Get first day of ISO week (Monday)
            first_day = datetime.strptime(f'{year}-W{week}-1', '%G-W%V-%u')
            start = first_day.strftime('%Y-%m-%d')
            end = (first_day + timedelta(days=7)).strftime('%Y-%m-%d')
        elif period_type == 'monthly':
            # Parse month: 2026-02
            year, month = period_key.split('-')
            start = f'{year}-{month}-01'
            # Get first day of next month
            if int(month) == 12:
                end = f'{int(year)+1}-01-01'
            else:
                end = f'{year}-{int(month)+1:02d}-01'
        else:
            # All time: use a very old start date
            start = '2000-01-01'
            end = '2100-01-01'
        return start, end

    def _calculate_activity_score(self, trades: int, entries: int, tags: int) -> float:
        """Calculate activity score (0-50)."""
        # Points with caps
        trade_pts = min(trades * 5, 100)
        entry_pts = min(entries * 10, 70)
        tag_pts = min(tags * 2, 30)

        # Raw total capped at 200, scaled to 0-50
        raw = min(trade_pts + entry_pts + tag_pts, 200)
        return raw / 4.0

    def _calculate_performance_score(
        self, win_rate, avg_r, pnl, pnl_percentile: float, closed_trades: int, min_trades: int
    ) -> float:
        """Calculate performance score (0-50)."""
        if closed_trades < min_trades:
            return 0.0

        # Convert to float to handle Decimal from DB
        win_rate = float(win_rate) if win_rate else 0.0
        avg_r = float(avg_r) if avg_r else 0.0

        has_r_data = avg_r != 0.0

        if has_r_data:
            # Full scoring: win_rate(20) + r_multiple(17.5) + pnl_percentile(12.5) = 50
            win_rate_pts = min(win_rate / 100 * 20, 20)
            r_clamped = max(-2.0, min(3.0, avg_r))
            r_normalized = (r_clamped + 2) / 5  # 0 to 1
            r_pts = r_normalized * 17.5
            pnl_pts = pnl_percentile / 100 * 12.5
        else:
            # No R-multiple data — redistribute to win_rate(30) + pnl_percentile(20) = 50
            win_rate_pts = min(win_rate / 100 * 30, 30)
            r_pts = 0.0
            pnl_pts = pnl_percentile / 100 * 20

        return win_rate_pts + r_pts + pnl_pts

    def _get_min_trades_for_period(self, period_type: str) -> int:
        """Get minimum trades required for performance score."""
        if period_type == 'weekly':
            return 1
        elif period_type == 'monthly':
            return 3
        else:
            return 5

    async def calculate_leaderboard(self, period_type: str) -> int:
        """Calculate leaderboard scores for all users in a period. Returns count of users scored."""
        period_key = self._get_period_key(period_type)
        start_date, end_date = self._get_period_boundaries(period_type, period_key)
        min_trades = self._get_min_trades_for_period(period_type)

        # Get all users with activity
        user_ids = self.db.get_all_user_ids_with_activity(start_date, end_date)

        # Collect all P&L values for percentile calculation
        all_pnl = []
        user_data = []

        for user_id in user_ids:
            activity = self.db.get_user_activity_metrics(user_id, start_date, end_date)
            performance = self.db.get_user_performance_metrics(user_id, start_date, end_date)

            user_data.append({
                'user_id': user_id,
                'activity': activity,
                'performance': performance,
            })
            all_pnl.append(performance['total_pnl'])

        # Calculate P&L percentiles
        sorted_pnl = sorted(all_pnl)

        for data in user_data:
            pnl = data['performance']['total_pnl']
            if len(sorted_pnl) > 1:
                # Find percentile rank using bisect for fair tie handling
                from bisect import bisect_left, bisect_right
                lo = bisect_left(sorted_pnl, pnl)
                hi = bisect_right(sorted_pnl, pnl)
                rank = (lo + hi - 1) / 2  # midpoint rank for ties
                pnl_percentile = (rank / (len(sorted_pnl) - 1)) * 100
            else:
                pnl_percentile = 50.0

            activity_score = self._calculate_activity_score(
                data['activity']['trades_logged'],
                data['activity']['journal_entries'],
                data['activity']['tags_used'],
            )

            performance_score = self._calculate_performance_score(
                data['performance']['win_rate'],
                data['performance']['avg_r_multiple'],
                data['performance']['total_pnl'],
                pnl_percentile,
                data['performance']['closed_trades'],
                min_trades,
            )

            total_score = activity_score + performance_score

            # Upsert score
            self.db.upsert_leaderboard_score(
                user_id=data['user_id'],
                period_type=period_type,
                period_key=period_key,
                trades_logged=data['activity']['trades_logged'],
                journal_entries=data['activity']['journal_entries'],
                tags_used=data['activity']['tags_used'],
                total_pnl=data['performance']['total_pnl'],
                win_rate=data['performance']['win_rate'],
                avg_r_multiple=data['performance']['avg_r_multiple'],
                closed_trades=data['performance']['closed_trades'],
                activity_score=activity_score,
                performance_score=performance_score,
                total_score=total_score,
            )

        # Update ranks
        self.db.update_leaderboard_ranks(period_type, period_key)

        return len(user_data)

    async def get_leaderboard(self, request: web.Request) -> web.Response:
        """GET /api/leaderboard - Get leaderboard rankings."""
        try:
            params = request.query
            period_type = params.get('period', 'weekly')
            if period_type not in ('weekly', 'monthly', 'all_time'):
                return self._error_response('Invalid period type', 400, request)

            limit = min(int(params.get('limit', 20)), 100)
            offset = int(params.get('offset', 0))
            recalculate = params.get('recalculate', 'false').lower() == 'true'

            period_key = self._get_period_key(period_type)

            # Optionally recalculate (admin/trigger)
            if recalculate:
                user = await self.auth.get_request_user(request)
                if user:  # Only allow recalculation for authenticated users
                    await self.calculate_leaderboard(period_type)

            # Get rankings
            rankings = self.db.get_leaderboard(period_type, period_key, limit, offset)
            total_participants = self.db.get_leaderboard_participant_count(period_type, period_key)

            # Get current user's rank if authenticated
            current_user_rank = None
            user = await self.auth.get_request_user(request)
            if user:
                current_user_rank = self.db.get_user_leaderboard_score(user['id'], period_type, period_key)

            return self._json_response({
                'success': True,
                'data': {
                    'rankings': rankings,
                    'currentUserRank': current_user_rank,
                    'totalParticipants': total_participants,
                    'periodType': period_type,
                    'periodKey': period_key,
                }
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_leaderboard error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_my_leaderboard(self, request: web.Request) -> web.Response:
        """GET /api/leaderboard/me - Get current user's leaderboard stats for all periods."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            result = {}
            for period_type in ('weekly', 'monthly', 'all_time'):
                period_key = self._get_period_key(period_type)
                score = self.db.get_user_leaderboard_score(user['id'], period_type, period_key)
                total = self.db.get_leaderboard_participant_count(period_type, period_key)
                result[period_type] = {
                    'score': score,
                    'totalParticipants': total,
                    'periodKey': period_key,
                }

            return self._json_response({
                'success': True,
                'data': result
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_my_leaderboard error: {e}")
            return self._error_response(str(e), 500, request)

    async def trigger_leaderboard_calculation(self, request: web.Request) -> web.Response:
        """POST /api/internal/leaderboard/calculate - Trigger leaderboard calculation (internal)."""
        try:
            params = request.query
            period_type = params.get('period', 'weekly')
            if period_type not in ('weekly', 'monthly', 'all_time'):
                return self._error_response('Invalid period type', 400, request)

            count = await self.calculate_leaderboard(period_type)
            self.logger.info(f"Calculated leaderboard for {period_type}: {count} users", emoji="🏆")

            return self._json_response({
                'success': True,
                'message': f'Calculated scores for {count} users',
                'period': period_type,
            }, request=request)

        except Exception as e:
            self.logger.error(f"trigger_leaderboard_calculation error: {e}")
            return self._error_response(str(e), 500, request)

    async def _leaderboard_scheduler(self):
        """Background task: recalculate all leaderboard periods every hour."""
        # Initial calculation on startup (wait for DB to be ready)
        await asyncio.sleep(15)
        self.logger.info("Leaderboard scheduler: running initial calculation", emoji="🏆")
        for period in ('weekly', 'monthly', 'all_time'):
            try:
                count = await self.calculate_leaderboard(period)
                self.logger.info(f"Leaderboard [{period}]: {count} users scored", emoji="🏆")
            except Exception as e:
                self.logger.error(f"Leaderboard [{period}] initial calc error: {e}")

        # Then recalculate every hour
        while True:
            await asyncio.sleep(3600)
            for period in ('weekly', 'monthly', 'all_time'):
                try:
                    count = await self.calculate_leaderboard(period)
                    self.logger.info(f"Leaderboard [{period}]: {count} users scored", emoji="🏆")
                except Exception as e:
                    self.logger.error(f"Leaderboard [{period}] scheduler error: {e}")

    # ==================== Risk Graph Service ====================

    async def _publish_risk_graph_event(self, user_id: int, event_type: str, data: dict):
        """Publish a risk graph event to Redis for SSE broadcast."""
        try:
            if not self._redis:
                self._redis = redis.from_url(self._redis_url)
            channel = f"risk_graph_updates:{user_id}"
            event = json.dumps({
                'type': event_type,
                'data': data,
                'ts': datetime.utcnow().isoformat()
            })
            await self._redis.publish(channel, event)
        except Exception as e:
            self.logger.error(f"Failed to publish risk graph event: {e}")

    async def list_risk_graph_strategies(self, request: web.Request) -> web.Response:
        """GET /api/risk-graph/strategies - List risk graph strategies for user."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            include_inactive = request.query.get('include_inactive', 'false') == 'true'
            strategies = self.db.list_risk_graph_strategies(user['id'], include_inactive)

            return self._json_response({
                'success': True,
                'data': [s.to_api_dict() for s in strategies],
                'count': len(strategies)
            }, request=request)

        except Exception as e:
            self.logger.error(f"list_risk_graph_strategies error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_risk_graph_strategy(self, request: web.Request) -> web.Response:
        """GET /api/risk-graph/strategies/:id - Get a single strategy."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            strategy_id = request.match_info['id']
            strategy = self.db.get_risk_graph_strategy(strategy_id, user['id'])

            if not strategy:
                return self._error_response('Strategy not found', 404, request)

            return self._json_response({
                'success': True,
                'data': strategy.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_risk_graph_strategy error: {e}")
            return self._error_response(str(e), 500, request)

    async def create_risk_graph_strategy(self, request: web.Request) -> web.Response:
        """POST /api/risk-graph/strategies - Create a new strategy."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            body = await request.json()

            from .models_v2 import RiskGraphStrategy

            strategy = RiskGraphStrategy(
                id=RiskGraphStrategy.new_id(),
                user_id=user['id'],
                symbol=body.get('symbol', 'SPX'),
                underlying=body.get('underlying', 'I:SPX'),
                strategy=body['strategy'],
                side=body['side'],
                strike=float(body['strike']),
                width=body.get('width'),
                dte=int(body['dte']),
                expiration=body['expiration'],
                debit=float(body['debit']) if body.get('debit') is not None else None,
                visible=body.get('visible', True),
                sort_order=body.get('sortOrder', 0),
                color=body.get('color'),
                label=body.get('label'),
                added_at=body.get('addedAt', int(datetime.utcnow().timestamp() * 1000)),
            )

            created = self.db.create_risk_graph_strategy(strategy)
            self.logger.info(f"Created risk graph strategy: {created.strategy} {created.strike}", emoji="📊")

            # Publish event for real-time sync
            await self._publish_risk_graph_event(user['id'], 'strategy_added', created.to_api_dict())

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201, request)

        except KeyError as e:
            return self._error_response(f'Missing required field: {e}', 400, request)
        except Exception as e:
            self.logger.error(f"create_risk_graph_strategy error: {e}")
            return self._error_response(str(e), 500, request)

    async def update_risk_graph_strategy(self, request: web.Request) -> web.Response:
        """PATCH /api/risk-graph/strategies/:id - Update a strategy."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            strategy_id = request.match_info['id']
            body = await request.json()

            # Convert camelCase to snake_case for updates
            updates = {}
            field_map = {
                'debit': 'debit',
                'visible': 'visible',
                'sortOrder': 'sort_order',
                'color': 'color',
                'label': 'label',
            }
            for api_key, db_key in field_map.items():
                if api_key in body:
                    updates[db_key] = body[api_key]

            change_reason = body.get('changeReason')

            updated = self.db.update_risk_graph_strategy(
                strategy_id, updates, user['id'], change_reason
            )

            if not updated:
                return self._error_response('Strategy not found', 404, request)

            # Publish event for real-time sync
            await self._publish_risk_graph_event(user['id'], 'strategy_updated', updated.to_api_dict())

            return self._json_response({
                'success': True,
                'data': updated.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"update_risk_graph_strategy error: {e}")
            return self._error_response(str(e), 500, request)

    async def delete_risk_graph_strategy(self, request: web.Request) -> web.Response:
        """DELETE /api/risk-graph/strategies/:id - Delete (soft) a strategy."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            strategy_id = request.match_info['id']
            hard_delete = request.query.get('hard', 'false') == 'true'

            deleted = self.db.delete_risk_graph_strategy(strategy_id, user['id'], hard_delete)

            if not deleted:
                return self._error_response('Strategy not found', 404, request)

            # Publish event for real-time sync
            await self._publish_risk_graph_event(user['id'], 'strategy_removed', {'id': strategy_id})

            return self._json_response({
                'success': True,
                'message': 'Strategy deleted'
            }, request=request)

        except Exception as e:
            self.logger.error(f"delete_risk_graph_strategy error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_risk_graph_strategy_versions(self, request: web.Request) -> web.Response:
        """GET /api/risk-graph/strategies/:id/versions - Get version history."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            strategy_id = request.match_info['id']

            # Verify ownership
            strategy = self.db.get_risk_graph_strategy(strategy_id, user['id'])
            if not strategy:
                return self._error_response('Strategy not found', 404, request)

            versions = self.db.get_risk_graph_strategy_versions(strategy_id)

            return self._json_response({
                'success': True,
                'data': [v.to_api_dict() for v in versions],
                'count': len(versions)
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_risk_graph_strategy_versions error: {e}")
            return self._error_response(str(e), 500, request)

    async def import_risk_graph_strategies(self, request: web.Request) -> web.Response:
        """POST /api/risk-graph/strategies/import - Bulk import strategies."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            body = await request.json()
            strategies_data = body.get('strategies', [])

            from .models_v2 import RiskGraphStrategy

            imported = []
            for data in strategies_data:
                strategy = RiskGraphStrategy(
                    id=RiskGraphStrategy.new_id(),
                    user_id=user['id'],
                    symbol=data.get('symbol', 'SPX'),
                    underlying=data.get('underlying', 'I:SPX'),
                    strategy=data['strategy'],
                    side=data['side'],
                    strike=float(data['strike']),
                    width=data.get('width'),
                    dte=int(data['dte']),
                    expiration=data['expiration'],
                    debit=float(data['debit']) if data.get('debit') is not None else None,
                    visible=data.get('visible', True),
                    sort_order=len(imported),
                    color=data.get('color'),
                    label=data.get('label'),
                    added_at=data.get('addedAt', int(datetime.utcnow().timestamp() * 1000)),
                )
                created = self.db.create_risk_graph_strategy(strategy)
                imported.append(created)

            self.logger.info(f"Imported {len(imported)} risk graph strategies", emoji="📊")

            return self._json_response({
                'success': True,
                'data': [s.to_api_dict() for s in imported],
                'count': len(imported)
            }, 201, request)

        except Exception as e:
            self.logger.error(f"import_risk_graph_strategies error: {e}")
            return self._error_response(str(e), 500, request)

    async def export_risk_graph_strategies(self, request: web.Request) -> web.Response:
        """GET /api/risk-graph/strategies/export - Export all strategies."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            strategies = self.db.list_risk_graph_strategies(user['id'], include_inactive=False)

            return self._json_response({
                'success': True,
                'data': {
                    'strategies': [s.to_api_dict() for s in strategies],
                    'exportedAt': datetime.utcnow().isoformat(),
                    'count': len(strategies)
                }
            }, request=request)

        except Exception as e:
            self.logger.error(f"export_risk_graph_strategies error: {e}")
            return self._error_response(str(e), 500, request)

    async def reorder_risk_graph_strategies(self, request: web.Request) -> web.Response:
        """POST /api/risk-graph/strategies/reorder - Update sort order."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            body = await request.json()
            strategy_order = body.get('order', [])

            if not strategy_order:
                return self._error_response('order array required', 400, request)

            self.db.reorder_risk_graph_strategies(user['id'], strategy_order)

            return self._json_response({
                'success': True,
                'message': 'Order updated'
            }, request=request)

        except Exception as e:
            self.logger.error(f"reorder_risk_graph_strategies error: {e}")
            return self._error_response(str(e), 500, request)

    # ==================== Risk Graph Templates ====================

    async def list_risk_graph_templates(self, request: web.Request) -> web.Response:
        """GET /api/risk-graph/templates - List templates for user."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            include_public = request.query.get('include_public', 'false') == 'true'
            templates = self.db.list_risk_graph_templates(user['id'], include_public)

            return self._json_response({
                'success': True,
                'data': [t.to_api_dict() for t in templates],
                'count': len(templates)
            }, request=request)

        except Exception as e:
            self.logger.error(f"list_risk_graph_templates error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_risk_graph_template(self, request: web.Request) -> web.Response:
        """GET /api/risk-graph/templates/:id - Get a single template."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            template_id = request.match_info['id']
            template = self.db.get_risk_graph_template(template_id, user['id'])

            if not template:
                return self._error_response('Template not found', 404, request)

            return self._json_response({
                'success': True,
                'data': template.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_risk_graph_template error: {e}")
            return self._error_response(str(e), 500, request)

    async def create_risk_graph_template(self, request: web.Request) -> web.Response:
        """POST /api/risk-graph/templates - Create a new template."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            body = await request.json()

            from .models_v2 import RiskGraphTemplate

            template = RiskGraphTemplate(
                id=RiskGraphTemplate.new_id(),
                user_id=user['id'],
                name=body['name'],
                description=body.get('description'),
                symbol=body.get('symbol', 'SPX'),
                strategy=body['strategy'],
                side=body['side'],
                strike_offset=int(body.get('strikeOffset', 0)),
                width=body.get('width'),
                dte_target=int(body['dteTarget']),
                debit_estimate=float(body['debitEstimate']) if body.get('debitEstimate') else None,
                is_public=body.get('isPublic', False),
            )

            created = self.db.create_risk_graph_template(template)
            self.logger.info(f"Created risk graph template: {created.name}", emoji="📋")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201, request)

        except KeyError as e:
            return self._error_response(f'Missing required field: {e}', 400, request)
        except Exception as e:
            self.logger.error(f"create_risk_graph_template error: {e}")
            return self._error_response(str(e), 500, request)

    async def update_risk_graph_template(self, request: web.Request) -> web.Response:
        """PATCH /api/risk-graph/templates/:id - Update a template."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            template_id = request.match_info['id']
            body = await request.json()

            # Convert camelCase to snake_case
            updates = {}
            field_map = {
                'name': 'name',
                'description': 'description',
                'symbol': 'symbol',
                'strategy': 'strategy',
                'side': 'side',
                'strikeOffset': 'strike_offset',
                'width': 'width',
                'dteTarget': 'dte_target',
                'debitEstimate': 'debit_estimate',
                'isPublic': 'is_public',
            }
            for api_key, db_key in field_map.items():
                if api_key in body:
                    updates[db_key] = body[api_key]

            updated = self.db.update_risk_graph_template(template_id, updates, user['id'])

            if not updated:
                return self._error_response('Template not found', 404, request)

            return self._json_response({
                'success': True,
                'data': updated.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"update_risk_graph_template error: {e}")
            return self._error_response(str(e), 500, request)

    async def delete_risk_graph_template(self, request: web.Request) -> web.Response:
        """DELETE /api/risk-graph/templates/:id - Delete a template."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            template_id = request.match_info['id']
            deleted = self.db.delete_risk_graph_template(template_id, user['id'])

            if not deleted:
                return self._error_response('Template not found', 404, request)

            return self._json_response({
                'success': True,
                'message': 'Template deleted'
            }, request=request)

        except Exception as e:
            self.logger.error(f"delete_risk_graph_template error: {e}")
            return self._error_response(str(e), 500, request)

    async def use_risk_graph_template(self, request: web.Request) -> web.Response:
        """POST /api/risk-graph/templates/:id/use - Create strategy from template."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            template_id = request.match_info['id']
            body = await request.json()

            template = self.db.get_risk_graph_template(template_id, user['id'])
            if not template:
                return self._error_response('Template not found', 404, request)

            # Get current spot price to calculate actual strike
            spot_price = float(body.get('spotPrice', 0))
            if not spot_price:
                return self._error_response('spotPrice required', 400, request)

            # Round to nearest strike increment (5 for SPX, 50 for NDX)
            increment = 50 if template.symbol == 'NDX' else 5
            strike = round((spot_price + template.strike_offset) / increment) * increment

            # Calculate expiration date
            from datetime import timedelta
            expiration = (datetime.utcnow() + timedelta(days=template.dte_target)).strftime('%Y-%m-%d')

            from .models_v2 import RiskGraphStrategy

            strategy = RiskGraphStrategy(
                id=RiskGraphStrategy.new_id(),
                user_id=user['id'],
                symbol=template.symbol,
                underlying=body.get('underlying', f'I:{template.symbol}'),
                strategy=template.strategy,
                side=template.side,
                strike=strike,
                width=template.width,
                dte=template.dte_target,
                expiration=expiration,
                debit=body.get('debit') or template.debit_estimate,
                visible=True,
                added_at=int(datetime.utcnow().timestamp() * 1000),
            )

            created = self.db.create_risk_graph_strategy(strategy)

            # Increment template use count
            self.db.increment_template_use_count(template_id)

            # Publish event for real-time sync
            await self._publish_risk_graph_event(user['id'], 'strategy_added', created.to_api_dict())

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, 201, request)

        except Exception as e:
            self.logger.error(f"use_risk_graph_template error: {e}")
            return self._error_response(str(e), 500, request)

    async def share_risk_graph_template(self, request: web.Request) -> web.Response:
        """POST /api/risk-graph/templates/:id/share - Generate share code."""
        try:
            user = await self.auth.get_request_user(request)
            if not user:
                return self._error_response('Authentication required', 401, request)

            template_id = request.match_info['id']
            share_code = self.db.generate_template_share_code(template_id, user['id'])

            if not share_code:
                return self._error_response('Template not found', 404, request)

            return self._json_response({
                'success': True,
                'data': {'shareCode': share_code}
            }, request=request)

        except Exception as e:
            self.logger.error(f"share_risk_graph_template error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_shared_risk_graph_template(self, request: web.Request) -> web.Response:
        """GET /api/risk-graph/templates/shared/:code - Get template by share code."""
        try:
            share_code = request.match_info['code']
            template = self.db.get_risk_graph_template_by_share_code(share_code)

            if not template:
                return self._error_response('Template not found', 404, request)

            return self._json_response({
                'success': True,
                'data': template.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_shared_risk_graph_template error: {e}")
            return self._error_response(str(e), 500, request)

    # ==================== Position API (TradeLog Service Layer) ====================

    async def _publish_trade_log_event(
        self,
        user_id: int,
        event_type: str,
        aggregate_id: str,
        aggregate_version: int,
        payload: dict,
        aggregate_type: str = 'position'
    ):
        """Publish a trade log event to Redis pub/sub for SSE broadcasting."""
        try:
            redis_conn = await self._get_redis()
            channel = f"trade_log_updates:{user_id}"

            # Create and store event for replay
            event = PositionEvent(
                event_id=PositionEvent.new_event_id(),
                event_type=event_type,
                aggregate_type=aggregate_type,
                aggregate_id=aggregate_id,
                aggregate_version=aggregate_version,
                user_id=user_id,
                payload=payload,
                occurred_at=datetime.utcnow().isoformat(),
            )
            stored_event = self.db.record_position_event(event)

            # Publish to Redis
            message = {
                'event_id': stored_event.event_id,
                'event_seq': stored_event.event_seq,
                'type': event_type,
                'aggregate_type': aggregate_type,
                'aggregate_id': aggregate_id,
                'aggregate_version': aggregate_version,
                'occurred_at': stored_event.occurred_at,
                'payload': payload,
            }
            await redis_conn.publish(channel, json.dumps(message))

        except Exception as e:
            self.logger.error(f"Failed to publish trade log event: {e}")

    async def _publish_lifecycle_event(
        self,
        user_id: int,
        event_type: str,
        payload: dict
    ):
        """Publish a log lifecycle event to Redis pub/sub for SSE broadcasting.

        Events follow the format:
        {
            "type": "log.lifecycle.archived",
            "timestamp": "2026-03-02T14:22:11Z",
            "user_id": 123,
            "payload": { "log_id": "uuid", ... }
        }
        """
        try:
            redis_conn = await self._get_redis()
            channel = "log_lifecycle_updates"

            message = {
                'type': f"log.lifecycle.{event_type}",
                'timestamp': datetime.utcnow().isoformat() + 'Z',
                'user_id': user_id,
                'payload': payload,
            }
            await redis_conn.publish(channel, json.dumps(message))
            self.logger.debug(f"Published lifecycle event: {event_type} for user {user_id}")

        except Exception as e:
            self.logger.error(f"Failed to publish lifecycle event: {e}")

    def _check_idempotency(self, request: web.Request, user_id: int):
        """Check idempotency key and return cached response if exists."""
        key = request.headers.get('Idempotency-Key')
        if not key:
            return None, None

        cached = self.db.check_idempotency_key(key, user_id)
        if cached:
            return key, cached
        return key, None

    def _store_idempotency(self, key: str, user_id: int, status: int, body: dict):
        """Store idempotency key with response."""
        if key:
            self.db.store_idempotency_key(key, user_id, status, body)

    @require_auth
    async def list_positions(self, request: web.Request) -> web.Response:
        """GET /api/positions - List positions for user."""
        try:
            user_id = request['user']['id']
            status = request.query.get('status')  # 'planned', 'open', 'closed'
            limit = int(request.query.get('limit', 100))
            offset = int(request.query.get('offset', 0))

            positions = self.db.list_positions(user_id, status=status, limit=limit, offset=offset)

            return self._json_response({
                'success': True,
                'data': [p.to_api_dict() for p in positions],
                'count': len(positions)
            }, request=request)

        except Exception as e:
            self.logger.error(f"list_positions error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def get_position(self, request: web.Request) -> web.Response:
        """GET /api/positions/:id - Get a single position with legs."""
        try:
            user_id = request['user']['id']
            position_id = request.match_info['id']

            position = self.db.get_position(position_id, user_id, include_legs=True, include_fills=True)

            if not position:
                return self._error_response('Position not found', 404, request)

            return self._json_response({
                'success': True,
                'data': position.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_position error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def get_position_snapshot(self, request: web.Request) -> web.Response:
        """GET /api/positions/:id/snapshot - Get complete position snapshot for RiskGraph."""
        try:
            user_id = request['user']['id']
            position_id = request.match_info['id']

            snapshot = self.db.get_position_snapshot(position_id, user_id)

            if not snapshot:
                return self._error_response('Position not found', 404, request)

            return self._json_response({
                'success': True,
                'data': snapshot
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_position_snapshot error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def create_position(self, request: web.Request) -> web.Response:
        """POST /api/positions - Create a new position with legs."""
        try:
            user_id = request['user']['id']

            # Check idempotency
            idem_key, cached = self._check_idempotency(request, user_id)
            if cached:
                return self._json_response(cached['body'], request=request, status=cached['status'])

            body = await request.json()

            # Validate required fields
            if not body.get('symbol'):
                return self._error_response('symbol is required', 400, request)
            if not body.get('legs') or not isinstance(body['legs'], list):
                return self._error_response('legs array is required', 400, request)

            # Create position
            position = Position(
                id=Position.new_id(),
                user_id=user_id,
                status=body.get('status', 'planned'),
                symbol=body['symbol'],
                underlying=body.get('underlying', f"I:{body['symbol']}"),
                tags=body.get('tags'),
                campaign_id=body.get('campaignId'),
            )

            # Create legs
            legs = []
            for leg_data in body['legs']:
                leg = Leg(
                    id=Leg.new_id(),
                    position_id=position.id,
                    instrument_type=leg_data.get('instrumentType', 'option'),
                    expiry=leg_data.get('expiry'),
                    strike=leg_data.get('strike'),
                    right=leg_data.get('right'),
                    quantity=leg_data.get('quantity', 0),
                )
                legs.append(leg)

            created = self.db.create_position(position, legs)

            response_body = {
                'success': True,
                'data': created.to_api_dict()
            }

            # Store idempotency
            self._store_idempotency(idem_key, user_id, 201, response_body)

            # Publish event
            await self._publish_trade_log_event(
                user_id=user_id,
                event_type='PositionCreated',
                aggregate_id=created.id,
                aggregate_version=created.version,
                payload=created.to_api_dict()
            )

            return self._json_response(response_body, request=request, status=201)

        except Exception as e:
            self.logger.error(f"create_position error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def update_position(self, request: web.Request) -> web.Response:
        """PATCH /api/positions/:id - Update position (If-Match version optional)."""
        try:
            user_id = request['user']['id']
            position_id = request.match_info['id']

            # Get version from If-Match header (optional - skip version check if not provided)
            version_header = request.headers.get('If-Match')
            version = None
            if version_header:
                try:
                    version = int(version_header)
                except ValueError:
                    return self._error_response('Invalid If-Match header', 400, request)

            body = await request.json()

            # Convert camelCase to snake_case for updates
            updates = {}
            if 'status' in body:
                updates['status'] = body['status']
            if 'visible' in body:
                updates['visible'] = body['visible']
            if 'tags' in body:
                updates['tags'] = body['tags']
            if 'campaignId' in body:
                updates['campaign_id'] = body['campaignId']
            if 'openedAt' in body:
                updates['opened_at'] = body['openedAt']
            if 'closedAt' in body:
                updates['closed_at'] = body['closedAt']

            updated = self.db.update_position(position_id, user_id, version, updates)

            if not updated:
                return self._error_response('Position not found', 404, request)

            # Publish event
            await self._publish_trade_log_event(
                user_id=user_id,
                event_type='PositionAdjusted',
                aggregate_id=updated.id,
                aggregate_version=updated.version,
                payload=updated.to_api_dict()
            )

            return self._json_response({
                'success': True,
                'data': updated.to_api_dict()
            }, request=request)

        except VersionConflictError as e:
            return self._json_response({
                'success': False,
                'error': 'Version conflict',
                'currentVersion': e.current_version
            }, request=request, status=409)

        except Exception as e:
            self.logger.error(f"update_position error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def delete_position(self, request: web.Request) -> web.Response:
        """DELETE /api/positions/:id - Delete a position."""
        try:
            user_id = request['user']['id']
            position_id = request.match_info['id']

            deleted = self.db.delete_position(position_id, user_id)

            if not deleted:
                return self._error_response('Position not found', 404, request)

            # Publish event
            await self._publish_trade_log_event(
                user_id=user_id,
                event_type='PositionDeleted',
                aggregate_id=position_id,
                aggregate_version=0,
                payload={'id': position_id}
            )

            return self._json_response({
                'success': True
            }, request=request)

        except Exception as e:
            self.logger.error(f"delete_position error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def record_fill(self, request: web.Request) -> web.Response:
        """POST /api/positions/:id/fills - Record a fill for a position leg."""
        try:
            user_id = request['user']['id']
            position_id = request.match_info['id']

            # Check idempotency
            idem_key, cached = self._check_idempotency(request, user_id)
            if cached:
                return self._json_response(cached['body'], request=request, status=cached['status'])

            body = await request.json()

            # Validate
            if not body.get('legId'):
                return self._error_response('legId is required', 400, request)
            if body.get('price') is None:
                return self._error_response('price is required', 400, request)
            if body.get('quantity') is None:
                return self._error_response('quantity is required', 400, request)

            # Verify position belongs to user
            position = self.db.get_position(position_id, user_id)
            if not position:
                return self._error_response('Position not found', 404, request)

            # Create fill
            fill = Fill(
                id=Fill.new_id(),
                leg_id=body['legId'],
                price=float(body['price']),
                quantity=int(body['quantity']),
                occurred_at=body.get('occurredAt', datetime.utcnow().isoformat()),
            )

            self.db.record_fill(fill)

            # Update position to 'open' if it was 'planned'
            if position.status == 'planned':
                try:
                    self.db.update_position(position_id, user_id, position.version, {
                        'status': 'open',
                        'opened_at': fill.occurred_at
                    })
                except VersionConflictError:
                    pass  # Another fill may have updated it, that's fine

            # Refresh position
            updated_position = self.db.get_position(position_id, user_id, include_legs=True, include_fills=True)

            response_body = {
                'success': True,
                'data': updated_position.to_api_dict() if updated_position else None
            }

            # Store idempotency
            self._store_idempotency(idem_key, user_id, 201, response_body)

            # Publish event
            await self._publish_trade_log_event(
                user_id=user_id,
                event_type='FillRecorded',
                aggregate_id=position_id,
                aggregate_version=updated_position.version if updated_position else position.version + 1,
                payload={
                    'fill': fill.to_api_dict(),
                    'position': updated_position.to_api_dict() if updated_position else None
                }
            )

            return self._json_response(response_body, request=request, status=201)

        except Exception as e:
            self.logger.error(f"record_fill error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def close_position(self, request: web.Request) -> web.Response:
        """POST /api/positions/:id/close - Close a position (requires If-Match version)."""
        try:
            user_id = request['user']['id']
            position_id = request.match_info['id']

            # Get version from If-Match header
            version_header = request.headers.get('If-Match')
            if not version_header:
                return self._error_response('If-Match header required for close', 428, request)

            try:
                version = int(version_header)
            except ValueError:
                return self._error_response('Invalid If-Match header', 400, request)

            closed = self.db.close_position(position_id, user_id, version)

            if not closed:
                return self._error_response('Position not found', 404, request)

            # Publish event
            await self._publish_trade_log_event(
                user_id=user_id,
                event_type='PositionClosed',
                aggregate_id=closed.id,
                aggregate_version=closed.version,
                payload=closed.to_api_dict()
            )

            return self._json_response({
                'success': True,
                'data': closed.to_api_dict()
            }, request=request)

        except VersionConflictError as e:
            return self._json_response({
                'success': False,
                'error': 'Version conflict',
                'currentVersion': e.current_version
            }, request=request, status=409)

        except Exception as e:
            self.logger.error(f"close_position error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def list_journal_entries_for_position(self, request: web.Request) -> web.Response:
        """GET /api/journal_entries?position_id=... - Get journal entries for a position."""
        try:
            user_id = request['user']['id']
            position_id = request.query.get('position_id')

            if not position_id:
                return self._error_response('position_id query parameter required', 400, request)

            # Verify position belongs to user
            position = self.db.get_position(position_id, user_id, include_legs=False)
            if not position:
                return self._error_response('Position not found', 404, request)

            # Get journal entries - reuse existing journal entry infrastructure
            # For now, filter journal entries by trade_refs that match position_id
            # This integrates with the existing journaling system

            # Note: The existing journal system uses JournalTradeRef for linking
            # For the new Position model, we'll store position_id in notes or a new field
            # For MVP, return empty array until full journal integration is done

            return self._json_response({
                'success': True,
                'data': [],
                'count': 0
            }, request=request)

        except Exception as e:
            self.logger.error(f"list_journal_entries_for_position error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def create_journal_entry_for_position(self, request: web.Request) -> web.Response:
        """POST /api/journal_entries - Create a journal entry for a position."""
        try:
            user_id = request['user']['id']
            body = await request.json()

            position_id = body.get('positionId') or body.get('position_id')
            if not position_id:
                return self._error_response('positionId is required', 400, request)

            object_of_reflection = body.get('objectOfReflection') or body.get('object_of_reflection')
            if not object_of_reflection:
                return self._error_response('objectOfReflection is required', 400, request)

            phase = body.get('phase', 'setup')

            # Verify position belongs to user
            position = self.db.get_position(position_id, user_id, include_legs=False)
            if not position:
                return self._error_response('Position not found', 404, request)

            # Create journal entry using existing infrastructure
            # For MVP, store as a journal entry with position reference in notes

            entry_id = str(__import__('uuid').uuid4())
            entry_data = {
                'id': entry_id,
                'positionId': position_id,
                'objectOfReflection': object_of_reflection,
                'biasFlags': body.get('biasFlags') or body.get('bias_flags'),
                'notes': body.get('notes'),
                'phase': phase,
                'createdAt': datetime.utcnow().isoformat(),
            }

            # Note: Full integration with JournalEntry model would go here
            # For MVP, return the created entry structure

            return self._json_response({
                'success': True,
                'data': entry_data
            }, request=request, status=201)

        except Exception as e:
            self.logger.error(f"create_journal_entry_for_position error: {e}")
            return self._error_response(str(e), 500, request)

    # ==================== ML Feedback Loop API ====================

    async def log_ml_decision(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/decisions - Log an ML scoring decision."""
        try:
            # Internal endpoint - basic localhost check
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            body = await request.json()

            # Validate required fields
            required = ['idea_id', 'original_score', 'final_score', 'selector_params_version', 'feature_snapshot_id']
            for field in required:
                if field not in body:
                    return self._error_response(f'{field} is required', 400, request)

            decision = MLDecision(
                idea_id=body['idea_id'],
                decision_time=datetime.utcnow(),
                model_id=body.get('model_id'),
                model_version=body.get('model_version'),
                selector_params_version=body['selector_params_version'],
                feature_snapshot_id=body['feature_snapshot_id'],
                original_score=float(body['original_score']),
                ml_score=float(body['ml_score']) if body.get('ml_score') is not None else None,
                final_score=float(body['final_score']),
                experiment_id=body.get('experiment_id'),
                experiment_arm=body.get('experiment_arm'),
                action_taken=body.get('action_taken', 'ranked'),
            )

            created = self.db.create_ml_decision(decision)

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, status=201, request=request)

        except Exception as e:
            self.logger.error(f"log_ml_decision error: {e}")
            return self._error_response(str(e), 500, request)

    async def list_ml_decisions(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/decisions - List ML decisions."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            params = request.query
            idea_id = params.get('idea_id')
            model_id = params.get('model_id')
            experiment_id = params.get('experiment_id')
            limit = int(params.get('limit', 100))

            decisions = self.db.list_ml_decisions(
                idea_id=idea_id,
                model_id=int(model_id) if model_id else None,
                experiment_id=int(experiment_id) if experiment_id else None,
                limit=limit
            )

            return self._json_response({
                'success': True,
                'data': [d.to_api_dict() for d in decisions],
                'count': len(decisions)
            }, request=request)

        except Exception as e:
            self.logger.error(f"list_ml_decisions error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_ml_decision_stats(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/decisions/stats - Get ML decision statistics."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            # Get stats directly from database
            stats = self.db.get_ml_decision_stats()

            return self._json_response({
                'success': True,
                'data': stats
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_ml_decision_stats error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_ml_decision(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/decisions/:id - Get a single ML decision."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            decision_id = int(request.match_info['id'])
            decision = self.db.get_ml_decision(decision_id)

            if not decision:
                return self._error_response('Decision not found', 404, request)

            return self._json_response({
                'success': True,
                'data': decision.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_ml_decision error: {e}")
            return self._error_response(str(e), 500, request)

    async def update_ml_decision_action(self, request: web.Request) -> web.Response:
        """PATCH /api/internal/ml/decisions/:id/action - Update action taken."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            decision_id = int(request.match_info['id'])
            body = await request.json()

            action = body.get('action_taken')
            if action not in ('ranked', 'presented', 'traded', 'dismissed'):
                return self._error_response('Invalid action_taken value', 400, request)

            updated = self.db.update_ml_decision_action(decision_id, action)
            if not updated:
                return self._error_response('Decision not found or invalid state transition', 404, request)

            return self._json_response({
                'success': True,
                'data': updated.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"update_ml_decision_action error: {e}")
            return self._error_response(str(e), 500, request)

    async def record_pnl_event(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/pnl-events - Record a P&L event."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            body = await request.json()

            required = ['idea_id', 'pnl_delta', 'underlying_price', 'event_type']
            for field in required:
                if field not in body:
                    return self._error_response(f'{field} is required', 400, request)

            event = PnLEvent(
                event_time=datetime.utcnow(),
                idea_id=body['idea_id'],
                trade_id=body.get('trade_id'),
                strategy_id=body.get('strategy_id'),
                pnl_delta=float(body['pnl_delta']),
                fees=float(body.get('fees', 0)),
                slippage=float(body.get('slippage', 0)),
                underlying_price=float(body['underlying_price']),
                event_type=body['event_type'],
            )

            created = self.db.create_pnl_event(event)

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, status=201, request=request)

        except Exception as e:
            self.logger.error(f"record_pnl_event error: {e}")
            return self._error_response(str(e), 500, request)

    async def list_pnl_events(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/pnl-events - List P&L events."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            params = request.query
            idea_id = params.get('idea_id')
            from_date = params.get('from')
            to_date = params.get('to')
            limit = int(params.get('limit', 100))

            events = self.db.list_pnl_events(
                idea_id=idea_id,
                from_date=from_date,
                to_date=to_date,
                limit=limit
            )

            return self._json_response({
                'success': True,
                'data': [e.to_api_dict() for e in events],
                'count': len(events)
            }, request=request)

        except Exception as e:
            self.logger.error(f"list_pnl_events error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_equity_curve(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/equity-curve - Get equity curve from P&L events."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            params = request.query
            from_date = params.get('from')
            to_date = params.get('to')

            curve = self.db.compute_equity_curve(from_date, to_date)

            return self._json_response({
                'success': True,
                'data': curve
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_equity_curve error: {e}")
            return self._error_response(str(e), 500, request)

    async def list_daily_performance(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/daily-performance - List daily performance records."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            params = request.query
            from_date = params.get('from')
            to_date = params.get('to')
            limit = int(params.get('limit', 30))

            records = self.db.list_daily_performance(from_date, to_date, limit)

            return self._json_response({
                'success': True,
                'data': [r.to_api_dict() for r in records],
                'count': len(records)
            }, request=request)

        except Exception as e:
            self.logger.error(f"list_daily_performance error: {e}")
            return self._error_response(str(e), 500, request)

    async def materialize_daily_performance(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/daily-performance/materialize - Materialize daily performance."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            body = await request.json()
            target_date = body.get('date')  # ISO format date string

            if not target_date:
                from datetime import date
                target_date = date.today().isoformat()

            result = self.db.materialize_daily_performance(target_date)

            return self._json_response({
                'success': True,
                'data': result.to_api_dict() if result else None
            }, request=request)

        except Exception as e:
            self.logger.error(f"materialize_daily_performance error: {e}")
            return self._error_response(str(e), 500, request)

    async def create_feature_snapshot(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/feature-snapshots - Create a feature snapshot."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            body = await request.json()

            required = ['tracked_idea_id', 'feature_set_version', 'feature_extractor_version', 'spot_price']
            for field in required:
                if field not in body:
                    return self._error_response(f'{field} is required', 400, request)

            snapshot = MLFeatureSnapshot(
                tracked_idea_id=body['tracked_idea_id'],
                snapshot_time=datetime.utcnow(),
                feature_set_version=body['feature_set_version'],
                feature_extractor_version=body['feature_extractor_version'],
                gex_calc_version=body.get('gex_calc_version'),
                vix_regime_classifier_version=body.get('vix_regime_classifier_version'),
                spot_price=float(body['spot_price']),
                spot_5m_return=body.get('spot_5m_return'),
                spot_15m_return=body.get('spot_15m_return'),
                spot_1h_return=body.get('spot_1h_return'),
                spot_1d_return=body.get('spot_1d_return'),
                intraday_high=body.get('intraday_high'),
                intraday_low=body.get('intraday_low'),
                range_position=body.get('range_position'),
                vix_level=body.get('vix_level'),
                vix_regime=body.get('vix_regime'),
                vix_term_slope=body.get('vix_term_slope'),
                iv_rank_30d=body.get('iv_rank_30d'),
                iv_percentile_30d=body.get('iv_percentile_30d'),
                gex_total=body.get('gex_total'),
                gex_call_wall=body.get('gex_call_wall'),
                gex_put_wall=body.get('gex_put_wall'),
                gex_gamma_flip=body.get('gex_gamma_flip'),
                spot_vs_call_wall=body.get('spot_vs_call_wall'),
                spot_vs_put_wall=body.get('spot_vs_put_wall'),
                spot_vs_gamma_flip=body.get('spot_vs_gamma_flip'),
                market_mode=body.get('market_mode'),
                bias_lfi=body.get('bias_lfi'),
                bias_direction=body.get('bias_direction'),
                minutes_since_open=body.get('minutes_since_open'),
                day_of_week=body.get('day_of_week'),
                is_opex_week=body.get('is_opex_week', False),
                days_to_monthly_opex=body.get('days_to_monthly_opex'),
                es_futures_premium=body.get('es_futures_premium'),
                tnx_level=body.get('tnx_level'),
                dxy_level=body.get('dxy_level'),
            )

            created = self.db.create_feature_snapshot(snapshot)

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, status=201, request=request)

        except Exception as e:
            self.logger.error(f"create_feature_snapshot error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_feature_snapshot(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/feature-snapshots/:id - Get a feature snapshot."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            snapshot_id = int(request.match_info['id'])
            snapshot = self.db.get_feature_snapshot(snapshot_id)

            if not snapshot:
                return self._error_response('Snapshot not found', 404, request)

            return self._json_response({
                'success': True,
                'data': snapshot.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_feature_snapshot error: {e}")
            return self._error_response(str(e), 500, request)

    async def list_ml_models(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/models - List ML models."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            params = request.query
            status = params.get('status')
            limit = int(params.get('limit', 20))

            models = self.db.list_ml_models(status=status, limit=limit)

            # Don't include model_blob in list response
            return self._json_response({
                'success': True,
                'data': [m.to_api_dict(include_blob=False) for m in models],
                'count': len(models)
            }, request=request)

        except Exception as e:
            self.logger.error(f"list_ml_models error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_ml_model(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/models/:id - Get a single ML model."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            model_id = int(request.match_info['id'])
            include_blob = request.query.get('include_blob', '').lower() == 'true'

            model = self.db.get_ml_model(model_id)

            if not model:
                return self._error_response('Model not found', 404, request)

            return self._json_response({
                'success': True,
                'data': model.to_api_dict(include_blob=include_blob)
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_ml_model error: {e}")
            return self._error_response(str(e), 500, request)

    async def register_ml_model(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/models - Register a new ML model."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            body = await request.json()

            required = ['model_name', 'model_type', 'model_blob', 'feature_list', 'hyperparameters']
            for field in required:
                if field not in body:
                    return self._error_response(f'{field} is required', 400, request)

            # Get next version for this model name
            next_version = self.db.get_next_model_version(body['model_name'])

            model = MLModel(
                model_name=body['model_name'],
                model_version=next_version,
                model_type=body['model_type'],
                model_blob=body['model_blob'],  # Base64 encoded
                feature_list=body['feature_list'],
                hyperparameters=body['hyperparameters'],
                train_auc=body.get('train_auc'),
                val_auc=body.get('val_auc'),
                train_samples=body.get('train_samples'),
                val_samples=body.get('val_samples'),
                status='validating',
            )

            created = self.db.create_ml_model(model)

            self.logger.info(f"Registered ML model {created.model_name} v{created.model_version}", emoji="🤖")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict(include_blob=False)
            }, status=201, request=request)

        except Exception as e:
            self.logger.error(f"register_ml_model error: {e}")
            return self._error_response(str(e), 500, request)

    async def deploy_ml_model(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/models/:id/deploy - Deploy model as champion."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            model_id = int(request.match_info['id'])
            deployed = self.db.deploy_ml_model(model_id)

            if not deployed:
                return self._error_response('Model not found or invalid state', 404, request)

            self.logger.info(f"Deployed ML model {deployed.model_name} v{deployed.model_version} as champion", emoji="🚀")

            return self._json_response({
                'success': True,
                'data': deployed.to_api_dict(include_blob=False)
            }, request=request)

        except Exception as e:
            self.logger.error(f"deploy_ml_model error: {e}")
            return self._error_response(str(e), 500, request)

    async def retire_ml_model(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/models/:id/retire - Retire a model."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            model_id = int(request.match_info['id'])
            retired = self.db.retire_ml_model(model_id)

            if not retired:
                return self._error_response('Model not found', 404, request)

            self.logger.info(f"Retired ML model {retired.model_name} v{retired.model_version}", emoji="📦")

            return self._json_response({
                'success': True,
                'data': retired.to_api_dict(include_blob=False)
            }, request=request)

        except Exception as e:
            self.logger.error(f"retire_ml_model error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_champion_model(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/models/champion - Get current champion model."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            include_blob = request.query.get('include_blob', '').lower() == 'true'
            regime = request.query.get('regime')

            champion = self.db.get_champion_model(regime=regime, include_blob=include_blob)

            if not champion:
                return self._json_response({
                    'success': True,
                    'data': None
                }, request=request)

            return self._json_response({
                'success': True,
                'data': champion.to_api_dict(include_blob=include_blob)
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_champion_model error: {e}")
            return self._error_response(str(e), 500, request)

    async def list_experiments(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/experiments - List ML experiments."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            params = request.query
            status = params.get('status')
            limit = int(params.get('limit', 20))

            experiments = self.db.list_experiments(status=status, limit=limit)

            return self._json_response({
                'success': True,
                'data': [e.to_api_dict() for e in experiments],
                'count': len(experiments)
            }, request=request)

        except Exception as e:
            self.logger.error(f"list_experiments error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_experiment(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/experiments/:id - Get a single experiment."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            experiment_id = int(request.match_info['id'])
            experiment = self.db.get_experiment(experiment_id)

            if not experiment:
                return self._error_response('Experiment not found', 404, request)

            return self._json_response({
                'success': True,
                'data': experiment.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_experiment error: {e}")
            return self._error_response(str(e), 500, request)

    async def create_experiment(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/experiments - Create a new A/B experiment."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            body = await request.json()

            required = ['experiment_name', 'challenger_model_id']
            for field in required:
                if field not in body:
                    return self._error_response(f'{field} is required', 400, request)

            # Get champion model for the experiment
            champion = self.db.get_champion_model()
            if not champion and not body.get('champion_model_id'):
                return self._error_response('No champion model available', 400, request)

            experiment = MLExperiment(
                experiment_name=body['experiment_name'],
                description=body.get('description'),
                champion_model_id=body.get('champion_model_id', champion.id if champion else None),
                challenger_model_id=body['challenger_model_id'],
                traffic_split=float(body.get('traffic_split', 0.10)),
                status='running',
                started_at=datetime.utcnow(),
            )

            created = self.db.create_experiment(experiment)

            self.logger.info(f"Created ML experiment '{created.experiment_name}'", emoji="🧪")

            return self._json_response({
                'success': True,
                'data': created.to_api_dict()
            }, status=201, request=request)

        except Exception as e:
            self.logger.error(f"create_experiment error: {e}")
            return self._error_response(str(e), 500, request)

    async def evaluate_experiment(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/experiments/:id/evaluate - Evaluate experiment results."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            experiment_id = int(request.match_info['id'])
            result = self.db.evaluate_experiment(experiment_id)

            if not result:
                return self._error_response('Experiment not found', 404, request)

            return self._json_response({
                'success': True,
                'data': result
            }, request=request)

        except Exception as e:
            self.logger.error(f"evaluate_experiment error: {e}")
            return self._error_response(str(e), 500, request)

    async def conclude_experiment(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/experiments/:id/conclude - Conclude experiment."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            experiment_id = int(request.match_info['id'])
            body = await request.json()

            winner = body.get('winner')
            promote_challenger = body.get('promote_challenger', False)

            concluded = self.db.conclude_experiment(experiment_id, winner, promote_challenger)

            if not concluded:
                return self._error_response('Experiment not found or already concluded', 404, request)

            self.logger.info(f"Concluded ML experiment {experiment_id}, winner: {winner}", emoji="🏆")

            return self._json_response({
                'success': True,
                'data': concluded.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"conclude_experiment error: {e}")
            return self._error_response(str(e), 500, request)

    async def abort_experiment(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/experiments/:id/abort - Abort experiment."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            experiment_id = int(request.match_info['id'])
            body = await request.json()
            reason = body.get('reason')

            aborted = self.db.abort_experiment(experiment_id, reason)

            if not aborted:
                return self._error_response('Experiment not found or not running', 404, request)

            self.logger.info(f"Aborted ML experiment {experiment_id}: {reason}", emoji="🛑")

            return self._json_response({
                'success': True,
                'data': aborted.to_api_dict()
            }, request=request)

        except Exception as e:
            self.logger.error(f"abort_experiment error: {e}")
            return self._error_response(str(e), 500, request)

    async def get_circuit_breaker_status(self, request: web.Request) -> web.Response:
        """GET /api/internal/ml/circuit-breakers - Get circuit breaker status."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            status = self.db.get_circuit_breaker_status()

            return self._json_response({
                'success': True,
                'data': status
            }, request=request)

        except Exception as e:
            self.logger.error(f"get_circuit_breaker_status error: {e}")
            return self._error_response(str(e), 500, request)

    async def check_circuit_breakers(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/circuit-breakers/check - Check all circuit breakers."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            result = self.db.check_circuit_breakers()

            return self._json_response({
                'success': True,
                'data': result
            }, request=request)

        except Exception as e:
            self.logger.error(f"check_circuit_breakers error: {e}")
            return self._error_response(str(e), 500, request)

    async def disable_ml(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/circuit-breakers/disable-ml - Disable ML scoring."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            self.db.set_ml_enabled(False)
            self.logger.warn("ML scoring disabled via kill switch", emoji="🛑")

            return self._json_response({
                'success': True,
                'message': 'ML scoring disabled'
            }, request=request)

        except Exception as e:
            self.logger.error(f"disable_ml error: {e}")
            return self._error_response(str(e), 500, request)

    async def enable_ml(self, request: web.Request) -> web.Response:
        """POST /api/internal/ml/circuit-breakers/enable-ml - Enable ML scoring."""
        try:
            peername = request.transport.get_extra_info('peername')
            if peername and peername[0] not in ('127.0.0.1', '::1', 'localhost'):
                return self._error_response('Internal endpoint only', 403, request)

            self.db.set_ml_enabled(True)
            self.logger.info("ML scoring enabled", emoji="✅")

            return self._json_response({
                'success': True,
                'message': 'ML scoring enabled'
            }, request=request)

        except Exception as e:
            self.logger.error(f"enable_ml error: {e}")
            return self._error_response(str(e), 500, request)

    # ==================== Health & App Setup ====================

    async def health_check(self, request: web.Request) -> web.Response:
        """GET /health - Health check endpoint."""
        return self._json_response({
            'success': True,
            'service': 'journal',
            'version': 'v2',
            'status': 'healthy',
            'ts': datetime.utcnow().isoformat()
        })

    @web.middleware
    async def cors_middleware(self, request: web.Request, handler):
        """Add CORS headers to all responses."""
        # Handle preflight
        if request.method == 'OPTIONS':
            return await self.handle_options(request)

        # Process request
        response = await handler(request)

        # Add CORS headers
        origin = self._get_cors_origin(request)
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'

        return response

    def create_app(self) -> web.Application:
        """Create the aiohttp application with routes."""
        app = web.Application(middlewares=[self.cors_middleware])

        # CORS preflight for all routes
        app.router.add_route('OPTIONS', '/{tail:.*}', self.handle_options)

        # Health check
        app.router.add_get('/health', self.health_check)

        # Trade Log CRUD (v2)
        app.router.add_get('/api/logs', self.list_logs)
        app.router.add_get('/api/logs/{id}', self.get_log)
        app.router.add_post('/api/logs', self.create_log)
        app.router.add_put('/api/logs/{id}', self.update_log)
        app.router.add_delete('/api/logs/{id}', self.delete_log)

        # Trade Log Lifecycle Management
        app.router.add_post('/api/logs/{id}/archive', self.archive_log)
        app.router.add_post('/api/logs/{id}/reactivate', self.reactivate_log)
        app.router.add_post('/api/logs/{id}/retire', self.schedule_retire_log)
        app.router.add_delete('/api/logs/{id}/retire', self.cancel_retire_log)
        app.router.add_get('/api/logs/for-import', self.get_logs_for_import)
        app.router.add_post('/api/logs/import-recommendation', self.get_import_recommendation)

        # User/Log health endpoints (for Vexy context feed)
        app.router.add_get('/api/users/with-logs', self.get_users_with_logs)
        app.router.add_get('/api/logs/health', self.get_logs_health_metrics)

        # Log-scoped trades (v2)
        app.router.add_get('/api/logs/{logId}/trades', self.list_trades)
        app.router.add_post('/api/logs/{logId}/trades', self.create_trade)

        # Trade operations
        app.router.add_get('/api/trades/{id}', self.get_trade)
        app.router.add_put('/api/trades/{id}', self.update_trade)
        app.router.add_delete('/api/trades/{id}', self.delete_trade)
        app.router.add_post('/api/trades/{id}/adjust', self.add_adjustment)
        app.router.add_post('/api/trades/{id}/close', self.close_trade)

        # Order Queue (Simulated Trading)
        app.router.add_get('/api/orders', self.list_orders)
        app.router.add_get('/api/orders/active', self.get_orders_active)
        app.router.add_post('/api/orders', self.create_order)
        app.router.add_delete('/api/orders/{id}', self.cancel_order)

        # Log-scoped analytics (v2)
        app.router.add_get('/api/logs/{logId}/analytics', self.get_log_analytics)
        app.router.add_get('/api/logs/{logId}/equity', self.get_log_equity)
        app.router.add_get('/api/logs/{logId}/drawdown', self.get_log_drawdown)
        app.router.add_get('/api/logs/{logId}/distribution', self.get_return_distribution)

        # Export/Import
        app.router.add_get('/api/logs/{logId}/export', self.export_trades)
        app.router.add_post('/api/logs/{logId}/import', self.import_trades)

        # Symbols registry
        app.router.add_get('/api/symbols', self.list_symbols)
        app.router.add_get('/api/symbols/{symbol}', self.get_symbol)
        app.router.add_post('/api/symbols', self.add_symbol)
        app.router.add_put('/api/symbols/{symbol}', self.update_symbol)
        app.router.add_delete('/api/symbols/{symbol}', self.delete_symbol)

        # Tags (Vocabulary System)
        app.router.add_get('/api/tags', self.list_tags)
        app.router.add_get('/api/tags/{id}', self.get_tag)
        app.router.add_post('/api/tags', self.create_tag)
        app.router.add_post('/api/tags/seed', self.seed_tags)
        app.router.add_put('/api/tags/{id}', self.update_tag)
        app.router.add_put('/api/tags/{id}/retire', self.retire_tag)
        app.router.add_put('/api/tags/{id}/restore', self.restore_tag)
        app.router.add_delete('/api/tags/{id}', self.delete_tag)

        # Settings
        app.router.add_get('/api/settings', self.get_settings)
        app.router.add_get('/api/settings/{key}', self.get_setting)
        app.router.add_put('/api/settings/{key}', self.set_setting)
        app.router.add_delete('/api/settings/{key}', self.delete_setting)

        # Journal Entries
        app.router.add_get('/api/journal/entries', self.list_journal_entries)
        app.router.add_get('/api/journal/entries/date/{date}', self.get_journal_entry_by_date)
        app.router.add_get('/api/journal/entries/{id}', self.get_journal_entry)
        app.router.add_post('/api/journal/entries', self.create_journal_entry)
        app.router.add_put('/api/journal/entries/{id}', self.update_journal_entry)
        app.router.add_delete('/api/journal/entries/{id}', self.delete_journal_entry)

        # Journal Retrospectives
        app.router.add_get('/api/journal/retrospectives', self.list_retrospectives)
        app.router.add_get('/api/journal/retrospectives/{type}/{periodStart}', self.get_retrospective_by_period)
        app.router.add_get('/api/journal/retrospectives/{id}', self.get_retrospective)
        app.router.add_post('/api/journal/retrospectives', self.create_retrospective)
        app.router.add_put('/api/journal/retrospectives/{id}', self.update_retrospective)
        app.router.add_delete('/api/journal/retrospectives/{id}', self.delete_retrospective)

        # Journal Trade References
        app.router.add_get('/api/journal/entries/{id}/trades', self.list_entry_trade_refs)
        app.router.add_post('/api/journal/entries/{id}/trades', self.add_entry_trade_ref)
        app.router.add_get('/api/journal/retrospectives/{id}/trades', self.list_retro_trade_refs)
        app.router.add_post('/api/journal/retrospectives/{id}/trades', self.add_retro_trade_ref)
        app.router.add_delete('/api/journal/trade-refs/{id}', self.delete_trade_ref)

        # Journal Attachments
        app.router.add_get('/api/journal/entries/{id}/attachments', self.list_entry_attachments)
        app.router.add_post('/api/journal/entries/{id}/attachments', self.upload_entry_attachment)
        app.router.add_get('/api/journal/retrospectives/{id}/attachments', self.list_retro_attachments)
        app.router.add_post('/api/journal/retrospectives/{id}/attachments', self.upload_retro_attachment)
        app.router.add_get('/api/journal/attachments/{id}', self.download_attachment)
        app.router.add_delete('/api/journal/attachments/{id}', self.delete_attachment)

        # Journal Calendar Views (Temporal Gravity)
        app.router.add_get('/api/journal/calendar/{year}/{month}', self.get_calendar_month)
        app.router.add_get('/api/journal/calendar/week/{weekStart}', self.get_calendar_week)

        # Playbook
        app.router.add_get('/api/playbook/entries', self.list_playbook_entries)
        app.router.add_get('/api/playbook/entries/{id}', self.get_playbook_entry)
        app.router.add_post('/api/playbook/entries', self.create_playbook_entry)
        app.router.add_put('/api/playbook/entries/{id}', self.update_playbook_entry)
        app.router.add_delete('/api/playbook/entries/{id}', self.delete_playbook_entry)
        app.router.add_get('/api/playbook/entries/{id}/sources', self.list_playbook_sources)
        app.router.add_post('/api/playbook/entries/{id}/sources', self.add_playbook_source)
        app.router.add_delete('/api/playbook/sources/{id}', self.delete_playbook_source)
        app.router.add_get('/api/playbook/flagged-material', self.get_flagged_playbook_material)

        # Alerts
        app.router.add_get('/api/alerts', self.list_alerts)
        app.router.add_get('/api/alerts/{id}', self.get_alert)
        app.router.add_post('/api/alerts', self.create_alert)
        app.router.add_patch('/api/alerts/{id}', self.update_alert)
        app.router.add_delete('/api/alerts/{id}', self.delete_alert)
        app.router.add_post('/api/alerts/{id}/reset', self.reset_alert)
        app.router.add_get('/api/internal/alerts', self.list_all_alerts_internal)

        # Prompt Alerts
        app.router.add_get('/api/prompt-alerts', self.list_prompt_alerts)
        app.router.add_get('/api/prompt-alerts/{id}', self.get_prompt_alert)
        app.router.add_post('/api/prompt-alerts', self.create_prompt_alert)
        app.router.add_patch('/api/prompt-alerts/{id}', self.update_prompt_alert)
        app.router.add_delete('/api/prompt-alerts/{id}', self.delete_prompt_alert)
        app.router.add_get('/api/internal/prompt-alerts', self.list_prompt_alerts_internal)

        # Algo Alerts
        app.router.add_get('/api/algo-alerts', self.list_algo_alerts)
        app.router.add_post('/api/algo-alerts', self.create_algo_alert)
        app.router.add_put('/api/algo-alerts/{id}', self.update_algo_alert_route)
        app.router.add_delete('/api/algo-alerts/{id}', self.delete_algo_alert_route)
        app.router.add_get('/api/algo-proposals', self.list_algo_proposals)
        app.router.add_post('/api/algo-proposals/{id}/approve', self.approve_algo_proposal)
        app.router.add_post('/api/algo-proposals/{id}/reject', self.reject_algo_proposal)
        app.router.add_get('/api/internal/algo-alerts', self.list_algo_alerts_internal)
        app.router.add_put('/api/internal/algo-alerts/{id}/status', self.update_algo_alert_status_internal)
        app.router.add_post('/api/internal/algo-proposals', self.create_algo_proposal_internal)

        # Legacy endpoints (backwards compatibility)
        app.router.add_get('/api/trades', self.legacy_list_trades)
        app.router.add_post('/api/trades', self.legacy_create_trade)
        app.router.add_get('/api/analytics', self.legacy_analytics)
        app.router.add_get('/api/analytics/equity', self.legacy_equity)

        # =================================================================
        # Trade Idea Tracking (Feedback Optimization Loop) - Internal API
        # =================================================================
        app.router.add_post('/api/internal/tracked-ideas', self.create_tracked_idea)
        app.router.add_get('/api/internal/tracked-ideas', self.list_tracked_ideas)
        app.router.add_get('/api/internal/tracked-ideas/analytics', self.get_tracking_analytics)
        app.router.add_get('/api/internal/selector-params', self.list_selector_params)
        app.router.add_get('/api/internal/selector-params/active', self.get_active_params)
        app.router.add_post('/api/internal/selector-params', self.create_selector_params)
        app.router.add_post('/api/internal/selector-params/{version}/activate', self.activate_selector_params)

        # =================================================================
        # Leaderboard API
        # =================================================================
        app.router.add_get('/api/leaderboard', self.get_leaderboard)
        app.router.add_get('/api/leaderboard/me', self.get_my_leaderboard)
        app.router.add_post('/api/internal/leaderboard/calculate', self.trigger_leaderboard_calculation)

        # =================================================================
        # Risk Graph Service API
        # =================================================================
        # Strategies
        app.router.add_get('/api/risk-graph/strategies', self.list_risk_graph_strategies)
        app.router.add_get('/api/risk-graph/strategies/{id}', self.get_risk_graph_strategy)
        app.router.add_post('/api/risk-graph/strategies', self.create_risk_graph_strategy)
        app.router.add_patch('/api/risk-graph/strategies/{id}', self.update_risk_graph_strategy)
        app.router.add_delete('/api/risk-graph/strategies/{id}', self.delete_risk_graph_strategy)
        app.router.add_get('/api/risk-graph/strategies/{id}/versions', self.get_risk_graph_strategy_versions)

        # Bulk operations
        app.router.add_post('/api/risk-graph/strategies/import', self.import_risk_graph_strategies)
        app.router.add_get('/api/risk-graph/strategies/export', self.export_risk_graph_strategies)
        app.router.add_post('/api/risk-graph/strategies/reorder', self.reorder_risk_graph_strategies)

        # Templates
        app.router.add_get('/api/risk-graph/templates', self.list_risk_graph_templates)
        app.router.add_get('/api/risk-graph/templates/{id}', self.get_risk_graph_template)
        app.router.add_post('/api/risk-graph/templates', self.create_risk_graph_template)
        app.router.add_patch('/api/risk-graph/templates/{id}', self.update_risk_graph_template)
        app.router.add_delete('/api/risk-graph/templates/{id}', self.delete_risk_graph_template)
        app.router.add_post('/api/risk-graph/templates/{id}/use', self.use_risk_graph_template)
        app.router.add_post('/api/risk-graph/templates/{id}/share', self.share_risk_graph_template)
        app.router.add_get('/api/risk-graph/templates/shared/{code}', self.get_shared_risk_graph_template)

        # ==================== Position API (TradeLog Service Layer) ====================
        app.router.add_get('/api/positions', self.list_positions)
        app.router.add_get('/api/positions/{id}', self.get_position)
        app.router.add_get('/api/positions/{id}/snapshot', self.get_position_snapshot)
        app.router.add_post('/api/positions', self.create_position)
        app.router.add_patch('/api/positions/{id}', self.update_position)
        app.router.add_delete('/api/positions/{id}', self.delete_position)
        app.router.add_post('/api/positions/{id}/fills', self.record_fill)
        app.router.add_post('/api/positions/{id}/close', self.close_position)

        # Journal entries for positions
        app.router.add_get('/api/journal_entries', self.list_journal_entries_for_position)
        app.router.add_post('/api/journal_entries', self.create_journal_entry_for_position)

        # =================================================================
        # ML Feedback Loop API
        # =================================================================
        # ML Decisions (logging)
        app.router.add_post('/api/internal/ml/decisions', self.log_ml_decision)
        app.router.add_get('/api/internal/ml/decisions', self.list_ml_decisions)
        app.router.add_get('/api/internal/ml/decisions/stats', self.get_ml_decision_stats)
        app.router.add_get('/api/internal/ml/decisions/{id}', self.get_ml_decision)
        app.router.add_patch('/api/internal/ml/decisions/{id}/action', self.update_ml_decision_action)

        # P&L Events
        app.router.add_post('/api/internal/ml/pnl-events', self.record_pnl_event)
        app.router.add_get('/api/internal/ml/pnl-events', self.list_pnl_events)
        app.router.add_get('/api/internal/ml/equity-curve', self.get_equity_curve)

        # Daily Performance
        app.router.add_get('/api/internal/ml/daily-performance', self.list_daily_performance)
        app.router.add_post('/api/internal/ml/daily-performance/materialize', self.materialize_daily_performance)

        # Feature Snapshots
        app.router.add_post('/api/internal/ml/feature-snapshots', self.create_feature_snapshot)
        app.router.add_get('/api/internal/ml/feature-snapshots/{id}', self.get_feature_snapshot)

        # ML Models
        app.router.add_get('/api/internal/ml/models', self.list_ml_models)
        app.router.add_get('/api/internal/ml/models/{id}', self.get_ml_model)
        app.router.add_post('/api/internal/ml/models', self.register_ml_model)
        app.router.add_post('/api/internal/ml/models/{id}/deploy', self.deploy_ml_model)
        app.router.add_post('/api/internal/ml/models/{id}/retire', self.retire_ml_model)
        app.router.add_get('/api/internal/ml/models/champion', self.get_champion_model)

        # Experiments
        app.router.add_get('/api/internal/ml/experiments', self.list_experiments)
        app.router.add_get('/api/internal/ml/experiments/{id}', self.get_experiment)
        app.router.add_post('/api/internal/ml/experiments', self.create_experiment)
        app.router.add_post('/api/internal/ml/experiments/{id}/evaluate', self.evaluate_experiment)
        app.router.add_post('/api/internal/ml/experiments/{id}/conclude', self.conclude_experiment)
        app.router.add_post('/api/internal/ml/experiments/{id}/abort', self.abort_experiment)

        # Circuit Breakers
        app.router.add_get('/api/internal/ml/circuit-breakers', self.get_circuit_breaker_status)
        app.router.add_post('/api/internal/ml/circuit-breakers/check', self.check_circuit_breakers)
        app.router.add_post('/api/internal/ml/circuit-breakers/disable-ml', self.disable_ml)
        app.router.add_post('/api/internal/ml/circuit-breakers/enable-ml', self.enable_ml)

        # Edge Lab
        app.router.add_get('/api/edge-lab/setups', self.list_edge_lab_setups)
        app.router.add_get('/api/edge-lab/setups/{id}', self.get_edge_lab_setup)
        app.router.add_post('/api/edge-lab/setups', self.create_edge_lab_setup)
        app.router.add_patch('/api/edge-lab/setups/{id}', self.update_edge_lab_setup)
        app.router.add_get('/api/edge-lab/setups/{setupId}/hypothesis', self.get_edge_lab_hypothesis)
        app.router.add_post('/api/edge-lab/hypotheses', self.create_edge_lab_hypothesis)
        app.router.add_post('/api/edge-lab/hypotheses/{id}/lock', self.lock_edge_lab_hypothesis)
        app.router.add_get('/api/edge-lab/setups/{setupId}/outcome', self.get_edge_lab_outcome)
        app.router.add_post('/api/edge-lab/outcomes', self.create_edge_lab_outcome)
        app.router.add_post('/api/edge-lab/outcomes/{id}/confirm', self.confirm_edge_lab_outcome)

        # Edge Lab Analytics
        app.router.add_get('/api/edge-lab/setups/{setupId}/suggest-outcome', self.suggest_edge_lab_outcome)
        app.router.add_get('/api/edge-lab/analytics/regime-correlation', self.get_regime_correlation)
        app.router.add_get('/api/edge-lab/analytics/bias-overlay', self.get_bias_overlay)
        app.router.add_get('/api/edge-lab/analytics/edge-score', self.get_edge_score)
        app.router.add_get('/api/edge-lab/analytics/edge-score/history', self.get_edge_score_history)
        app.router.add_get('/api/edge-lab/analytics/signatures', self.get_signature_sample_sizes)
        app.router.add_post('/api/edge-lab/analytics/compute', self.compute_edge_lab_analytics)
        app.router.add_get('/api/edge-lab/analytics/dashboard', self.get_edge_lab_dashboard)
        app.router.add_post('/api/internal/edge-lab/materialize', self.materialize_edge_lab_metrics)

        return app

    # ==================== Edge Lab Handlers ====================

    @require_auth
    async def list_edge_lab_setups(self, request: web.Request) -> web.Response:
        """GET /api/edge-lab/setups — List setups for authenticated user."""
        try:
            user_id = request['user']['id']
            limit = int(request.query.get('limit', '50'))
            offset = int(request.query.get('offset', '0'))
            filters = {}
            for key in ['status', 'regime', 'position_structure', 'structure_signature',
                         'start_date', 'end_date']:
                val = request.query.get(key)
                if val:
                    filters[key] = val
            setups = self.db.list_edge_lab_setups(user_id, limit, offset, filters or None)
            return self._json_response({
                'success': True,
                'data': [s.to_api_dict() for s in setups],
                'count': len(setups),
            }, request=request)
        except Exception as e:
            self.logger.error(f"list_edge_lab_setups error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def get_edge_lab_setup(self, request: web.Request) -> web.Response:
        """GET /api/edge-lab/setups/{id} — Get a single setup."""
        try:
            user_id = request['user']['id']
            setup_id = request.match_info['id']
            setup = self.db.get_edge_lab_setup(setup_id, user_id)
            if not setup:
                return self._error_response('Setup not found', 404, request)

            # Include hypothesis and outcome if they exist
            result = setup.to_api_dict()
            hypothesis = self.db.get_hypothesis_for_setup(setup_id, user_id)
            if hypothesis:
                result['hypothesis'] = hypothesis.to_api_dict()
            outcome = self.db.get_outcome_for_setup(setup_id, user_id)
            if outcome:
                result['outcome'] = outcome.to_api_dict()

            return self._json_response({'success': True, 'data': result}, request=request)
        except Exception as e:
            self.logger.error(f"get_edge_lab_setup error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def create_edge_lab_setup(self, request: web.Request) -> web.Response:
        """POST /api/edge-lab/setups — Create a new setup."""
        try:
            user_id = request['user']['id']
            body = await request.json()

            # Validate required fields
            required = ['setup_date', 'regime', 'gex_posture', 'vol_state',
                         'time_structure', 'heatmap_color', 'position_structure',
                         'width_bucket', 'directional_bias']
            for field in required:
                camel = self._snake_to_camel(field)
                if not body.get(field) and not body.get(camel):
                    return self._error_response(f'{field} is required', 400, request)

            setup = EdgeLabSetup(
                id=EdgeLabSetup.new_id(),
                user_id=user_id,
                setup_date=body.get('setup_date') or body.get('setupDate'),
                regime=body.get('regime'),
                gex_posture=body.get('gex_posture') or body.get('gexPosture'),
                vol_state=body.get('vol_state') or body.get('volState'),
                time_structure=body.get('time_structure') or body.get('timeStructure'),
                heatmap_color=body.get('heatmap_color') or body.get('heatmapColor'),
                position_structure=body.get('position_structure') or body.get('positionStructure'),
                width_bucket=body.get('width_bucket') or body.get('widthBucket'),
                directional_bias=body.get('directional_bias') or body.get('directionalBias'),
                trade_id=body.get('trade_id') or body.get('tradeId'),
                position_id=body.get('position_id') or body.get('positionId'),
                entry_logic=body.get('entry_logic') or body.get('entryLogic'),
                exit_logic=body.get('exit_logic') or body.get('exitLogic'),
                entry_defined=1 if body.get('entry_defined') or body.get('entryDefined') else 0,
                exit_defined=1 if body.get('exit_defined') or body.get('exitDefined') else 0,
            )
            setup.structure_signature = setup.compute_structure_signature()

            # Auto-populate bias state from readiness log
            setup_date = body.get('setup_date') or body.get('setupDate')
            readiness = self.db.get_readiness_for_date(user_id, setup_date)
            if readiness:
                import json as _json
                setup.bias_state_json = _json.dumps({
                    'sleep': readiness.get('sleep'),
                    'focus': readiness.get('focus'),
                    'distractions': readiness.get('distractions'),
                    'body_state': readiness.get('body_state'),
                    'friction': readiness.get('friction'),
                })

            result = self.db.create_edge_lab_setup(setup)
            return self._json_response(
                {'success': True, 'data': result.to_api_dict()},
                request=request, status=201,
            )
        except Exception as e:
            self.logger.error(f"create_edge_lab_setup error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def update_edge_lab_setup(self, request: web.Request) -> web.Response:
        """PATCH /api/edge-lab/setups/{id} — Update a setup."""
        try:
            user_id = request['user']['id']
            setup_id = request.match_info['id']
            body = await request.json()

            # Convert camelCase to snake_case for known fields
            updates = {}
            field_map = {
                'tradeId': 'trade_id', 'positionId': 'position_id',
                'gexPosture': 'gex_posture', 'volState': 'vol_state',
                'timeStructure': 'time_structure', 'heatmapColor': 'heatmap_color',
                'positionStructure': 'position_structure', 'widthBucket': 'width_bucket',
                'directionalBias': 'directional_bias', 'entryLogic': 'entry_logic',
                'exitLogic': 'exit_logic', 'entryDefined': 'entry_defined',
                'exitDefined': 'exit_defined', 'biasStateJson': 'bias_state_json',
            }
            for key, val in body.items():
                snake = field_map.get(key, key)
                updates[snake] = val

            # Recompute signature if structural fields changed
            struct_fields = {'regime', 'gex_posture', 'vol_state', 'time_structure',
                             'heatmap_color', 'position_structure', 'width_bucket',
                             'directional_bias'}
            if struct_fields & updates.keys():
                existing = self.db.get_edge_lab_setup(setup_id, user_id)
                if existing:
                    for f in struct_fields:
                        if f not in updates:
                            updates[f] = getattr(existing, f)
                    temp = EdgeLabSetup(
                        id='', user_id=0, setup_date='',
                        regime=updates.get('regime', ''),
                        gex_posture=updates.get('gex_posture', ''),
                        vol_state=updates.get('vol_state', ''),
                        time_structure=updates.get('time_structure', ''),
                        heatmap_color=updates.get('heatmap_color', ''),
                        position_structure=updates.get('position_structure', ''),
                        width_bucket=updates.get('width_bucket', ''),
                        directional_bias=updates.get('directional_bias', ''),
                    )
                    updates['structure_signature'] = temp.compute_structure_signature()

            result = self.db.update_edge_lab_setup(setup_id, user_id, updates)
            if not result:
                return self._error_response('Setup not found', 404, request)
            return self._json_response({'success': True, 'data': result.to_api_dict()}, request=request)
        except ValueError as e:
            return self._error_response(str(e), 400, request)
        except Exception as e:
            self.logger.error(f"update_edge_lab_setup error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def get_edge_lab_hypothesis(self, request: web.Request) -> web.Response:
        """GET /api/edge-lab/setups/{setupId}/hypothesis"""
        try:
            user_id = request['user']['id']
            setup_id = request.match_info['setupId']
            hypothesis = self.db.get_hypothesis_for_setup(setup_id, user_id)
            if not hypothesis:
                return self._error_response('Hypothesis not found', 404, request)
            return self._json_response({'success': True, 'data': hypothesis.to_api_dict()}, request=request)
        except Exception as e:
            self.logger.error(f"get_edge_lab_hypothesis error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def create_edge_lab_hypothesis(self, request: web.Request) -> web.Response:
        """POST /api/edge-lab/hypotheses — Create a hypothesis for a setup."""
        try:
            user_id = request['user']['id']
            body = await request.json()

            setup_id = body.get('setup_id') or body.get('setupId')
            if not setup_id:
                return self._error_response('setup_id is required', 400, request)

            # Verify setup exists and belongs to user
            setup = self.db.get_edge_lab_setup(setup_id, user_id)
            if not setup:
                return self._error_response('Setup not found', 404, request)

            # Check no existing hypothesis
            existing = self.db.get_hypothesis_for_setup(setup_id, user_id)
            if existing:
                return self._error_response('Hypothesis already exists for this setup', 409, request)

            for field in ['thesis', 'convexity_source', 'failure_condition']:
                camel = self._snake_to_camel(field)
                if not body.get(field) and not body.get(camel):
                    return self._error_response(f'{field} is required', 400, request)

            hypothesis = EdgeLabHypothesis(
                id=EdgeLabHypothesis.new_id(),
                setup_id=setup_id,
                user_id=user_id,
                thesis=body.get('thesis'),
                convexity_source=body.get('convexity_source') or body.get('convexitySource'),
                failure_condition=body.get('failure_condition') or body.get('failureCondition'),
                max_risk_defined=1 if body.get('max_risk_defined') or body.get('maxRiskDefined') else 0,
            )
            result = self.db.create_edge_lab_hypothesis(hypothesis)
            return self._json_response(
                {'success': True, 'data': result.to_api_dict()},
                request=request, status=201,
            )
        except Exception as e:
            self.logger.error(f"create_edge_lab_hypothesis error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def lock_edge_lab_hypothesis(self, request: web.Request) -> web.Response:
        """POST /api/edge-lab/hypotheses/{id}/lock — Lock a hypothesis."""
        try:
            user_id = request['user']['id']
            hypothesis_id = request.match_info['id']
            result = self.db.lock_hypothesis(hypothesis_id, user_id)
            if not result:
                return self._error_response('Hypothesis not found', 404, request)
            return self._json_response({'success': True, 'data': result.to_api_dict()}, request=request)
        except ValueError as e:
            return self._error_response(str(e), 400, request)
        except Exception as e:
            self.logger.error(f"lock_edge_lab_hypothesis error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def get_edge_lab_outcome(self, request: web.Request) -> web.Response:
        """GET /api/edge-lab/setups/{setupId}/outcome"""
        try:
            user_id = request['user']['id']
            setup_id = request.match_info['setupId']
            outcome = self.db.get_outcome_for_setup(setup_id, user_id)
            if not outcome:
                return self._error_response('Outcome not found', 404, request)
            return self._json_response({'success': True, 'data': outcome.to_api_dict()}, request=request)
        except Exception as e:
            self.logger.error(f"get_edge_lab_outcome error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def create_edge_lab_outcome(self, request: web.Request) -> web.Response:
        """POST /api/edge-lab/outcomes — Create an outcome for a setup."""
        try:
            user_id = request['user']['id']
            body = await request.json()

            setup_id = body.get('setup_id') or body.get('setupId')
            if not setup_id:
                return self._error_response('setup_id is required', 400, request)

            setup = self.db.get_edge_lab_setup(setup_id, user_id)
            if not setup:
                return self._error_response('Setup not found', 404, request)

            # Check no existing outcome
            existing = self.db.get_outcome_for_setup(setup_id, user_id)
            if existing:
                return self._error_response('Outcome already exists for this setup', 409, request)

            outcome_type = body.get('outcome_type') or body.get('outcomeType')
            if not outcome_type:
                return self._error_response('outcome_type is required', 400, request)
            if outcome_type not in EdgeLabOutcome.VALID_OUTCOME_TYPES:
                return self._error_response(
                    f'Invalid outcome_type. Must be one of: {", ".join(EdgeLabOutcome.VALID_OUTCOME_TYPES)}',
                    400, request,
                )

            outcome = EdgeLabOutcome(
                id=EdgeLabOutcome.new_id(),
                setup_id=setup_id,
                user_id=user_id,
                outcome_type=outcome_type,
                system_suggestion=body.get('system_suggestion') or body.get('systemSuggestion'),
                suggestion_confidence=body.get('suggestion_confidence') or body.get('suggestionConfidence'),
                suggestion_reasoning=body.get('suggestion_reasoning') or body.get('suggestionReasoning'),
                hypothesis_valid=body.get('hypothesis_valid') if body.get('hypothesis_valid') is not None
                    else body.get('hypothesisValid'),
                structure_resolved=body.get('structure_resolved') if body.get('structure_resolved') is not None
                    else body.get('structureResolved'),
                exit_per_plan=body.get('exit_per_plan') if body.get('exit_per_plan') is not None
                    else body.get('exitPerPlan'),
                notes=body.get('notes'),
                # pnl_result is recorded for reference ONLY — never used in Edge Score
                pnl_result=body.get('pnl_result') or body.get('pnlResult'),
            )
            result = self.db.create_edge_lab_outcome(outcome)
            return self._json_response(
                {'success': True, 'data': result.to_api_dict()},
                request=request, status=201,
            )
        except Exception as e:
            self.logger.error(f"create_edge_lab_outcome error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def confirm_edge_lab_outcome(self, request: web.Request) -> web.Response:
        """POST /api/edge-lab/outcomes/{id}/confirm — Confirm an outcome."""
        try:
            user_id = request['user']['id']
            outcome_id = request.match_info['id']
            result = self.db.confirm_outcome(outcome_id, user_id)
            if not result:
                return self._error_response('Outcome not found', 404, request)
            return self._json_response({'success': True, 'data': result.to_api_dict()}, request=request)
        except ValueError as e:
            return self._error_response(str(e), 400, request)
        except Exception as e:
            self.logger.error(f"confirm_edge_lab_outcome error: {e}")
            return self._error_response(str(e), 500, request)

    @staticmethod
    def _snake_to_camel(name: str) -> str:
        """Convert snake_case to camelCase."""
        parts = name.split('_')
        return parts[0] + ''.join(p.capitalize() for p in parts[1:])

    # ==================== Edge Lab Analytics Handlers ====================

    @require_auth
    async def suggest_edge_lab_outcome(self, request: web.Request) -> web.Response:
        """GET /api/edge-lab/setups/{setupId}/suggest-outcome"""
        try:
            user_id = request['user']['id']
            setup_id = request.match_info['setupId']
            result = self.edge_lab_analytics.suggest_outcome_type(setup_id, user_id)
            return self._json_response({'success': True, 'data': result}, request=request)
        except Exception as e:
            self.logger.error(f"suggest_edge_lab_outcome error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def get_regime_correlation(self, request: web.Request) -> web.Response:
        """GET /api/edge-lab/analytics/regime-correlation?start=&end="""
        try:
            user_id = request['user']['id']
            start = request.query.get('start')
            end = request.query.get('end')
            if not start or not end:
                return self._error_response('start and end query params required', 400, request)
            result = self.edge_lab_analytics.compute_regime_correlation(user_id, start, end)
            return self._json_response({'success': True, 'data': result}, request=request)
        except Exception as e:
            self.logger.error(f"get_regime_correlation error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def get_bias_overlay(self, request: web.Request) -> web.Response:
        """GET /api/edge-lab/analytics/bias-overlay?start=&end="""
        try:
            user_id = request['user']['id']
            start = request.query.get('start')
            end = request.query.get('end')
            if not start or not end:
                return self._error_response('start and end query params required', 400, request)
            result = self.edge_lab_analytics.compute_bias_overlay(user_id, start, end)
            return self._json_response({'success': True, 'data': result}, request=request)
        except Exception as e:
            self.logger.error(f"get_bias_overlay error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def get_edge_score(self, request: web.Request) -> web.Response:
        """GET /api/edge-lab/analytics/edge-score?start=&end=&scope="""
        try:
            user_id = request['user']['id']
            start = request.query.get('start')
            end = request.query.get('end')
            if not start or not end:
                return self._error_response('start and end query params required', 400, request)
            scope = request.query.get('scope', 'all')
            result = self.edge_lab_analytics.compute_edge_score(user_id, start, end, scope)
            return self._json_response({'success': True, 'data': result}, request=request)
        except Exception as e:
            self.logger.error(f"get_edge_score error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def get_edge_score_history(self, request: web.Request) -> web.Response:
        """GET /api/edge-lab/analytics/edge-score/history?days="""
        try:
            user_id = request['user']['id']
            days = int(request.query.get('days', '90'))
            result = self.edge_lab_analytics.get_edge_score_history(user_id, days)
            return self._json_response({'success': True, 'data': result}, request=request)
        except Exception as e:
            self.logger.error(f"get_edge_score_history error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def get_signature_sample_sizes(self, request: web.Request) -> web.Response:
        """GET /api/edge-lab/analytics/signatures"""
        try:
            user_id = request['user']['id']
            result = self.edge_lab_analytics.get_signature_sample_sizes(user_id)
            return self._json_response({'success': True, 'data': result}, request=request)
        except Exception as e:
            self.logger.error(f"get_signature_sample_sizes error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def compute_edge_lab_analytics(self, request: web.Request) -> web.Response:
        """POST /api/edge-lab/analytics/compute — Trigger full analytics computation."""
        try:
            user_id = request['user']['id']
            body = await request.json()
            start = body.get('start')
            end = body.get('end')
            if not start or not end:
                return self._error_response('start and end are required', 400, request)

            results = {
                'regime_correlation': self.edge_lab_analytics.compute_regime_correlation(user_id, start, end),
                'bias_overlay': self.edge_lab_analytics.compute_bias_overlay(user_id, start, end),
                'edge_score': self.edge_lab_analytics.compute_edge_score(user_id, start, end),
                'signatures': self.edge_lab_analytics.get_signature_sample_sizes(user_id),
            }
            return self._json_response({'success': True, 'data': results}, request=request)
        except Exception as e:
            self.logger.error(f"compute_edge_lab_analytics error: {e}")
            return self._error_response(str(e), 500, request)

    @require_auth
    async def get_edge_lab_dashboard(self, request: web.Request) -> web.Response:
        """GET /api/edge-lab/analytics/dashboard — Fast read from precomputed metrics."""
        try:
            user_id = request['user']['id']
            today = datetime.utcnow().strftime('%Y-%m-%d')
            ninety_ago = (datetime.utcnow() - timedelta(days=90)).strftime('%Y-%m-%d')

            # Read precomputed metrics
            edge_score_metric = self.db.get_metric(user_id, 'edge_score', ninety_ago, today)
            regime_metric = self.db.get_metric(user_id, 'regime_correlation', ninety_ago, today)
            bias_metric = self.db.get_metric(user_id, 'bias_overlay', ninety_ago, today)

            result = {
                'edge_score': edge_score_metric.to_api_dict() if edge_score_metric else None,
                'regime_correlation': regime_metric.to_api_dict() if regime_metric else None,
                'bias_overlay': bias_metric.to_api_dict() if bias_metric else None,
                'score_history': self.edge_lab_analytics.get_edge_score_history(user_id, 90),
                'signatures': self.edge_lab_analytics.get_signature_sample_sizes(user_id),
            }
            return self._json_response({'success': True, 'data': result}, request=request)
        except Exception as e:
            self.logger.error(f"get_edge_lab_dashboard error: {e}")
            return self._error_response(str(e), 500, request)

    async def materialize_edge_lab_metrics(self, request: web.Request) -> web.Response:
        """POST /api/internal/edge-lab/materialize — Manual trigger for metrics materialization."""
        try:
            count = await self._materialize_edge_lab_for_all_users()
            return self._json_response({'success': True, 'users_processed': count}, request=request)
        except Exception as e:
            self.logger.error(f"materialize_edge_lab_metrics error: {e}")
            return self._error_response(str(e), 500, request)

    async def _materialize_edge_lab_for_all_users(self) -> int:
        """Materialize Edge Lab metrics for all users with sufficient data."""
        import time as _time
        from .models_v2 import EdgeLabMetric

        conn = self.db._get_conn()
        cursor = conn.cursor()
        try:
            # Find users with edge lab data
            cursor.execute("""
                SELECT DISTINCT user_id, COUNT(*) as outcome_count
                FROM edge_lab_outcomes
                WHERE is_confirmed = 1
                GROUP BY user_id
            """)
            users = cursor.fetchall()
        finally:
            cursor.close()
            conn.close()

        today = datetime.utcnow().strftime('%Y-%m-%d')
        ninety_ago = (datetime.utcnow() - timedelta(days=90)).strftime('%Y-%m-%d')
        processed = 0

        for row in users:
            user_id = row[0]
            outcome_count = row[1]

            # Skip users with < 10 confirmed outcomes — no point materializing noise
            if outcome_count < 10:
                continue

            start_time = _time.time()

            try:
                # Compute and save edge score
                score_result = self.edge_lab_analytics.compute_edge_score(
                    user_id, ninety_ago, today
                )
                score_metric = EdgeLabMetric(
                    user_id=user_id,
                    metric_type='edge_score',
                    scope='all',
                    window_start=ninety_ago,
                    window_end=today,
                    payload=json.dumps(score_result),
                    sample_size=outcome_count,
                )
                self.db.save_metric(score_metric)

                # Compute and save regime correlation
                regime_result = self.edge_lab_analytics.compute_regime_correlation(
                    user_id, ninety_ago, today
                )
                regime_metric = EdgeLabMetric(
                    user_id=user_id,
                    metric_type='regime_correlation',
                    scope='all',
                    window_start=ninety_ago,
                    window_end=today,
                    payload=json.dumps(regime_result, default=str),
                    sample_size=outcome_count,
                )
                self.db.save_metric(regime_metric)

                # Compute and save bias overlay
                bias_result = self.edge_lab_analytics.compute_bias_overlay(
                    user_id, ninety_ago, today
                )
                bias_metric = EdgeLabMetric(
                    user_id=user_id,
                    metric_type='bias_overlay',
                    scope='all',
                    window_start=ninety_ago,
                    window_end=today,
                    payload=json.dumps(bias_result, default=str),
                    sample_size=outcome_count,
                )
                self.db.save_metric(bias_metric)

                elapsed_ms = int((_time.time() - start_time) * 1000)
                self.logger.info(
                    f"Edge Lab materialized for user {user_id}: "
                    f"{outcome_count} outcomes, {elapsed_ms}ms"
                )
                processed += 1

            except Exception as e:
                self.logger.error(f"Edge Lab materialization failed for user {user_id}: {e}")

        return processed

    async def _edge_lab_scheduler(self):
        """Background task: materialize Edge Lab metrics daily for all users."""
        while True:
            try:
                await asyncio.sleep(86400)  # 24 hours
                self.logger.info("Edge Lab scheduler: starting daily materialization")
                count = await self._materialize_edge_lab_for_all_users()
                self.logger.info(f"Edge Lab scheduler: materialized {count} users")
            except asyncio.CancelledError:
                break
            except Exception as e:
                self.logger.error(f"Edge Lab scheduler error: {e}")
                await asyncio.sleep(3600)  # retry in 1 hour on error

    async def start(self) -> web.AppRunner:
        """Start the API server and return the runner for cleanup."""
        app = self.create_app()
        runner = web.AppRunner(app)
        await runner.setup()

        site = web.TCPSite(runner, '0.0.0.0', self.port)
        await site.start()

        self.logger.ok(f"FOTW Trade Log API (v2) running on port {self.port}", emoji="📒")
        return runner


async def run(config: Dict[str, Any], logger) -> None:
    """Entry point for orchestrator."""
    orchestrator = JournalOrchestrator(config, logger)
    runner = await orchestrator.start()

    # Start leaderboard scheduler as background task
    leaderboard_task = asyncio.create_task(
        orchestrator._leaderboard_scheduler(),
        name="leaderboard-scheduler",
    )

    # Start Edge Lab metrics scheduler as background task
    edge_lab_task = asyncio.create_task(
        orchestrator._edge_lab_scheduler(),
        name="edge-lab-scheduler",
    )

    try:
        # Run forever until cancelled
        while True:
            await asyncio.sleep(1)
    except asyncio.CancelledError:
        logger.info("Orchestrator cancelled", emoji="🛑")
        leaderboard_task.cancel()
        edge_lab_task.cancel()
    finally:
        logger.info("Shutting down API server", emoji="🛑")
        await runner.cleanup()
